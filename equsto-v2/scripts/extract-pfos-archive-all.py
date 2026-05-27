#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
arşiv/projeler altındaki tüm .xlsx / .xls / .pdf → pfos-archive-extract.json

Kullanım:
  python scripts/extract-pfos-archive-all.py
  python scripts/extract-pfos-archive-all.py --root "C:/D Disk/EQUSTO-CURSOR/arşiv/projeler"
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

# ── zone eşleme (extract-pfos-referans-projeler ile uyumlu) ─────────────────
ZONE_LABEL_MAP = [
    (r"^a\s*-\s*kuru|a-\s*kuru\s*depo", "kuru_depo"),
    (r"^f\s*-\s*derin|f-\s*derin\s*dondur", "derin_dondurucu"),
    (r"^c\s*-\s*soguk|^c\s*-\s*soğuk|panel\s*tip\s*soguk", "soguk_oda"),
    (r"^z\s*-\s*cop\s*soguk", "soguk_oda"),
    (r"^e\s*-\s*hazirlik|^e\s*-\s*hazırlık", "sebze_hazirlik"),
    (r"^s\s*-\s*servis|servis\s*teshir", "pastane"),
    (r"^m\s*-\s*servis\s*mutf", "ana_mutfak"),
    (r"^m\s*-\s*sicak|sicak\s*mutfak", "ana_mutfak"),
    (r"^m\s*-\s*acik|acik\s*mutfak", "show_mutfagi"),
    (r"^d\s*-\s*bulasik|^d\s*-\s*bulaşik|personel\s*bulasik", "bulasikhane"),
    (r"^b\s*-\s*bar|tatli\s*bar", "bar"),
    (r"^c\s*-\s*mutfak", "ana_mutfak"),
    (r"lounge\s*mutfak", "bar"),
    (r"banket\s*hatti|banket\s*arabasi", "acik_bufe"),
    (r"ana\s*mutfak|main\s*kitchen|hot\s*kitchen|sicak\s*mutfak", "ana_mutfak"),
    (r"sebze\s*hazir|sebze\s*hazırl|vegetable", "sebze_hazirlik"),
    (r"et\s*hazir|et\s*hazırl|kasap|butcher", "et_hazirlik"),
    (r"hazirlik\s*mutf|hazırlık\s*mutf", "sebze_hazirlik"),
    (r"kuru\s*depo|kuru\s*gida|dry\s*storage", "kuru_depo"),
    (r"so[gğ]uk\s*mutfak|cold\s*kitchen|so[gğ]uk\s*oda|so[gğ]uk\s*depo", "soguk_oda"),
    (r"derin\s*dondur|deep\s*freez|dondurucu\s*depo|dondurucu\s*oda", "derin_dondurucu"),
    (r"bula[sş]ik|yikama\s*mutf|dishwash", "bulasikhane"),
    (r"pastane|patisserie|pastry|unlu\s*mam|tatli\s*uretim", "pastane"),
    (r"\bbar\b|beverage|kahve\s*hatti", "bar"),
    (r"a[cç]ik\s*b[uü]fe|buffet|brunch|kahvalti", "acik_bufe"),
    (r"show\s*mutf|te[sş]hir\s*mutfak", "show_mutfagi"),
    (r"izgara|grill\s*line|ocakba[sş]i|kebab|komurlu", "izgara_meze"),
    (r"personel\s*yemek", "bulasikhane"),
]

PROJECT_META = {
    "2017-050": {"baslik": "DoubleTree Hilton Topkapı", "konsept": "Hotel", "dukkan": "5 Yıldız Otel"},
    "2017-044": {
        "baslik": "THC Bakü",
        "konsept": "Restaurant",
        "dukkan": "All Dining Cafe (TheHouse Cafe, Happymoons vb)",
        "dukkan_kisa": "All Day Cafe",
    },
    "2017-120": {"baslik": "Sütiş Mersin", "konsept": "Restaurant", "dukkan": "Türk Restoran"},
    "2017-204": {"baslik": "Vadistanbul", "konsept": "Restaurant", "dukkan": "Food Court / Çoklu outlet"},
}

POZ_RE = re.compile(r"^[A-Z]\d+$|^[A-Z]{1,3}\d+$", re.I)
SKIP = re.compile(
    r"^(poz|marka|ürün|urun|ölçü|olcu|adet|toplam|proje|teklif|tarih|"
    r"yetkili|telefon|faks|e-posta|iskonto|kdv|eur|try|usd|fiyat|"
    r"genel toplam|ara toplam|not:|sayfa)",
    re.I,
)


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    tr = str.maketrans("ığüşöçâîû", "igusocaiu")
    s = s.translate(tr)
    return re.sub(r"\s+", " ", s)


def map_zone(text: str) -> str | None:
    n = norm(text)
    if len(n) < 3 or len(n) > 120:
        return None
    if not re.search(
        r"mutfak|depo|hazir|bula[sş]|bar|pastane|bufe|izgara|banket|"
        r"soguk|sicak|lounge|yikama|dondurucu|büfe",
        n,
    ):
        return None
    for pat, key in ZONE_LABEL_MAP:
        if re.search(pat, n, re.I):
            return key
    return None


def project_id_from_folder(name: str) -> str:
    m = re.match(r"^(\d{4}-\d+)", name.strip())
    return m.group(1) if m else name.strip()[:32]


def cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def is_section_row(poz: str, name: str, olcu: str, adet) -> bool:
    if not name or len(name) < 4:
        return False
    if SKIP.search(norm(name)):
        return False
    if poz and POZ_RE.match(poz.replace(" ", "")):
        return False
    if map_zone(name):
        return True
    # Başlık satırı: poz yok, ölçü/adet boş, BÜYÜK HARF ağırlıklı
    if not poz and not olcu and adet in (None, "", 0):
        upper = sum(1 for c in name if c.isupper())
        if upper >= max(3, len(name) * 0.35) and map_zone(name):
            return True
        if upper >= max(3, len(name) * 0.35) and re.search(
            r"mutfak|depo|hazirlik|bula[sş]|bar|izgara|so[gğ]uk|sicak|bufe|banket",
            norm(name),
        ):
            return True
    return False


def extract_xlsx_openpyxl(path: Path) -> dict:
    import openpyxl

    zones: dict[str, dict] = {}
    current_zone: str | None = None
    meta = {"proje_adi": "", "teklif_no": ""}
    sheets = []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"error": str(e), "zones": {}, "zone_order": []}

    for sname in wb.sheetnames:
        ws = wb[sname]
        sheet_lines = 0
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cells = list(row) + [None] * 8
            b, c, d, e, f = (
                cell_str(cells[1]),
                cell_str(cells[2]),
                cell_str(cells[3]),
                cell_str(cells[4]),
                cells[5],
            )
            if norm(b) in ("proje", "proje adi") and (d or c):
                meta["proje_adi"] = d or c
            if "teklif" in norm(b) and d:
                meta["teklif_no"] = d
            # Standart Equsto: ürün adı D; bazı eski formlar (Sütiş vb.): B bölüm, C ürün
            if d:
                name = d
                poz = b or c
                olcu = e
                adet = f
            elif c and len(c) > 3:
                name = c
                poz = b
                olcu = cells[4] if len(cells) > 4 else e
                adet = cells[5] if len(cells) > 5 else f
            else:
                name = b
                poz = ""
                olcu = e
                adet = f
            if not name:
                continue
            zk = map_zone(name)
            if is_section_row(poz, name, olcu, adet) or (zk and not poz):
                if not zk:
                    continue
                current_zone = zk
                if zk not in zones:
                    zones[zk] = {
                        "zone_key": zk,
                        "labels_found": [],
                        "line_count": 0,
                        "sample_lines": [],
                    }
                if name not in zones[zk]["labels_found"]:
                    zones[zk]["labels_found"].append(name)
                continue
            if not current_zone:
                continue
            if poz or len(name) > 8:
                zones[current_zone]["line_count"] += 1
                sheet_lines += 1
                if len(zones[current_zone]["sample_lines"]) < 12:
                    zones[current_zone]["sample_lines"].append(name[:200])
        sheets.append({"name": sname, "line_count": sheet_lines})
    wb.close()
    return {
        "meta": meta,
        "zones": zones,
        "zone_order": list(zones.keys()),
        "sheets": sheets,
    }


def extract_xlsx_legacy_zip(path: Path) -> dict:
    """openpyxl başarısız .xls veya bozuk dosyalar için zip xml."""
    from extract_pfos_referans_projeler import extract_xlsx  # type: ignore

    return extract_xlsx(path)


def extract_pdf_pypdf(path: Path, max_pages: int = 30) -> dict:
    try:
        from pypdf import PdfReader
    except ImportError:
        return {"error": "pypdf not installed", "zones": {}, "zone_order": []}

    zones: dict[str, dict] = {}
    current: str | None = None
    try:
        reader = PdfReader(str(path), strict=False)
        text_parts = []
        for page in reader.pages[:max_pages]:
            try:
                t = page.extract_text() or ""
                text_parts.append(t)
            except Exception:
                continue
        full = "\n".join(text_parts)
        page_count = len(reader.pages)
    except Exception as e:
        return {"error": str(e), "zones": {}, "zone_order": []}

    for line in full.splitlines():
        line = line.strip()
        if len(line) < 4:
            continue
        zk = map_zone(line)
        if zk:
            current = zk
            if zk not in zones:
                zones[zk] = {
                    "zone_key": zk,
                    "labels_found": [line[:120]],
                    "line_count": 0,
                    "sample_lines": [],
                }
            elif line not in zones[zk]["labels_found"]:
                zones[zk]["labels_found"].append(line[:120])
            continue
        if current and len(line) > 10 and not SKIP.search(norm(line)):
            zones[current]["line_count"] += 1
            if len(zones[current]["sample_lines"]) < 8:
                zones[current]["sample_lines"].append(line[:200])

    return {
        "pages": min(page_count, max_pages),
        "zones": zones,
        "zone_order": list(zones.keys()),
        "text_chars": len(full),
    }


def merge_zones(into: dict, add: dict) -> None:
    for zk, z in add.items():
        if not zk:
            continue
        if zk not in into:
            into[zk] = {
                "zone_key": zk,
                "labels_found": [],
                "line_count": 0,
                "sample_lines": [],
                "sources": [],
            }
        t = into[zk]
        for lb in z.get("labels_found") or []:
            if lb not in t["labels_found"]:
                t["labels_found"].append(lb)
        t["line_count"] += z.get("line_count") or 0
        for s in z.get("sample_lines") or []:
            if len(t["sample_lines"]) < 15 and s not in t["sample_lines"]:
                t["sample_lines"].append(s)


def scan_archive(root: Path) -> dict:
    projects: dict[str, dict] = {}
    stats = {"xlsx_ok": 0, "xlsx_err": 0, "pdf_ok": 0, "pdf_err": 0, "skipped": 0}

    for fpath in sorted(root.rglob("*")):
        if not fpath.is_file():
            continue
        suf = fpath.suffix.lower()
        if suf not in (".xlsx", ".xls", ".pdf"):
            continue
        if fpath.name.startswith("~$"):
            stats["skipped"] += 1
            continue

        rel = fpath.relative_to(root)
        folder = rel.parts[0] if len(rel.parts) > 1 else "_root"
        pid = project_id_from_folder(folder)

        if pid not in projects:
            base = PROJECT_META.get(pid, {})
            projects[pid] = {
                "id": pid,
                "folder": folder,
                "baslik": base.get("baslik") or folder,
                "konsept": base.get("konsept", ""),
                "dukkan": base.get("dukkan", ""),
                "files": [],
                "zones": {},
                "zone_order": [],
            }
        proj = projects[pid]
        entry = {"path": str(rel).replace("\\", "/"), "type": suf[1:], "name": fpath.name}

        if suf in (".xlsx", ".xls"):
            if suf == ".xls":
                entry["status"] = "skipped_xls"
                stats["skipped"] += 1
                proj["files"].append(entry)
                continue
            ex = extract_xlsx_openpyxl(fpath)
            if ex.get("error"):
                entry["status"] = "error"
                entry["error"] = ex["error"]
                stats["xlsx_err"] += 1
            else:
                entry["status"] = "ok"
                entry["zone_order"] = ex.get("zone_order") or []
                entry["line_count"] = sum(
                    (ex.get("zones") or {}).get(z, {}).get("line_count", 0)
                    for z in entry["zone_order"]
                )
                merge_zones(proj["zones"], ex.get("zones") or {})
                if ex.get("meta", {}).get("proje_adi"):
                    proj["proje_adi"] = ex["meta"]["proje_adi"]
                stats["xlsx_ok"] += 1
        else:
            ex = extract_pdf_pypdf(fpath)
            if ex.get("error"):
                entry["status"] = "error"
                entry["error"] = ex["error"]
                stats["pdf_err"] += 1
            else:
                entry["status"] = "ok"
                entry["zone_order"] = ex.get("zone_order") or []
                merge_zones(proj["zones"], ex.get("zones") or {})
                stats["pdf_ok"] += 1

        proj["files"].append(entry)

    # zone_order + meta
    out_projects = []
    zone_freq: dict[str, int] = defaultdict(int)
    for pid in sorted(projects.keys()):
        p = projects[pid]
        meta = PROJECT_META.get(pid, {})
        for k, v in meta.items():
            if v:
                p[k] = v
        p["zones"] = {k: v for k, v in p["zones"].items() if k}
        p["zone_order"] = list(p["zones"].keys())
        p["zone_count"] = len(p["zone_order"])
        p["file_count"] = len(p["files"])
        p["status"] = "ok" if p["zones"] else "files_only"
        for zk in p["zone_order"]:
            if zk:
                zone_freq[zk] += 1
        out_projects.append(p)

    return {
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "root": str(root),
        "stats": stats,
        "project_count": len(out_projects),
        "projects": out_projects,
        "aggregate": {
            "zone_frequency": dict(
                sorted((k, v) for k, v in zone_freq.items() if k)
            )
        },
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--root",
        default=str(Path(__file__).resolve().parents[2] / "arşiv" / "projeler")
        if (Path(__file__).resolve().parents[2] / "arşiv" / "projeler").is_dir()
        else r"C:/D Disk/EQUSTO-CURSOR/arşiv/projeler",
    )
    ap.add_argument(
        "--out",
        default=str(
            Path(__file__).resolve().parents[1] / "public" / "data" / "pfos-archive-extract.json"
        ),
    )
    ap.add_argument(
        "--referans-out",
        default=str(
            Path(__file__).resolve().parents[1] / "public" / "data" / "pfos-referans-projeler.json"
        ),
    )
    args = ap.parse_args()
    root = Path(args.root)
    if not root.is_dir():
        # fallback
        root = Path(r"C:/D Disk/EQUSTO-CURSOR/arşiv/projeler")
    print(f"Scanning {root} ...")
    data = scan_archive(root)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Wrote {out} — {data['project_count']} projects, "
        f"xlsx ok={data['stats']['xlsx_ok']} pdf ok={data['stats']['pdf_ok']}"
    )

    # 4 referans projeyi güncelle
    ref_ids = set(PROJECT_META.keys())
    referans = {
        "version": 2,
        "generated_at": data["generated_at"],
        "source": "extract-pfos-archive-all.py",
        "note": "Öncelikli 4 referans + tüm arşiv: pfos-archive-extract.json",
        "projects": [p for p in data["projects"] if p["id"] in ref_ids],
        "aggregate": data["aggregate"],
    }
    for p in referans["projects"]:
        p["status"] = "ok" if p.get("zones") else "partial"
    Path(args.referans_out).write_text(
        json.dumps(referans, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Updated {args.referans_out} ({len(referans['projects'])} referans)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
