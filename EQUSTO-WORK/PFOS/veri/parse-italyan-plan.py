# -*- coding: utf-8 -*-
"""ITALYAN-PLAN.pdf → JSON + Excel (toplam m² kullanıcı notu: 180)."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERI = Path(__file__).resolve().parent
PROJE_VERI = VERI / "proje-veri"
PLAN_PDF = PROJE_VERI / "ITALYAN-PLAN.pdf"
EKIPMAN_PDF = PROJE_VERI / "03-italyan.pdf"

OUT_JSON = VERI / "ITALYAN-PLAN.json"
OUT_ALAN_XLSX = VERI / "ITALYAN-PLAN-alan-listesi.xlsx"
OUT_BIRLESIK_XLSX = VERI / "ITALYAN-PLAN-birlesik.xlsx"

TOTAL_M2_USER = 180.0
PLAN_PDF_HEADER_M2 = 57.7

# Plan üzerindeki m² etiketleri (toplam 83) → ekipman bölümleri
MAHALS = [
    {
        "id": "kuru-depo",
        "label": "Kuru depo",
        "bolum": "A",
        "planM2": 6,
        "pozlar": ["A1", "A2"],
    },
    {
        "id": "soguk-oda",
        "label": "Soğuk oda",
        "bolum": "B",
        "planM2": 6,
        "pozlar": ["B1", "B2", "B3"],
    },
    {
        "id": "derin-dondurucu",
        "label": "Derin dondurucu",
        "bolum": "C",
        "planM2": 7,
        "pozlar": ["C1", "C2"],
    },
    {
        "id": "sebze-hazirlik",
        "label": "Sebze hazırlık",
        "bolum": "D",
        "planM2": 7,
        "pozlar": ["D1", "D2", "D3", "D4", "D5", "D6"],
    },
    {
        "id": "et-hazirlik",
        "label": "Et hazırlık",
        "bolum": "E",
        "planM2": 14,
        "pozlar": ["E1", "E2", "E2A", "E3", "E4", "E5", "E6", "E7"],
    },
    {
        "id": "hamur-hazirlik",
        "label": "Hamur hazırlık",
        "bolum": "F",
        "planM2": 10,
        "pozlar": ["F1", "F2", "F3", "F4", "F5", "F6", "F7", "F8", "F9"],
    },
    {
        "id": "pisirme",
        "label": "Pişirme",
        "bolum": "G",
        "planM2": 12,
        "pozlar": [
            "G1", "G2", "G3", "G4", "G5", "G6", "G7", "G8", "G9", "G10",
            "G11", "G12", "G12A", "G13", "G14", "G15", "G16",
            "Y1", "Y2",
        ],
    },
    {
        "id": "soguk-hazirlik",
        "label": "Soğuk hazırlık",
        "bolum": "H",
        "planM2": 6,
        "pozlar": ["H1", "H2", "H3", "H3A", "H4", "H5", "H6", "H7", "H8", "H9"],
    },
    {
        "id": "on-mutfak-bar",
        "label": "Ön mutfak / Bar",
        "bolum": "J",
        "planM2": 15,
        "pozlar": [
            "J1", "J2", "J3", "J4", "J5", "J6", "J7", "J8", "J9", "J10",
            "J11", "J12", "J13", "J14", "J15", "J16", "J17",
        ],
    },
    {
        "id": "bulasik-yikama",
        "label": "Bulaşık yıkama",
        "bolum": "K",
        "planM2": None,
        "planM2Not": "Plan üzerinde ayrı m² etiketi yok; mahal etiketi mevcut",
        "pozlar": ["K1", "K2", "K3", "K4", "K5", "K6", "K7", "K8", "K9", "K10", "K11"],
    },
]

POZ_RE = re.compile(r"^[A-Z]\d{1,2}A?$|^Y\d$")


def read_pdf_text(path: Path) -> str:
    try:
        import fitz  # pymupdf
    except ImportError:
        return path.read_text(encoding="utf-8", errors="ignore")
    doc = fitz.open(path)
    parts = [page.get_text() for page in doc]
    doc.close()
    return "\n".join(parts)


def extract_plan_poz(text: str) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for raw in text.split():
        tok = raw.strip().replace("\t", "")
        if not tok or tok in ("**", "m2", "03"):
            continue
        if POZ_RE.match(tok) and tok not in seen:
            seen.add(tok)
            out.append(tok)
    def sort_key(p: str):
        m = re.search(r"\d+", p)
        return (p[0], int(m.group()) if m else 0, p)

    return sorted(out, key=sort_key)


def parse_ekipman(pdf_text: str) -> dict[str, dict]:
    """03-italyan.pdf → poz → {ad, olcu, adet, bolum} (PyMuPDF: alanlar satır satır)."""
    items: dict[str, dict] = {}
    bolum = ""
    lines = [ln.strip() for ln in pdf_text.splitlines() if ln.strip()]
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("P.NO") or line.startswith("03-") or line.startswith("--"):
            i += 1
            continue
        if re.match(r"^[A-Z]- ", line):
            bolum = line.split("-", 1)[0].strip()
            i += 1
            continue
        if re.match(r"^[A-Z]\d{1,2}A?$", line) or re.match(r"^Y\d$", line):
            poz = line
            i += 1
            ad_lines: list[str] = []
            while i < len(lines) and not re.match(r"^[A-Z]\d{1,2}A?$", lines[i]) and not re.match(
                r"^Y\d$", lines[i]
            ) and not re.match(r"^[A-Z]- ", lines[i]):
                if lines[i] in ("P.NO", "ÜRÜN ADI", "ÖLÇÜ", "AD.") or lines[i].startswith("03-"):
                    i += 1
                    continue
                ad_lines.append(lines[i])
                i += 1
            if len(ad_lines) >= 2:
                olcu = ad_lines[-2]
                adet_s = ad_lines[-1]
                ad = " ".join(ad_lines[:-2]) if len(ad_lines) > 2 else ad_lines[0]
                if adet_s.isdigit():
                    items[poz] = {
                        "bolum": bolum,
                        "ad": ad,
                        "olcu": olcu or "—",
                        "adet": int(adet_s),
                    }
            continue
        i += 1
    return items


def scale_m2(plan_m2: float | None, factor: float) -> float | None:
    if plan_m2 is None:
        return None
    return round(plan_m2 * factor, 1)


def build_json(plan_text: str, ekipman: dict[str, dict]) -> dict:
    annotated = [m["planM2"] for m in MAHALS if m["planM2"] is not None]
    plan_sum = sum(annotated)
    factor = TOTAL_M2_USER / plan_sum if plan_sum else 1.0

    plan_poz = extract_plan_poz(plan_text)
    mahaller = []
    for m in MAHALS:
        scaled = scale_m2(m["planM2"], factor)
        poz_detay = []
        for poz in m["pozlar"]:
            eq = ekipman.get(poz)
            poz_detay.append({
                "poz": poz,
                "planda": poz in plan_poz,
                "ekipman": eq,
            })
        mahaller.append({
            **m,
            "scaledM2": scaled,
            "pozDetay": poz_detay,
        })

    only_plan = sorted(set(plan_poz) - set(ekipman))
    only_ekipman = sorted(set(ekipman) - set(plan_poz))

    return {
        "version": "1.0",
        "konsept": "03-italyan",
        "label": "İTALYAN",
        "kaynak": {
            "plan": str(PLAN_PDF.name),
            "ekipman": str(EKIPMAN_PDF.name),
        },
        "meta": {
            "totalM2": TOTAL_M2_USER,
            "totalM2Kaynak": "kullanici-notu",
            "planPdfHeaderM2": PLAN_PDF_HEADER_M2,
            "planAnnotatedSumM2": plan_sum,
            "scaleFactor": round(factor, 4),
        },
        "mahaller": [
            {k: v for k, v in m.items() if k != "pozDetay"}
            | {"scaledM2": m["scaledM2"], "pozSayisi": len(m["pozlar"])}
            for m in mahaller
        ],
        "mahallerDetay": mahaller,
        "planPozlari": plan_poz,
        "uyumsuzluk": {
            "plandaEkipmandaYok": only_plan,
            "ekipmandaPlandaYok": only_ekipman,
        },
    }


def style_hdr(cell, text: str):
    cell.value = text
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="001E50")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_alan_xlsx(data: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alan"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "03 — İTALYAN · Plan alan listesi"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="001E50")
    ws.merge_cells("A1:F1")
    ws["A2"] = "Toplam m² (proje)"
    ws["B2"] = data["meta"]["totalM2"]
    ws["A3"] = "Plan PDF başlık m²"
    ws["B3"] = data["meta"]["planPdfHeaderM2"]
    ws["A4"] = "Plan etiket toplamı"
    ws["B4"] = data["meta"]["planAnnotatedSumM2"]

    row = 6
    headers = ["Sıra", "Mahal", "Böl.", "Plan m²", "Ölçekli m² (180)", "Pozlar"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col)
        style_hdr(c, h)
        c.border = border

    row += 1
    for i, m in enumerate(data["mahaller"], 1):
        det = next(x for x in data["mahallerDetay"] if x["id"] == m["id"])
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=m["label"]).border = border
        ws.cell(row=row, column=3, value=m["bolum"]).border = border
        ws.cell(row=row, column=4, value=m.get("planM2") or "—").border = border
        ws.cell(row=row, column=5, value=m.get("scaledM2") or "—").border = border
        ws.cell(row=row, column=6, value=", ".join(det["pozlar"])).border = border
        row += 1

    widths = [6, 22, 6, 10, 14, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"
    wb.save(OUT_ALAN_XLSX)


def write_birlesik_xlsx(data: dict, ekipman: dict[str, dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan + Ekipman"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sec_fill = PatternFill("solid", fgColor="E8EEF5")

    ws["A1"] = "03 — İTALYAN · Plan poz ↔ ekipman"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="001E50")
    ws.merge_cells("A1:G1")

    row = 3
    headers_list = ["Mahal", "Böl.", "Poz", "Planda", "Ürün adı", "Ölçü", "Ad."]
    for col, h in enumerate(headers_list, 1):
        c = ws.cell(row=row, column=col)
        style_hdr(c, h)
        c.border = border
    row += 1

    for m in data["mahallerDetay"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"{m['bolum']} — {m['label']}")
        c.font = Font(bold=True, color="001E50")
        c.fill = sec_fill
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1
        for pd in m["pozDetay"]:
            poz = pd["poz"]
            eq = pd.get("ekipman") or {}
            ws.cell(row=row, column=1, value=m["label"]).border = border
            ws.cell(row=row, column=2, value=m["bolum"]).border = border
            ws.cell(row=row, column=3, value=poz).border = border
            ws.cell(row=row, column=4, value="✓" if pd["planda"] else "—").border = border
            ws.cell(row=row, column=5, value=eq.get("ad", "—")).border = border
            ws.cell(row=row, column=6, value=eq.get("olcu", "—")).border = border
            ws.cell(row=row, column=7, value=eq.get("adet", "—")).border = border
            row += 1

    for i, w in enumerate([20, 6, 8, 8, 52, 18, 6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    wb.save(OUT_BIRLESIK_XLSX)


def main():
    plan_text = read_pdf_text(PLAN_PDF)
    ekipman_text = read_pdf_text(EKIPMAN_PDF)
    ekipman = parse_ekipman(ekipman_text)

    data = build_json(plan_text, ekipman)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_alan_xlsx(data)
    write_birlesik_xlsx(data, ekipman)
    print(OUT_JSON)
    print(OUT_ALAN_XLSX)
    print(OUT_BIRLESIK_XLSX)
    print(f"toplam m2: {data['meta']['totalM2']}, mahal: {len(data['mahaller'])}")


if __name__ == "__main__":
    main()
