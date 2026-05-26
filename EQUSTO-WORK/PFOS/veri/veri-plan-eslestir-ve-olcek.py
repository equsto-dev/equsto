# -*- coding: utf-8 -*-
"""
PFOS veri/: plan PDF + ekipman listesi eşleştir → *-PLAN-birlesik.xlsx
+ PFOS-VERI-ESLESTIRME.xlsx + PFOS-OLCEK-MATRISI.xlsx/.json
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERI = Path(__file__).resolve().parent
PROJE_VERI = VERI / "proje-veri"

# Plan + liste çiftleri (veri kökü veya alt klasör)
PAIRS: list[dict] = [
    {"slug": "03-italyan", "label": "İTALYAN", "plan": "ITALYAN-PLAN.pdf", "liste": "03-italyan-ekipman-listesi.xlsx", "projeM2": 180},
    {"slug": "06-sushi", "label": "SUSHI", "plan": "SUSHI.-PLANpdf.pdf", "liste": "06-SUSHI-ekipman-listesi.xlsx"},
    {"slug": "08-hamburger", "label": "HAMBURGER", "plan": "8 HAMBURGER-PLAN.pdf", "liste": "08-HAMBURGER-ekipman-listesi.xlsx"},
    {"slug": "11-birahane", "label": "BİRAHANE", "plan": "11 BIRAHANE-PLAN.pdf", "liste": "11-BIRAHANE-ekipman-listesi.xlsx"},
    {"slug": "13-hotdog", "label": "HOTDOG", "plan": "13 HOTDOG-PLAN.pdf", "liste": "13-HOTDOG-ekipman-listesi.xlsx"},
    {"slug": "14-pastane", "label": "PASTANE", "plan": "14 PASTANE-PLAN.pdf", "liste": "14-PASTANE-ekipman-listesi.xlsx"},
    {"slug": "17-tavukcu", "label": "TAVUKÇU", "plan": "17 TAVUKCU-PLAN.pdf", "liste": "17-TAVUKCU-ekipman-listesi.xlsx"},
    {"slug": "19-thehouse-cafe", "label": "THEHOUSE CAFE", "plan": "19 THEHOUSE CAFE-PLAN.pdf", "liste": "19-THEHOUSE-CAFE-ekipman-listesi.xlsx"},
    {"slug": "20-dondurmaci-krep", "label": "DONDURMACI-KREP", "plan": "20 DONDURMACI - KREP-PLAN.pdf", "liste": "20-DONDURMACI---KREP-ekipman-listesi.xlsx"},
    {"slug": "04-pideci", "label": "PİDECİ", "plan": "PIDECI-PLAN.pdf", "liste": "PIDECI-ekipman-listesi.xlsx"},
    {"slug": "restoran", "label": "RESTORAN", "plan": "RESTORAN-PLAN.pdf", "liste": "RESTORAN-ekipman-listesi.xlsx"},
    # Alt klasör — aynı plan, m² bantına göre ayrı ekipman listesi (matris + birleşik)
    {"slug": "02-balikci-80-150", "label": "BALIKÇI 80-150", "plan": "2 BALIKCI-PLAN.pdf", "liste": "BALIKCI/80-150 m2 BALIKCI-ekipman-listesi.xlsx", "band": "80-150"},
    {"slug": "02-balikci-150-250", "label": "BALIKÇI 150-250", "plan": "2 BALIKCI-PLAN.pdf", "liste": "BALIKCI/150-250 m2 BALIKCI-ekipman-listesi.xlsx", "band": "150-250"},
    {"slug": "01-steakhouse-80-150", "label": "STEAKHOUSE 80-150", "plan": "STEAKHOUSE/STEAKHOUSE-PLAN.pdf", "liste": "STEAKHOUSE/80-150 m2-steakhouse-ekipman-listesi.xlsx", "band": "80-150"},
    {"slug": "01-steakhouse-150-250", "label": "STEAKHOUSE 150-250", "plan": "STEAKHOUSE/STEAKHOUSE-PLAN.pdf", "liste": "STEAKHOUSE/150-250 M2-steakhouse-ekipman-listesi.xlsx", "band": "150-250"},
]

# Liste var, plan yok
LISTE_ONLY = [
    {"slug": "07-sarkuteri", "label": "ŞARKÜTERİ", "liste": "7 ŞARKÜTERİ.pdf"},
]

WIZARD_M2_PRESETS = [20, 45, 80, 120, 150, 180, 250, 350, 500, 750, 1000]

POZ_RE = re.compile(r"^[A-Z]\d{1,2}A?$|^\d{1,3}$")


def read_pdf(path: Path) -> str:
    import fitz

    doc = fitz.open(path)
    t = "\n".join(p.get_text() for p in doc)
    doc.close()
    return t


def extract_plan_m2_values(text: str) -> list[float]:
    vals = []
    for ln in text.splitlines():
        for m in re.finditer(r"(\d+(?:[.,]\d+)?)\s*m\s*[²2]", ln, re.I):
            v = float(m.group(1).replace(",", "."))
            if 3 <= v <= 2000:
                vals.append(v)
    return vals


def extract_plan_poz(text: str) -> list[str]:
    """Plan üzerindeki poz etiketleri (A1, C24, J11A …)."""
    seen: set[str] = set()
    for m in re.finditer(r"\b([A-Z]\d{1,2}A?)\b", text):
        seen.add(m.group(1))
    return sorted(seen, key=lambda p: (p[0], int(re.search(r"\d+", p).group())))


def _is_poz(s: str) -> bool:
    return bool(re.match(r"^[A-Z]\d{1,2}A?$", s) or re.match(r"^\d{1,3}$", s))


def read_ekipman_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    bolum = ""
    bolum_ad = ""
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not any(row):
            continue
        a, b, c, d, e = (row + (None,) * 5)[:5]
        if str(a or "").strip().upper() in ("TOPLAM ADET", "PNO", "P.NO"):
            if str(a or "").strip().upper() == "TOPLAM ADET":
                break
            continue
        if a and not b and isinstance(a, str) and "-" in str(a):
            bolum_ad = str(a).strip()
            bolum = bolum_ad.split("-", 1)[0].strip()
            continue
        poz, ad, olcu, adet_raw = None, None, None, None
        # Düzen A: Böl. | Poz | Ürün | Ölçü | Ad.
        if b and c and _is_poz(str(b).strip()):
            poz = str(b).strip()
            ad, olcu, adet_raw = str(c).strip(), d, e
        # Düzen B (Steakhouse vb.): Poz | Ürün | Ölçü | Adet — poz A sütununda
        elif a and b and _is_poz(str(a).strip()):
            poz = str(a).strip()
            ad, olcu, adet_raw = str(b).strip(), c, d
        if not poz or not ad:
            continue
        adet = adet_raw
        if isinstance(adet, float):
            adet = int(adet)
        rows.append({
            "bolum": bolum,
            "bolumAd": bolum_ad,
            "poz": poz,
            "ad": ad,
            "olcu": str(olcu).strip() if olcu else "—",
            "adet": adet if adet is not None else "—",
        })
    wb.close()
    return rows


def total_adet(rows: list[dict]) -> int:
    t = 0
    for r in rows:
        if isinstance(r.get("adet"), int):
            t += r["adet"]
    return t


def merge_pair(pair: dict) -> dict | None:
    plan_path = PROJE_VERI / pair["plan"]
    liste_path = PROJE_VERI / pair["liste"]
    if not plan_path.exists() or not liste_path.exists():
        return None

    plan_text = read_pdf(plan_path)
    m2_vals = extract_plan_m2_values(plan_text)
    plan_m2_header = m2_vals[0] if m2_vals else None
    plan_m2_sum = sum(m2_vals[:12]) if m2_vals else None  # rough annotated sum cap

    referans_m2 = pair.get("projeM2") or plan_m2_header
    if pair.get("band"):
        parts = pair["band"].split("-")
        if len(parts) == 2:
            try:
                referans_m2 = (int(parts[0]) + int(parts[1])) / 2
            except ValueError:
                pass

    plan_poz = set(extract_plan_poz(plan_text))
    ekipman = read_ekipman_xlsx(liste_path)
    if not ekipman and liste_path.suffix.lower() == ".pdf":
        ekipman = []  # skip pdf parse inline — xlsx preferred

    eq_poz = {r["poz"] for r in ekipman}
    only_plan = sorted(plan_poz - eq_poz)
    only_eq = sorted(eq_poz - plan_poz)
    matched = len(eq_poz & plan_poz)

    slug = pair["slug"]
    safe = slug.upper().replace(" ", "-")
    out_birlesik = VERI / f"{safe}-PLAN-birlesik.xlsx"

    write_birlesik(out_birlesik, pair["label"], ekipman, plan_poz)

    return {
        "slug": slug,
        "label": pair["label"],
        "band": pair.get("band"),
        "planDosya": pair["plan"],
        "listeDosya": pair["liste"],
        "birlesikDosya": out_birlesik.name,
        "planM2Baslik": plan_m2_header,
        "planM2EtiketToplami": round(plan_m2_sum, 1) if plan_m2_sum and plan_m2_sum < 500 else None,
        "referansM2": referans_m2,
        "referansM2Kaynak": "proje-notu" if pair.get("projeM2") else ("m2-bant" if pair.get("band") else "plan-baslik"),
        "ekipmanKalem": len(ekipman),
        "toplamAdet": total_adet(ekipman),
        "planPozSayisi": len(plan_poz),
        "eslesenPoz": matched,
        "plandaListeYok": only_eq,
        "plandaFazla": only_plan,
        "eslesmeOrani": round(matched / len(eq_poz), 3) if eq_poz else 0,
        "durum": "tam" if not only_plan and len(only_eq) <= 3 else "kısmi",
    }


def write_birlesik(path: Path, title: str, ekipman: list[dict], plan_poz: set[str]):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="001E50")
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan + Ekipman"
    ws["A1"] = f"{title} · Plan poz ↔ ekipman"
    ws["A1"].font = Font(size=14, bold=True, color="001E50")
    ws.merge_cells("A1:G1")
    row = 3
    for col, h in enumerate(("Böl.", "Poz", "Planda", "Ürün adı", "Ölçü", "Ad.", "Not"), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.border = border
    row += 1
    for eq in ekipman:
        planda = "✓" if eq["poz"] in plan_poz else "—"
        notu = "" if planda == "✓" else "Listede; planda yok"
        for col, val in enumerate(
            (eq.get("bolum"), eq["poz"], planda, eq["ad"], eq["olcu"], eq["adet"], notu), 1
        ):
            ws.cell(row=row, column=col, value=val).border = border
        row += 1
    for i, w in enumerate((6, 8, 8, 44, 16, 6, 18), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    wb.save(path)


def olcek_satirlari(referans_m2: float | None, presets: list[int]) -> list[dict]:
    if not referans_m2 or referans_m2 <= 0:
        return [{"hedefM2": p, "alanCarpani": None, "adetCarpani": None} for p in presets]
    out = []
    for p in presets:
        c = round(p / referans_m2, 3)
        # PFOS kural özeti: alan doğrusal; adet kademeli (küçük alanlarda yumuşat)
        adet_c = round(c**0.85, 3) if c < 1.5 else round(min(c, 2.5), 3)
        out.append({
            "hedefM2": p,
            "alanCarpani": c,
            "adetCarpaniOneri": adet_c,
            "not": "alan=doğrusal; adet≈referansAdet×adetCarpani (yuvarla); soğuk oda/panel tek kalabilir",
        })
    return out


def write_olcek_matrisi(konseptler: list[dict]):
    data = {
        "version": "1.0",
        "aciklama": "PFOS veri arşivi referans m² → wizard hedef m² ölçek önerisi (taslak, canlı motor değil)",
        "wizardM2": {"min": 20, "max": 1000, "presets": WIZARD_M2_PRESETS},
        "kurallar": {
            "alanDagitimi": "bolumM2_hedef = bolumM2_referans × alanCarpani",
            "adetGenel": "adet_hedef = max(1, round(adet_referans × adetCarpaniOneri))",
            "singleton": "panel soğuk oda, davlumbaz, kasa: referans 1 ise hedef≤1.5× referansta 1 kalır",
            "depoRaf": "istif rafı / raf: tam orantılı",
        },
        "konseptler": [],
    }
    for k in konseptler:
        ref = k.get("referansM2")
        data["konseptler"].append({**k, "olcekBantlari": olcek_satirlari(ref, WIZARD_M2_PRESETS)})

    (VERI / "PFOS-OLCEK-MATRISI.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    wb = Workbook()
    # Sayfa 1 — özet
    ws = wb.active
    ws.title = "Ozet"
    headers = [
        "Slug", "Konsept", "m² bant", "Referans m²", "Kaynak", "Plan m² (başlık)",
        "Kalem", "Toplam adet", "Plan poz", "Eşleşen", "Eşleşme %", "Durum",
        "Plan dosya", "Liste", "Birleşik",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="001E50")
    r = 2
    for k in konseptler:
        ws.cell(row=r, column=1, value=k["slug"])
        ws.cell(row=r, column=2, value=k["label"])
        ws.cell(row=r, column=3, value=k.get("band") or "")
        ws.cell(row=r, column=4, value=k.get("referansM2"))
        ws.cell(row=r, column=5, value=k.get("referansM2Kaynak"))
        ws.cell(row=r, column=6, value=k.get("planM2Baslik"))
        ws.cell(row=r, column=7, value=k.get("ekipmanKalem"))
        ws.cell(row=r, column=8, value=k.get("toplamAdet"))
        ws.cell(row=r, column=9, value=k.get("planPozSayisi"))
        ws.cell(row=r, column=10, value=k.get("eslesenPoz"))
        ws.cell(row=r, column=11, value=k.get("eslesmeOrani"))
        ws.cell(row=r, column=12, value=k.get("durum"))
        ws.cell(row=r, column=13, value=k.get("planDosya"))
        ws.cell(row=r, column=14, value=k.get("listeDosya"))
        ws.cell(row=r, column=15, value=k.get("birlesikDosya"))
        r += 1

    # Sayfa 2 — ölçek matrisi (uzun)
    ws2 = wb.create_sheet("OlcekMatrisi")
    ws2.cell(row=1, column=1, value="Konsept")
    ws2.cell(row=1, column=2, value="Referans m²")
    for i, p in enumerate(WIZARD_M2_PRESETS, 3):
        ws2.cell(row=1, column=i, value=f"{p} m²")
    r = 2
    for k in konseptler:
        ref = k.get("referansM2")
        ws2.cell(row=r, column=1, value=k["label"])
        ws2.cell(row=r, column=2, value=ref)
        for i, p in enumerate(WIZARD_M2_PRESETS, 3):
            if ref:
                ws2.cell(row=r, column=i, value=round(p / ref, 3))
            else:
                ws2.cell(row=r, column=i, value="—")
        r += 1

    wb.save(VERI / "PFOS-OLCEK-MATRISI.xlsx")


def write_eslestirme_ozet(rows: list[dict], liste_only: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Eslestirme"
    hdr = [
        "Tip", "Slug", "Label", "Plan", "Liste", "Birleşik", "Referans m²",
        "Kalem", "Eşleşme", "Durum", "Not",
    ]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="001E50")
    r = 2
    for k in rows:
        ws.cell(row=r, column=1, value="plan+liste")
        ws.cell(row=r, column=2, value=k["slug"])
        ws.cell(row=r, column=3, value=k["label"])
        ws.cell(row=r, column=4, value=k.get("planDosya"))
        ws.cell(row=r, column=5, value=k.get("listeDosya"))
        ws.cell(row=r, column=6, value=k.get("birlesikDosya"))
        ws.cell(row=r, column=7, value=k.get("referansM2"))
        ws.cell(row=r, column=8, value=k.get("ekipmanKalem"))
        ws.cell(row=r, column=9, value=k.get("eslesmeOrani"))
        ws.cell(row=r, column=10, value=k.get("durum"))
        notlar = []
        if k.get("plandaListeYok"):
            notlar.append(f"planda yok: {len(k['plandaListeYok'])} poz")
        if k.get("plandaFazla"):
            notlar.append(f"planda fazla: {k['plandaFazla'][:5]}")
        ws.cell(row=r, column=11, value="; ".join(notlar)[:200])
        r += 1
    for lo in liste_only:
        ws.cell(row=r, column=1, value="liste-only")
        ws.cell(row=r, column=2, value=lo["slug"])
        ws.cell(row=r, column=3, value=lo["label"])
        ws.cell(row=r, column=5, value=lo["liste"])
        ws.cell(row=r, column=11, value="Plan PDF yok")
        r += 1
    wb.save(VERI / "PFOS-VERI-ESLESTIRME.xlsx")


def main():
    results: list[dict] = []
    for pair in PAIRS:
        row = merge_pair(pair)
        if row:
            results.append(row)
            print(f"OK {row['birlesikDosya']} | {row['label']} | ref {row['referansM2']} m2 | eslesme {row['eslesmeOrani']}")
        else:
            print(f"SKIP {pair['slug']}")

    write_eslestirme_ozet(results, LISTE_ONLY)
    write_olcek_matrisi(results)
    print("\nPFOS-VERI-ESLESTIRME.xlsx")
    print("PFOS-OLCEK-MATRISI.xlsx / .json")


if __name__ == "__main__":
    main()
