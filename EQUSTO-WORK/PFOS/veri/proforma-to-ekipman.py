# -*- coding: utf-8 -*-
"""MeffTech / proforma Excel (NO, MALIN CİNSİ, ÖLÇÜ, AD) → PFOS ekipman listesi."""
from __future__ import annotations

import importlib.util
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

VERI = Path(__file__).resolve().parent

_spec = importlib.util.spec_from_file_location(
    "pdf_ekipman_batch", VERI / "pdf-ekipman-batch.py"
)
_batch = importlib.util.module_from_spec(_spec)
assert _spec.loader
_spec.loader.exec_module(_batch)
write_ekipman_xlsx = _batch.write_ekipman_xlsx
PROJE_VERI = VERI / "proje-veri"

POZ_RE = re.compile(r"^[A-Z]\d{1,2}A?$", re.I)
HDR_NO = re.compile(r"^NO$", re.I)


def _cell(row: tuple, idx: int):
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_proforma_sheet(ws) -> tuple[str, list[dict]]:
    """TEKLİF FORMATI benzeri sayfadan ekipman satırları."""
    title = "ESPRESSOLAB MERSİN"
    ref = None
    rows: list[dict] = []
    bolum = "A"
    bolum_ad = "A- ESPRESSOLAB"
    in_table = False

    for row in ws.iter_rows(values_only=True):
        cells = tuple(row) if row else ()
        if not any(c is not None and str(c).strip() for c in cells):
            continue

        a, b = _cell(cells, 0), _cell(cells, 1)
        if a and HDR_NO.match(a):
            in_table = True
            continue
        if not in_table:
            if b and "ESPRESSOLAB" in b.upper():
                title = b.strip()
            for i, c in enumerate(cells):
                if c and str(c).strip().upper().startswith("REF"):
                    v = _cell(cells, i + 2) or _cell(cells, 11)
                    if v:
                        ref = v
                    break
            continue

        if b and not a and b.upper() not in ("TOPLAM:", "ÖZEL İSKONTO:", "GENEL TOPLAM:"):
            if not POZ_RE.match(b):
                bolum_ad = f"A- {b.strip()}"
            continue

        if a and POZ_RE.match(a):
            olcu = _cell(cells, 4) or "—"
            adet_raw = cells[8] if len(cells) > 8 else 1
            if adet_raw is None:
                adet: int | str = 1
            elif isinstance(adet_raw, (int, float)):
                adet = int(adet_raw)
            else:
                s = str(adet_raw).strip()
                adet = int(s) if s.isdigit() else s

            rows.append({
                "bolum": bolum,
                "bolumAd": bolum_ad,
                "poz": a.upper(),
                "ad": (b or "").strip(),
                "olcu": "—" if olcu in (None, "", "-") else olcu,
                "adet": adet,
            })

    if ref:
        title = f"{title} ({ref})"
    return title, rows


def pick_sheet(wb):
    for name in wb.sheetnames:
        if "TEKL" in name.upper() or "FORMAT" in name.upper():
            return wb[name]
    return wb.active


def convert(src: Path, out: Path | None = None) -> Path:
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = pick_sheet(wb)
    title, rows = parse_proforma_sheet(ws)
    wb.close()

    if not rows:
        raise SystemExit(f"Ekipman satırı bulunamadı: {src}")

    if out is None:
        stem = src.stem.lower().replace(" ", "-")
        out = PROJE_VERI / f"{stem}-ekipman-listesi.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)
    kaynak = f"{src.parent.name}/{src.name}"
    write_ekipman_xlsx(out, title, rows, kaynak)
    return out


def main():
    if len(sys.argv) < 2:
        print("Kullanım: python proforma-to-ekipman.py <proforma.xlsx> [çıktı.xlsx]")
        raise SystemExit(1)
    src = Path(sys.argv[1]).resolve()
    out = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else None
    path = convert(src, out)
    print("OK ->", path)


if __name__ == "__main__":
    main()
