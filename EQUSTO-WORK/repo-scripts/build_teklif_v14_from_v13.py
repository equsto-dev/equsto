# -*- coding: utf-8 -*-
"""
v13 arşiv şablonundan temiz equsto_teklif_v14.xlsx üretir.
Üst bilgi (1–4) + logo v13'ten korunur; satır 5+ sıfırdan yazılır.

  python scripts/build_teklif_v14_from_v13.py
"""
from __future__ import annotations

import re
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "arşiv" / "teklif formatı" / "equsto_teklif_v13.xlsx"
ARCHIVE_V14 = REPO / "arşiv" / "teklif formatı" / "equsto_teklif_v14.xlsx"
PUBLIC = REPO / "public" / "data" / "templates" / "equsto_teklif_v14.xlsx"
DIST = REPO / "dist" / "data" / "templates" / "equsto_teklif_v14.xlsx"

COLS = 13  # A..M
COL = {
    "bol": 1,
    "poz": 2,
    "ek": 3,
    "stok": 4,
    "tanim": 5,
    "marka": 6,
    "olcu": 7,
    "elk": 8,
    "gaz": 9,
    "adet": 10,
    "satis": 11,
    "toplam": 12,
    "doviz": 13,
}


def _unmerge_all_from(ws, min_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= min_row:
            try:
                ws.unmerge_cells(str(merged))
            except KeyError:
                pass


def _merge(ws, row: int, c1: int, c2: int) -> None:
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def _poz_code(grup: str, poz: str) -> str:
    g = (grup or "A").strip().upper()[:1] or "A"
    p = re.sub(r"\D", "", str(poz or "")) or "1"
    return g + p.zfill(2)


def _read_v13_product(ws, data_row: int, spec_row: int) -> tuple[dict, tuple] | None:
    stok = ws.cell(data_row, 5).value
    tanim = ws.cell(data_row, 6).value
    if not stok and not tanim:
        return None
    return (
        {
            "bol": "01",
            "poz": _poz_code(
                str(ws.cell(data_row, 2).value or "A"),
                str(ws.cell(data_row, 3).value or ""),
            ),
            "ek": ws.cell(data_row, 4).value,
            "stok": stok,
            "tanim": tanim,
            "marka": ws.cell(data_row, 7).value,
            "olcu": ws.cell(data_row, 8).value,
            "elk": ws.cell(data_row, 9).value,
            "gaz": ws.cell(data_row, 10).value,
            "adet": ws.cell(data_row, 11).value,
            "satis": ws.cell(data_row, 12).value,
            "doviz": ws.cell(data_row, 14).value or "EUR",
        },
        (
            ws.cell(spec_row, 1).value or "\U0001f4f7\nFoto\u011fraf",
            ws.cell(spec_row, 8).value or "",
        ),
    )


def _read_terms(ws) -> list[str]:
    terms: list[str] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        s = str(v).strip()
        if s == "ŞARTLARIMIZ":
            terms = [s]
        elif terms and (s.startswith("  ") or re.match(r"\s*\d+\.", s)):
            terms.append(s)
        elif terms and s.startswith("Banka:"):
            break
    return terms if len(terms) > 1 else [
        "ŞARTLARIMIZ",
        "  01.   Teklifimiz 7 (YEDİ) gün geçerlidir.",
        "  17.   Equsto.com yapay zekadan yardım alır; hata yapabilir. Nihai teyit satıcı onayındadır.",
    ]


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = COLS) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _write_data_row(ws, row: int, item: dict) -> None:
    ws.cell(row, COL["bol"]).value = item["bol"]
    ws.cell(row, COL["poz"]).value = item["poz"]
    ws.cell(row, COL["ek"]).value = item.get("ek") or ""
    ws.cell(row, COL["stok"]).value = item["stok"]
    ws.cell(row, COL["tanim"]).value = item["tanim"]
    ws.cell(row, COL["marka"]).value = item["marka"]
    ws.cell(row, COL["olcu"]).value = item["olcu"]
    ws.cell(row, COL["elk"]).value = item.get("elk") if item.get("elk") is not None else 0
    ws.cell(row, COL["elk"]).number_format = "0.0"
    ws.cell(row, COL["gaz"]).value = item.get("gaz") if item.get("gaz") is not None else 0
    ws.cell(row, COL["gaz"]).number_format = "0.0"
    ws.cell(row, COL["adet"]).value = item.get("adet") or 1
    ws.cell(row, COL["satis"]).value = item.get("satis")
    ws.cell(row, COL["satis"]).number_format = "#,##0.00"
    j = ws.cell(row, COL["adet"]).column_letter
    k = ws.cell(row, COL["satis"]).column_letter
    ws.cell(row, COL["toplam"]).value = f"={j}{row}*{k}{row}"
    ws.cell(row, COL["toplam"]).number_format = "#,##0.00"
    ws.cell(row, COL["doviz"]).value = item.get("doviz") or "EUR"


def _write_spec_row(ws, row: int, foto: str, acik: str, style_row: int) -> None:
    _copy_row_style(ws, style_row, row)
    _merge(ws, row, 1, 6)
    _merge(ws, row, 7, COLS)
    ws.cell(row, 1).value = foto
    ws.cell(row, 1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.cell(row, 7).value = acik
    ws.cell(row, 7).alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )


def build_v14() -> None:
    if not SRC.is_file():
        raise SystemExit("Kaynak yok: " + str(SRC))

    wb = load_workbook(SRC)
    ws = wb.active

    terms = _read_terms(ws)
    section_title = ws.cell(5, 1).value or "01. MUTFAK"

    products: list[tuple[dict, tuple]] = []
    for dr, sr in ((6, 7), (8, 9), (10, 11)):
        p = _read_v13_product(ws, dr, sr)
        if p:
            products.append(p)

    for i, (item, _) in enumerate(products):
        item["bol"] = "01"
        item["poz"] = "A" + str(i + 1).zfill(2)

    # Satır 5+ ve tüm birleştirmeleri temizle (eski v13 kalıntısı kalmasın)
    _unmerge_all_from(ws, 5)
    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)

    font_bold = Font(name="Arial", size=9, bold=True)
    font_hdr = Font(name="Arial", size=8, bold=True)

    headers = [
        "Böl.",
        "Poz",
        "EK",
        "Stok no",
        "Tanımı",
        "Marka",
        "Ölçü",
        "Elk. kW",
        "Gaz kW",
        "Adet",
        "Satış",
        "Toplam Satış",
        "Döviz",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, value=h)
        c.font = font_hdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(4, 14).value = None

    # Bölüm
    r = 5
    _copy_row_style(ws, 5, r)
    _unmerge_all_from(ws, r)
    _merge(ws, r, 1, COLS)
    ws.cell(r, 1).value = section_title
    ws.cell(r, 1).font = font_bold
    r += 1

    elk_parts: list[str] = []
    gaz_parts: list[str] = []
    sum_l: list[str] = []
    adet_refs: list[str] = []

    for item, (foto, acik) in products:
        _copy_row_style(ws, 6, r)
        _write_data_row(ws, r, item)
        j = ws.cell(r, COL["adet"]).column_letter
        h = ws.cell(r, COL["elk"]).column_letter
        i = ws.cell(r, COL["gaz"]).column_letter
        l = ws.cell(r, COL["toplam"]).column_letter
        elk_parts.append(f"{h}{r}*{j}{r}")
        gaz_parts.append(f"{i}{r}*{j}{r}")
        sum_l.append(f"{l}{r}")
        adet_refs.append(f"{j}{r}")
        r += 1

        _write_spec_row(ws, r, str(foto), str(acik), 7)
        ws.row_dimensions[r].height = 120
        r += 1

    # Gazlı toplam bağlantı
    _copy_row_style(ws, 6, r)
    ws.cell(r, COL["tanim"]).value = "Gazlı cihaz toplam bağlantısı (kW)"
    ws.cell(r, COL["tanim"]).font = font_bold
    ws.cell(r, COL["tanim"]).alignment = Alignment(horizontal="right", vertical="center")
    if gaz_parts:
        ws.cell(r, COL["gaz"]).value = "=" + "+".join(gaz_parts)
    ws.cell(r, COL["gaz"]).number_format = "0.0"
    r += 1

    # Sütun toplamları
    _copy_row_style(ws, 6, r)
    ws.cell(r, COL["olcu"]).value = "Sütun toplamları →"
    ws.cell(r, COL["olcu"]).alignment = Alignment(horizontal="right", vertical="center")
    if elk_parts:
        ws.cell(r, COL["elk"]).value = "=" + "+".join(elk_parts)
    ws.cell(r, COL["elk"]).number_format = "0.0"
    if gaz_parts:
        ws.cell(r, COL["gaz"]).value = "=" + "+".join(gaz_parts)
    ws.cell(r, COL["gaz"]).number_format = "0.0"
    if adet_refs:
        ws.cell(r, COL["adet"]).value = "=SUM(" + ",".join(adet_refs) + ")"
    r += 1

    # GENEL TOPLAM
    _copy_row_style(ws, 6, r)
    ws.cell(r, COL["adet"]).value = "GENEL TOPLAM"
    ws.cell(r, COL["adet"]).font = font_bold
    if sum_l:
        ws.cell(r, COL["toplam"]).value = "=" + "+".join(sum_l)
    ws.cell(r, COL["toplam"]).number_format = "#,##0.00"
    ws.cell(r, COL["doviz"]).value = "EUR"
    r += 1

    r += 1  # dip ile şartlar arası boşluk

    # ŞARTLARIMIZ + boş satır + maddeler + boş + banka
    _merge(ws, r, 1, COLS)
    ws.cell(r, 1).value = terms[0]
    ws.cell(r, 1).font = font_bold
    r += 1
    r += 1  # başlıktan sonra boş satır
    for t in terms[1:]:
        _merge(ws, r, 1, COLS)
        ws.cell(r, 1).value = t
        ws.row_dimensions[r].height = 14.1
        r += 1
    r += 1  # şartlar sonrası boş satır
    _merge(ws, r, 1, COLS)
    ws.cell(
        r,
        1,
        value="Banka: Garanti BBVA  |  Şube: Kağıthane  |  IBAN: TR00 0006 2000 0000 0000 0000 00",
    )
    r += 1
    _merge(ws, r, 1, COLS)
    ws.cell(
        r,
        1,
        value="Adres: —  |  Tel: —  |  E-posta: info@equsto.com  |  VKN: —",
    )
    foot = r + 2
    ws.cell(foot, 1).value = "Form No: EQS-TKL-014 (v14)"
    try:
        ws.unmerge_cells(f"I{foot}:N{foot}")
    except ValueError:
        pass
    _merge(ws, foot, 8, 14)
    ws.cell(foot, 8).value = "EQUSTO  |  info@equsto.com  |  Sayfa 1"
    ws.cell(foot, 8).alignment = Alignment(horizontal="right", vertical="center")

    ARCHIVE_V14.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ARCHIVE_V14)
    print("Yazıldı:", ARCHIVE_V14, "satır", ws.max_row, "birleşik", len(ws.merged_cells.ranges))
    for dst in (PUBLIC, DIST):
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(ARCHIVE_V14, dst)
        print("Kopyalandı:", dst)


def main() -> int:
    try:
        build_v14()
    except Exception as e:
        print("HATA:", e, file=sys.stderr)
        import traceback

        traceback.print_exc()
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
