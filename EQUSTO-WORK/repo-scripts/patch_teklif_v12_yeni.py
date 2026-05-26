# -*- coding: utf-8 -*-
"""
DEPRECATED — insert_cols ile bozulma riski. Bunun yerine:

  python scripts/build_teklif_v12_template.py

equsto_teklif_v12_yeni.xlsx — Elk/Gaz kW sütunları, dip toplam, logo.
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
ARCHIVE_YENI = REPO / "arşiv" / "teklif formatı" / "equsto_teklif_v12_yeni.xlsx"
LOGO_CANDIDATES = [
    REPO / "public" / "assets" / "equsto-logo-teklif.png",
    Path(
        r"C:\Users\User\.cursor\projects\c-D-Disk-EQUSTO-CURSOR-public\assets\c__Users_User_AppData_Roaming_Cursor_User_workspaceStorage_empty-window_images_equsto-logo-268b247b-7153-4e55-8ac0-28589723059a.png"
    ),
    REPO / "public" / "images" / "equsto-logo.png",
]
PUBLIC_LOGO = REPO / "public" / "assets" / "equsto-logo-teklif.png"
TARGETS = [
    ARCHIVE_YENI,
    REPO / "public" / "data" / "templates" / "equsto_teklif_v12.xlsx",
    REPO / "dist" / "data" / "templates" / "equsto_teklif_v12.xlsx",
]

COLS = 16  # A..P
ADET_COL = 13  # M
ELK_COL = 11  # K
GAZ_COL = 12  # L
SATIS_COL = 14  # N
TOPLAM_COL = 15  # O
DOVIZ_COL = 16  # P
HDR_ROW = 4
DATA_ROW = 6
SPEC_ROW = 7
KW_TOTAL_ROW = 18
SUBTOTAL_ROW = 19
GRAND_ROW = 20


def resolve_logo() -> Path:
    for p in LOGO_CANDIDATES:
        if p.is_file():
            PUBLIC_LOGO.parent.mkdir(parents=True, exist_ok=True)
            if p.resolve() != PUBLIC_LOGO.resolve():
                shutil.copy2(p, PUBLIC_LOGO)
            return PUBLIC_LOGO if PUBLIC_LOGO.is_file() else p
    raise FileNotFoundError("Equsto logo PNG bulunamadı")


def save_row_heights(ws) -> dict[int, float | None]:
    return {
        r: copy(ws.row_dimensions[r].height)
        for r in range(1, ws.max_row + 1)
        if ws.row_dimensions[r].height is not None
    }


def restore_row_heights(ws, heights: dict[int, float | None]) -> None:
    for r, h in heights.items():
        if h is not None:
            ws.row_dimensions[r].height = h


def patch_workbook(path: Path, logo: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    heights = save_row_heights(ws)

    # Adet öncesi 2 sütun: Elk kW, Gaz kW
    ws.insert_cols(11, 2)

    headers = [
        "Böl.",
        "Grup",
        "Poz",
        "EK",
        "Stok no",
        "Tanımı",
        "Kaynak",
        "Boy",
        "En",
        "Yük.",
        "Elk.\nkW",
        "Gaz\nkW",
        "Adet",
        "Satış",
        "Toplam Satış",
        "Döviz",
    ]
    for i, label in enumerate(headers, start=1):
        c = ws.cell(row=HDR_ROW, column=i, value=label)
        c.font = Font(name="Arial", size=8, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Örnek ürün (FBE20T)
    ws.cell(row=DATA_ROW, column=ELK_COL, value=26)
    ws.cell(row=DATA_ROW, column=GAZ_COL, value="")
    ws.cell(row=DATA_ROW, column=ADET_COL, value=1)
    ws.cell(row=DATA_ROW, column=SATIS_COL, value=8303.82).number_format = "#,##0.00"
    ws.cell(row=DATA_ROW, column=TOPLAM_COL, value=f"=M{DATA_ROW}*N{DATA_ROW}")
    ws.cell(row=DATA_ROW, column=TOPLAM_COL).number_format = "#,##0.00"
    ws.cell(row=DATA_ROW, column=DOVIZ_COL, value="EUR")

    for rng in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(rng))
        except (KeyError, ValueError):
            pass

    ws.merge_cells(start_row=SPEC_ROW, start_column=1, end_row=SPEC_ROW, end_column=7)
    ws.merge_cells(start_row=SPEC_ROW, start_column=8, end_row=SPEC_ROW, end_column=COLS)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=COLS)

    # Dip toplam — sütun kW + tutar
    ws.cell(row=KW_TOTAL_ROW, column=8, value="Sütun toplamları →").font = Font(
        name="Arial", size=9, bold=True
    )
    ws.cell(row=KW_TOTAL_ROW, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=KW_TOTAL_ROW, column=ELK_COL, value=f"=K{DATA_ROW}*M{DATA_ROW}")
    ws.cell(row=KW_TOTAL_ROW, column=ELK_COL).number_format = "0.0"
    ws.cell(row=KW_TOTAL_ROW, column=GAZ_COL, value=f"=L{DATA_ROW}*M{DATA_ROW}")
    ws.cell(row=KW_TOTAL_ROW, column=GAZ_COL).number_format = "0.0"

    ws.cell(row=SUBTOTAL_ROW, column=6, value="A. TOPLAM").font = Font(
        name="Arial", size=9, bold=True
    )
    ws.cell(row=SUBTOTAL_ROW, column=TOPLAM_COL, value=f"=O{DATA_ROW}")
    ws.cell(row=SUBTOTAL_ROW, column=TOPLAM_COL).number_format = "#,##0.00"
    ws.cell(row=SUBTOTAL_ROW, column=DOVIZ_COL, value="EUR")

    ws.cell(row=GRAND_ROW, column=6, value="GENEL TOPLAM  (KDV HARİÇ)").font = Font(
        name="Arial", size=9, bold=True
    )
    ws.cell(row=GRAND_ROW, column=TOPLAM_COL, value=f"=O{SUBTOTAL_ROW}")
    ws.cell(row=GRAND_ROW, column=TOPLAM_COL).number_format = "#,##0.00"
    ws.cell(row=GRAND_ROW, column=DOVIZ_COL, value="EUR")

    # Logo (M1:P2)
    ws["M1"] = None
    ws.merge_cells(start_row=1, start_column=13, end_row=2, end_column=COLS)
    img = XLImage(str(logo))
    img.width = 140
    img.height = 36
    ws.add_image(img, "M1")

    # Sütun genişlikleri
    widths = {
        "A": 4.5,
        "B": 5,
        "C": 4.5,
        "D": 3.5,
        "E": 12,
        "F": 28,
        "G": 10,
        "H": 7,
        "I": 7,
        "J": 7,
        "K": 7,
        "L": 7,
        "M": 5.5,
        "N": 11,
        "O": 12,
        "P": 5.5,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    restore_row_heights(ws, heights)
    wb.save(path)
    print("Güncellendi:", path)


def main() -> int:
    logo = resolve_logo()
    if not ARCHIVE_YENI.is_file():
        print("Dosya yok:", ARCHIVE_YENI)
        return 1
    patch_workbook(ARCHIVE_YENI, logo)
    for dst in TARGETS[1:]:
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(ARCHIVE_YENI, dst)
        print("Kopyalandı:", dst)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
