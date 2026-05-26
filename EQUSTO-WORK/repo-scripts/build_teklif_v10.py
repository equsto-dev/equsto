# -*- coding: utf-8 -*-
"""
equsto_teklif_v10.xlsx doldurur; ürün fotoğrafı varsa openpyxl ile gömer, yoksa silüet.

Kullanım:
  pip install openpyxl pillow
  python scripts/build_teklif_v10.py --input scripts/data/teklif-v10-sample.json -o out.xlsx

Fotoğraf (MVP):
  public/storage/products/{STOK_NO}/main.jpg
  veya --storage-root /var/www/equsto/storage/products

Satır JSON alanları: stok_no, poz, ad, marka, model, aciklama, adet, olcu, elk, gaz, birim,
  specs (liste veya metin), photo_path (opsiyonel)
"""
from __future__ import annotations

import argparse
import json
import sys
from copy import copy
from pathlib import Path
from typing import Any, Optional

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from teklif_v10_photos import (
    IMG_HEIGHT,
    IMG_WIDTH,
    REPO_ROOT,
    ensure_placeholder,
    resolve_photo_or_placeholder,
    resolve_storage_root,
)

PRODUCT_BLOCK_START = 5
PRODUCT_BLOCK_ROWS = 14
DATA_TEMPLATE_ROW = 6
SPEC_TEMPLATE_ROW = 7
SECTION_TEMPLATE_ROW = 5
COLS = 14
DEFAULT_TEMPLATE = REPO_ROOT / "public" / "data" / "templates" / "equsto_teklif_v14.xlsx"


def pad3(n: int) -> str:
    return f"{n:03d}"


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = COLS) -> None:
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst._style = copy(src._style)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def embed_photo(ws, anchor: str, stok_no: str, photo_path: Optional[str], storage_root: Path, placeholder: Path) -> None:
    path = resolve_photo_or_placeholder(
        stok_no,
        photo_path=photo_path,
        storage_root=storage_root,
        placeholder=placeholder,
    )
    cell = anchor.split("!")[-1] if "!" in anchor else anchor
    ws[cell].value = None
    img = XLImage(str(path))
    img.width, img.height = IMG_WIDTH, IMG_HEIGHT
    ws.add_image(img, anchor)


def fill_header(ws, ctx: dict[str, Any]) -> None:
    ws["K1"] = ctx.get("proformaNo") or ctx.get("teklifNo") or "EQS-DRAFT"
    ws["C2"] = ctx.get("proje") or ctx.get("isinAdi") or "Equsto Proje Fabrikası Teklifi"
    ws["H2"] = ctx.get("musteri") or "—"
    ws["K2"] = ctx.get("tarih") or ""
    ws["J3"] = ctx.get("kur") or 1
    ws["K3"] = ctx.get("kurLabel") or "TRY (PFOS net liste, KDV hariç — kur=1)"


def format_olcu_mm(olcu: str) -> str:
    import re

    m = re.search(
        r"(\d+(?:[.,]\d+)?)\s*[x×X*]\s*(\d+(?:[.,]\d+)?)\s*[x×X*]\s*(\d+(?:[.,]\d+)?)",
        str(olcu or ""),
    )
    if not m:
        return "—"
    en, boy, yuk = m.group(1), m.group(2), m.group(3)
    return f"{en}×{boy}×{yuk}"


def tanim_baslik(row: dict[str, Any]) -> str:
    stok = str(row.get("stok_no") or row.get("tip_kodu") or "").strip()
    ad = str(row.get("ad") or "").strip().upper()
    parts = ["A."]
    if ad:
        parts.append(ad)
    if stok and stok.upper() not in ad:
        parts.append(stok)
    return " ".join(parts)


def spec_text(row: dict[str, Any]) -> str:
    specs = row.get("specs")
    if isinstance(specs, list):
        return "\n".join(f"•  {s}" for s in specs if s)
    if specs:
        return str(specs)
    lines = []
    if row.get("ad"):
        lines.append(f"•  {row['ad']}")
    if row.get("olcu"):
        lines.append(f"•  Ölçü: {row['olcu']}")
    if row.get("stok_no"):
        lines.append(f"•  Stok / tip: {row['stok_no']}")
    if row.get("marka"):
        lines.append(f"•  Marka: {row['marka']}")
    return "\n".join(lines) if lines else "•  PFOS katalog kalemi"


def build_product_block(
    ws,
    zones: list[dict[str, Any]],
    storage_root: Path,
    placeholder: Path,
    embed_images: bool = True,
) -> None:
    ws.delete_rows(PRODUCT_BLOCK_START, PRODUCT_BLOCK_ROWS)
    row_num = PRODUCT_BLOCK_START
    poz_global = 0
    sum_refs: list[str] = []

    kur = float(ws["J3"].value or 1) or 1

    for zi, zone in enumerate(zones):
        ws.insert_rows(row_num)
        copy_row_style(ws, SECTION_TEMPLATE_ROW, row_num)
        try:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=COLS)
        except ValueError:
            pass
        ws.cell(row=row_num, column=1).value = f"{zi + 1:02d}. {str(zone.get('label', 'BÖLÜM')).upper()}"
        row_num += 1

        for r in zone.get("rows") or []:
            poz_global += 1
            adet = int(r.get("adet") or 1)
            birim_tl = float(r.get("birim") or 0)
            birim_eur = round(birim_tl / kur, 2) if kur > 1 and birim_tl else birim_tl
            olcu_cell = format_olcu_mm(r.get("olcu") or "")

            ws.insert_rows(row_num)
            copy_row_style(ws, DATA_TEMPLATE_ROW, row_num)
            dr = row_num
            ws.cell(row=dr, column=1).value = f"{zi + 1:02d}"
            ws.cell(row=dr, column=2).value = "A"
            ws.cell(row=dr, column=3).value = f"{poz_global:02d}"
            ws.cell(row=dr, column=4).value = ""
            ws.cell(row=dr, column=5).value = r.get("stok_no") or r.get("tip_kodu") or ""
            ws.cell(row=dr, column=6).value = tanim_baslik(r)
            ws.cell(row=dr, column=7).value = r.get("marka") or ""
            ws.cell(row=dr, column=8).value = dims[0]
            ws.cell(row=dr, column=9).value = dims[1]
            ws.cell(row=dr, column=10).value = dims[2]
            ws.cell(row=dr, column=11).value = adet
            ws.cell(row=dr, column=12).value = birim_eur
            ws.cell(row=dr, column=13).value = f"=K{dr}*L{dr}"
            ws.cell(row=dr, column=14).value = "EUR"
            sum_refs.append(f"M{dr}")
            row_num += 1

            ws.insert_rows(row_num)
            copy_row_style(ws, SPEC_TEMPLATE_ROW, row_num)
            sr = row_num
            stok = str(r.get("stok_no") or r.get("tip_kodu") or "")
            if embed_images:
                embed_photo(
                    ws,
                    f"A{sr}",
                    stok,
                    r.get("photo_path"),
                    storage_root,
                    placeholder,
                )
            else:
                ws.cell(row=sr, column=1).value = "📷\nFotoğraf"
            try:
                ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=7)
                ws.merge_cells(start_row=sr, start_column=8, end_row=sr, end_column=COLS)
            except ValueError:
                pass
            ws.cell(row=sr, column=8).value = spec_text(r)
            ws.cell(row=sr, column=8).alignment = copy(ws.cell(row=SPEC_TEMPLATE_ROW, column=8).alignment)
            row_num += 1

    ws.insert_rows(row_num)
    copy_row_style(ws, DATA_TEMPLATE_ROW, row_num)
    ws.cell(row=row_num, column=6).value = "A. TOPLAM"
    sum_formula = "+".join(sum_refs) if sum_refs else "0"
    ws.cell(row=row_num, column=13).value = f"={sum_formula}"
    total_row = row_num
    row_num += 1

    ws.insert_rows(row_num)
    copy_row_style(ws, DATA_TEMPLATE_ROW, row_num)
    ws.cell(row=row_num, column=6).value = "GENEL TOPLAM  (KDV HARİÇ)"
    ws.cell(row=row_num, column=13).value = f"=M{total_row}"


def load_payload(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if "zones" in data:
        return data
    rows = data.get("rows") or []
    return {
        "ctx": data.get("ctx") or {},
        "zones": [{"label": data.get("zoneLabel") or "MUTFAK", "rows": rows}],
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Equsto teklif v10 Excel üretici")
    ap.add_argument("--input", "-i", type=Path, help="JSON: ctx + zones veya rows")
    ap.add_argument("--output", "-o", type=Path, default=Path("equsto-teklif-out.xlsx"))
    ap.add_argument("--template", "-t", type=Path, default=DEFAULT_TEMPLATE)
    ap.add_argument("--storage-root", type=Path, help="products/{STOK}/main.jpg kökü")
    ap.add_argument("--placeholder", type=Path, help="Silüet PNG")
    ap.add_argument("--no-images", action="store_true", help="Foto gömme (sadece metin)")
    ap.add_argument("--demo", action="store_true", help="Örnek JSON ile üret")
    args = ap.parse_args()

    if args.demo:
        sample = REPO_ROOT / "scripts" / "data" / "teklif-v10-sample.json"
        if not sample.is_file():
            print("Demo dosyası yok:", sample, file=sys.stderr)
            return 1
        args.input = sample

    if not args.input or not args.input.is_file():
        ap.error("--input veya --demo gerekli")

    if not args.template.is_file():
        print("Şablon bulunamadı:", args.template, file=sys.stderr)
        return 1

    storage_root = resolve_storage_root(args.storage_root)
    storage_root.mkdir(parents=True, exist_ok=True)
    placeholder = ensure_placeholder(args.placeholder or (REPO_ROOT / "public" / "data" / "templates" / "teklif-photo-placeholder.png"))

    payload = load_payload(args.input)
    wb = load_workbook(args.template)
    ws = wb.worksheets[0]
    fill_header(ws, payload.get("ctx") or {})
    build_product_block(
        ws,
        payload.get("zones") or [],
        storage_root,
        placeholder,
        embed_images=not args.no_images,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print("Yazıldı:", args.output.resolve())
    print("Foto kökü:", storage_root.resolve())
    return 0


if __name__ == "__main__":
    sys.exit(main())
