# -*- coding: utf-8 -*-
"""
equsto_teklif_v12.xlsx — Inoksan proforma (14 sütun: ölçü tek hücre, Elk/Gaz kW).
Şablonu sıfırdan üretir; arşiv + public/data/templates kopyalanır.

  python scripts/build_teklif_v12_template.py
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side

REPO = Path(__file__).resolve().parent.parent
ARCHIVE = REPO / "arşiv" / "teklif formatı" / "equsto_teklif_v12.xlsx"
ARCHIVE_V13 = REPO / "arşiv" / "teklif formatı" / "equsto_teklif_v13.xlsx"
PUBLIC = REPO / "public" / "data" / "templates" / "equsto_teklif_v13.xlsx"
DIST = REPO / "dist" / "data" / "templates" / "equsto_teklif_v13.xlsx"
LOGO = REPO / "public" / "assets" / "equsto-logo-teklif.png"

COLS = 14  # A..N — ölçü: En×Boy×Yük tek sütun (H)
LAST_COL = "N"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_H = Border(top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=THIN)

FONT_HDR = Font(name="Arial", size=8, bold=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_META = Font(name="Arial", size=9)
FONT_META_B = Font(name="Arial", size=9, bold=True)
FONT_SECTION = Font(name="Arial", size=9, bold=True)
FONT_DATA = Font(name="Arial", size=9)
FONT_DATA_B = Font(name="Arial", size=9, bold=True)
FONT_SPEC = Font(name="Arial", size=8)
FONT_FOOT = Font(name="Arial", size=8)

ALIGN_CC = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CL = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CR = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_SPEC = Alignment(horizontal="left", vertical="top", wrap_text=True)

HEADERS = [
    "Böl.",
    "Grup",
    "Poz",
    "EK",
    "Stok no",
    "Tanımı",
    "Marka",
    "Ölçü",
    "Elk.\nkW",
    "Gaz\nkW",
    "Adet",
    "Satış",
    "Toplam Satış",
    "Döviz",
]

SAMPLE_SPEC = (
    "Kombi fırın, dokunmatik kontrol panelli, elektrik, 10xGN2/1 tepsi kap., 26kW 400V 50-60Hz\n"
    "•  10 × GN2/1 tepsi kapasitesi, akıllı buhar jeneratörlü\n"
    "•  Dokunmatik kontrol paneli, çoklu dil\n"
    "•  Otomatik yıkama programları\n"
    "•  PID/PWM hassas ısı kontrolü\n"
    "•  Çift kademeli kapı kilit, Low-E çift cam\n"
    "•  Faydalı iç hacim ve raflar arası mesafe proje ölçüsüne göre teyit edilir"
)

SECTION_ROW = 5
DATA_ROW = 6
SPEC_ROW = 7
KW_TOTAL_ROW = 18
SUBTOTAL_ROW = 19
GRAND_ROW = 20
PRODUCT_LAST_ROW = 17


def set_row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


def style_range(ws, row: int, c1: int, c2: int, font=None, alignment=None, border=None) -> None:
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border


def merge_row(ws, row: int, c1: int, c2: int) -> None:
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def add_logo(ws, logo_path: Path) -> None:
    if not logo_path.is_file():
        return
    ws.merge_cells(start_row=1, start_column=11, end_row=2, end_column=COLS)
    img = XLImage(str(logo_path))
    img.width = 140
    img.height = 36
    ws.add_image(img, "K1")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Teklif"

    widths = {
        "A": 4.5,
        "B": 5,
        "C": 4.5,
        "D": 3.5,
        "E": 12,
        "F": 28,
        "G": 10,
        "H": 14,
        "I": 7,
        "J": 7,
        "K": 5.5,
        "L": 11,
        "M": 12,
        "N": 5.5,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # —— Üst bilgi ——
    merge_row(ws, 1, 1, 8)
    ws["A1"] = "PROFORMA FATURA"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_CL

    ws["I1"] = "Sayı:"
    ws["I1"].font = FONT_META_B
    merge_row(ws, 1, 10, 10)
    ws["J1"] = "EQS-2026-001"
    ws["J1"].font = FONT_META
    ws["J1"].alignment = ALIGN_CL

    merge_row(ws, 2, 1, 2)
    ws["A2"] = "Proje:"
    ws["A2"].font = FONT_META_B
    merge_row(ws, 2, 3, 6)
    ws["C2"] = "Coca-Cola İçecek — FBE20T Isparta"
    ws["C2"].font = FONT_META

    ws["G2"] = "Müşteri:"
    ws["G2"].font = FONT_META_B
    ws["H2"] = "Alper Bey"
    ws["H2"].font = FONT_META

    ws["I2"] = "Tarih:"
    ws["I2"].font = FONT_META_B
    ws["J2"] = "22.04.2026"
    ws["J2"].font = FONT_META
    ws["J2"].alignment = ALIGN_CL

    set_row_height(ws, 3, 37.5)
    merge_row(ws, 3, 1, 8)
    ws["A3"] = "TCMB Efektif Satış Kuru – 22.04.2026"
    ws["A3"].font = FONT_META
    ws["I3"] = "EUR/TRY"
    ws["I3"].font = FONT_META
    ws["I3"].alignment = ALIGN_CR
    ws["J3"] = 52.8238
    ws["J3"].number_format = '"₺"#,##0.00'
    ws["J3"].font = FONT_META
    ws["J3"].alignment = ALIGN_CR
    merge_row(ws, 3, 11, COLS)

    add_logo(ws, LOGO)

    # —— Tablo başlıkları ——
    set_row_height(ws, 4, 27.95)
    for i, label in enumerate(HEADERS, start=1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONT_HDR
        c.alignment = ALIGN_CC
        c.border = BORDER_H
    style_range(ws, 4, 8, COLS, alignment=ALIGN_CR)

    # —— Bölüm + ürün ——
    set_row_height(ws, SECTION_ROW, 18)
    merge_row(ws, SECTION_ROW, 1, COLS)
    ws.cell(row=SECTION_ROW, column=1, value="01. MUTFAK")
    ws.cell(row=SECTION_ROW, column=1).font = FONT_SECTION
    ws.cell(row=SECTION_ROW, column=1).alignment = ALIGN_CL
    ws.cell(row=SECTION_ROW, column=1).border = BORDER_H

    set_row_height(ws, DATA_ROW, 16)
    ws.cell(row=DATA_ROW, column=1, value="01").alignment = ALIGN_CC
    ws.cell(row=DATA_ROW, column=2, value="A").alignment = ALIGN_CC
    ws.cell(row=DATA_ROW, column=3, value="01").alignment = ALIGN_CC
    ws.cell(row=DATA_ROW, column=4, value="").alignment = ALIGN_CC
    ws.cell(row=DATA_ROW, column=5, value="INO-FBE20T").font = FONT_DATA_B
    ws.cell(row=DATA_ROW, column=5).alignment = ALIGN_CC
    ws.cell(row=DATA_ROW, column=6, value="A. SICAK MUTFAK INO-FBE20T").font = FONT_DATA_B
    ws.cell(row=DATA_ROW, column=6).alignment = ALIGN_CL
    ws.cell(row=DATA_ROW, column=7, value="İnoksan").alignment = ALIGN_CL
    ws.cell(row=DATA_ROW, column=8, value="1070×1180×1250").alignment = ALIGN_CR
    ws.cell(row=DATA_ROW, column=9, value=26).alignment = ALIGN_CR
    ws.cell(row=DATA_ROW, column=9).number_format = "0.0"
    ws.cell(row=DATA_ROW, column=10, value=0).alignment = ALIGN_CR
    ws.cell(row=DATA_ROW, column=10).number_format = "0.0"
    ws.cell(row=DATA_ROW, column=11, value=1).alignment = ALIGN_CC
    ws.cell(row=DATA_ROW, column=12, value=8303.82).number_format = "#,##0.00"
    ws.cell(row=DATA_ROW, column=12).alignment = ALIGN_CR
    ws.cell(row=DATA_ROW, column=13, value=f"=K{DATA_ROW}*L{DATA_ROW}")
    ws.cell(row=DATA_ROW, column=13).number_format = "#,##0.00"
    ws.cell(row=DATA_ROW, column=13).alignment = ALIGN_CR
    ws.cell(row=DATA_ROW, column=14, value="EUR").alignment = ALIGN_CC
    style_range(ws, DATA_ROW, 1, COLS, font=FONT_DATA, border=BORDER_BOTTOM)

    set_row_height(ws, SPEC_ROW, 120)
    merge_row(ws, SPEC_ROW, 1, 7)
    ws.cell(row=SPEC_ROW, column=1, value="📷\nFotoğraf")
    ws.cell(row=SPEC_ROW, column=1).font = FONT_SPEC
    ws.cell(row=SPEC_ROW, column=1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    merge_row(ws, SPEC_ROW, 8, COLS)
    ws.cell(row=SPEC_ROW, column=8, value=SAMPLE_SPEC)
    ws.cell(row=SPEC_ROW, column=8).font = FONT_SPEC
    ws.cell(row=SPEC_ROW, column=8).alignment = ALIGN_SPEC
    ws.cell(row=SPEC_ROW, column=8).border = BORDER_BOTTOM

    # —— Dip toplamlar (yalnızca A–P) ——
    ws.cell(row=KW_TOTAL_ROW, column=8, value="Sütun toplamları →").font = FONT_DATA_B
    ws.cell(row=KW_TOTAL_ROW, column=8).alignment = ALIGN_CR
    ws.cell(row=KW_TOTAL_ROW, column=9, value=f"=I{DATA_ROW}*K{DATA_ROW}")
    ws.cell(row=KW_TOTAL_ROW, column=9).number_format = "0.0"
    ws.cell(row=KW_TOTAL_ROW, column=10, value=f"=J{DATA_ROW}*K{DATA_ROW}")
    ws.cell(row=KW_TOTAL_ROW, column=10).number_format = "0.0"
    ws.cell(row=KW_TOTAL_ROW, column=11, value=f"=SUM(K{DATA_ROW}:K{PRODUCT_LAST_ROW})")
    ws.cell(row=KW_TOTAL_ROW, column=11).alignment = ALIGN_CC

    ws.cell(row=SUBTOTAL_ROW, column=6, value="A. TOPLAM").font = FONT_DATA_B
    ws.cell(row=SUBTOTAL_ROW, column=13, value=f"=M{DATA_ROW}")
    ws.cell(row=SUBTOTAL_ROW, column=13).number_format = "#,##0.00"
    ws.cell(row=SUBTOTAL_ROW, column=13).alignment = ALIGN_CR
    ws.cell(row=SUBTOTAL_ROW, column=14, value="EUR").alignment = ALIGN_CC

    ws.cell(row=GRAND_ROW, column=6, value="GENEL TOPLAM  (KDV HARİÇ)").font = FONT_DATA_B
    ws.cell(row=GRAND_ROW, column=13, value=f"=M{SUBTOTAL_ROW}")
    ws.cell(row=GRAND_ROW, column=13).number_format = "#,##0.00"
    ws.cell(row=GRAND_ROW, column=13).alignment = ALIGN_CR
    ws.cell(row=GRAND_ROW, column=14, value="EUR").alignment = ALIGN_CC

    # —— Şartlar ——
    sart_row = 22
    merge_row(ws, sart_row, 1, COLS)
    ws.cell(row=sart_row, column=1, value="ŞARTLARIMIZ").font = FONT_SECTION
    terms = [
        "  01.   Teklifimiz 7 (YEDİ) gün geçerlidir.",
        "  02.   Fiyatlarımıza KDV dahil değildir, faturada ayrıca eklenecektir.",
        "  03.   Faturamız TL olarak kesilecektir. Tutarlar TCMB Efektif Satış Kuru üzerinden hesaplanmıştır.",
        "  04.   Ödeme; siparişte %50 peşin banka havalesi, kalanı mal tesliminden önce banka havalesi şeklindedir.",
        "  05.   Ödeme şartlarının yerine getirilmesi ile birlikte teklif sipariş statüsüne geçer.",
        "  06.   Montaj satıcıya aittir. Her türlü tesisat ve sarf malzemesi alıcıya aittir.",
        "  07.   Nakliye ve nakliye sigortası satıcıya aittir.",
        "  08.   Her türlü yatay ve dikey taşımacılık alıcıya aittir. Kamyon üstü teslimdir.",
        "  09.   Teslim yeri müşteri adresidir.",
        "  10.   Teslim süresi: kesin siparişinizi takiben 6-8 hafta (üretim programına göre teyit).",
        "  11.   Soğuk odalarda dış ünite mesafesi 10-12 m olarak fiyatlandırılmıştır.",
        "  12.   İş kapsamında değişiklik olması durumunda karşılıklı mutabakatla teklif revize edilir.",
        "  13.   Ölçü bekler çözümünün siparişten sonraki 1 ay içinde tamamlanması gerekir.",
        "  14.   Zamanında ödenmeyen bedel için aylık %5 vade farkı uygulanır.",
        "  15.   Depoda 1 aydan fazla bekleyen mallar için aylık sipariş bedelinin %5'i depo kirasıdır.",
        "  16.   Dijital mutabakatlar yazılı mutabakat gibi sonuç doğurur.",
    ]
    r = sart_row + 1
    for t in terms:
        merge_row(ws, r, 1, COLS)
        ws.cell(row=r, column=1, value=t).font = FONT_SPEC
        ws.cell(row=r, column=1).alignment = ALIGN_CL
        set_row_height(ws, r, 14)
        r += 1

    bank_row = r + 1
    merge_row(ws, bank_row, 1, COLS)
    ws.cell(
        row=bank_row,
        column=1,
        value="Banka: Garanti BBVA  |  Şube: Kağıthane  |  IBAN: TR00 0006 2000 0000 0000 0000 00",
    ).font = FONT_SPEC

    addr_row = bank_row + 1
    merge_row(ws, addr_row, 1, COLS)
    ws.cell(
        row=addr_row,
        column=1,
        value="Adres: —  |  Tel: —  |  E-posta: info@equsto.com  |  VKN: —",
    ).font = FONT_SPEC

    foot_row = addr_row + 2
    ws.cell(row=foot_row, column=1, value="Form No: EQS-TKL-012 (v12)").font = FONT_FOOT
    merge_row(ws, foot_row, 9, COLS)
    ws.cell(
        row=foot_row,
        column=9,
        value="EQUSTO  |  info@equsto.com  |  Sayfa 1",
    ).font = FONT_FOOT
    ws.cell(row=foot_row, column=9).alignment = Alignment(horizontal="right", vertical="center")

    return wb


def _save_wb(wb: Workbook, path: Path) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        print("Yazıldı:", path)
        return True
    except PermissionError:
        print("UYARI: kilitli — atlandı:", path)
        return False


def main() -> int:
    wb = build_workbook()
    source = ARCHIVE
    if _save_wb(wb, ARCHIVE):
        source = ARCHIVE
    if _save_wb(wb, ARCHIVE_V13):
        source = ARCHIVE_V13
    for dst in (PUBLIC, DIST):
        dst.parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(source, dst)
            print("Kopyalandı:", dst)
        except OSError as e:
            print("Kopya hatası:", dst, e)
    return 0


if __name__ == "__main__":
    sys.exit(main())
