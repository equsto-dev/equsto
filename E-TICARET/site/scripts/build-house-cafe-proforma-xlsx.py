#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""19- THE HOUSE CAFE → steakhouse (2018-199-3) PROFORMA + Sayfa1 şablonu."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
OUT_DESKTOP = Path.home() / "Desktop" / "19-house-cafe.xlsx"
OUT_DATA = ROOT / "public" / "data" / "geo" / "19-house-cafe.xlsx"

# (bölüm slug PROFORMA, bölüm başlık, poz, ad, ölçü, adet)
ROWS: list[tuple[str, str, str, str, str, int]] = [
    ("özel üretim", "A- ÖZEL ÜRETİM", "1", "PASLANMAZ TEZGAH", "-", 1),
    ("soğuk oda", "B- SOĞUK ODA", "2", "PANEL SOĞUK ODA", "250*200*220", 1),
    ("soğuk oda", "B- SOĞUK ODA", "3", "PANEL BUZ ODASI", "200*150*220", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "4", "KONVEKSİYONLU FIRIN", "-", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "5", "MAYALAMA KABİNİ", "-", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "6", "4 BRÜLÖRLÜ KUZİNE", "-", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "7", "IZGARA (PLEYT)", "-", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "8", "FRİTÖZ (2X10 LT)", "-", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "9", "MAKARNA HAŞLAMA", "-", 1),
    ("pişirici ekipmanlar", "C- PİŞİRİCİ EKİPMANLAR", "10", "DAVLUMBAZ", "-", 1),
    ("hazırlık ekipmanları", "D- HAZIRLIK EKİPMANLARI", "11", "HAMUR YOĞURMA MAKİNASI", "-", 1),
    ("hazırlık ekipmanları", "D- HAZIRLIK EKİPMANLARI", "12", "SEBZE DOĞRAMA MAKİNASI", "-", 1),
    ("hazırlık ekipmanları", "D- HAZIRLIK EKİPMANLARI", "13", "ET KIYMA MAKİNASI", "-", 1),
    ("hazırlık ekipmanları", "D- HAZIRLIK EKİPMANLARI", "14", "EL BLENDERI", "-", 1),
    ("hazırlık ekipmanları", "D- HAZIRLIK EKİPMANLARI", "15", "TERAZİ (30 KG)", "-", 1),
    ("bulaşıkhane", "E- BULAŞIKHANE", "16", "BULAŞIK MAKİNASI (GİYOTİN TİP)", "-", 1),
    ("bulaşıkhane", "E- BULAŞIKHANE", "17", "BULAŞIK MAKİNASI GİRİŞ TEZGAHI", "-", 1),
    ("bulaşıkhane", "E- BULAŞIKHANE", "18", "BULAŞIK MAKİNASI ÇIKIŞ TEZGAHI", "-", 1),
    ("bulaşıkhane", "E- BULAŞIKHANE", "19", "ÖN YIKAMA DUŞU", "-", 1),
    ("bulaşıkhane", "E- BULAŞIKHANE", "20", "DUVAR RAFI", "-", 1),
    ("bulaşıkhane", "E- BULAŞIKHANE", "21", "EVYE", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "22", "BARDAK YIKAMA MAKİNASI", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "23", "BUZ MAKİNASI", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "24", "PORTAKAL SIKMA MAKİNASI", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "25", "KATI MEYVE SIKACAĞI", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "26", "BLENDER", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "27", "BAR BUZDOLABI", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "28", "BAR BUZDOLABI", "-", 1),
    ("bar ekipmanları", "F- BAR EKİPMANLARI", "29", "BAR BUZDOLABI", "-", 1),
    ("soğutucular", "G- SOĞUTUCULAR", "30", "DİK TİP BUZDOLABI (TEK KAPILI)", "-", 1),
    ("soğutucular", "G- SOĞUTUCULAR", "31", "DİK TİP DONDURUCU (TEK KAPILI)", "-", 1),
    ("soğutucular", "G- SOĞUTUCULAR", "32", "TEZGAH ALTI BUZDOLABI (2 KAPILI)", "-", 1),
    ("soğutucular", "G- SOĞUTUCULAR", "33", "TEZGAH ALTI BUZDOLABI (3 KAPILI)", "-", 1),
    ("soğutucular", "G- SOĞUTUCULAR", "34", "SALATA HAZIRLIK BUZDOLABI", "-", 1),
    ("soğutucular", "G- SOĞUTUCULAR", "35", "PİZZA HAZIRLIK BUZDOLABI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "36", "ÇALIŞMA TEZGAHI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "37", "ÇALIŞMA TEZGAHI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "38", "ÇALIŞMA TEZGAHI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "39", "ÇALIŞMA TEZGAHI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "40", "ÇALIŞMA TEZGAHI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "41", "EVYELİ TEZGAH", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "42", "EVYELİ TEZGAH", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "43", "DUVAR RAFI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "44", "DUVAR RAFI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "45", "İSTİF RAFI", "-", 1),
    ("tezgahlar", "H- TEZGAHLAR", "46", "İSTİF RAFI", "-", 1),
    ("diğer ekipmanlar", "I- DİĞER EKİPMANLAR", "47", "SİNEK ÖLDÜRÜCÜ", "-", 1),
    ("diğer ekipmanlar", "I- DİĞER EKİPMANLAR", "48", "ÇÖP KOVASI", "-", 1),
]

TITLE = "19- THE HOUSE CAFE"


def write_workbook(path: Path) -> None:
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "PROFORMA"

    ws_p["C1"] = TITLE
    ws_p["C1"].font = Font(bold=True, size=14, color="FF0000")

    row = 3
    cur_slug: str | None = None
    for slug, _heading, _poz, ad, _olcu, _adet in ROWS:
        if slug != cur_slug:
            ws_p.cell(row=row, column=3, value=slug)
            row += 1
            cur_slug = slug
        ws_p.cell(row=row, column=5, value="-")
        ws_p.cell(row=row, column=7, value=ad)
        row += 1

    ws_s = wb.create_sheet("Sayfa1")
    headers = [
        "Poz",
        "Ürün",
        None,
        None,
        None,
        "Ölçü",
        "Adet",
        "Birim EUR",
        "Tutar",
        "İndirimli",
        "İnd. birim",
        "İnd. tutar",
    ]
    for col, h in enumerate(headers, 1):
        if h:
            ws_s.cell(row=1, column=col, value=h)

    for i, (_slug, _heading, poz, ad, olcu, adet) in enumerate(ROWS, 2):
        ws_s.cell(row=i, column=1, value=poz)
        ws_s.cell(row=i, column=2, value=ad)
        ws_s.cell(row=i, column=6, value=olcu if olcu != "-" else None)
        ws_s.cell(row=i, column=7, value=adet)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    write_workbook(OUT_DESKTOP)
    write_workbook(OUT_DATA)
    print(f"OK {OUT_DESKTOP}")
    print(f"OK {OUT_DATA} ({len(ROWS)} kalem)")


if __name__ == "__main__":
    main()
