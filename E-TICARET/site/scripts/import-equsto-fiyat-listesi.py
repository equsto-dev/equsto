#!/usr/bin/env python3
"""
EQUSTO Fiyat Listesi 2026 — tüm sekmeler → PFOS + e-ticaret JSON + görseller + shop hazırlık

Kaynak: c:\\D Disk\\FİYAT LİSTELERİ\\EQUSTO - FİYAT LİSTESİ.xlsx

  python scripts/import-equsto-fiyat-listesi.py
  python scripts/import-equsto-fiyat-listesi.py --dry-run
  python scripts/import-equsto-fiyat-listesi.py --sheet KCT02
"""
from __future__ import annotations

import argparse
import json
import re
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(r"c:\D Disk\FİYAT LİSTELERİ\EQUSTO - FİYAT LİSTESİ.xlsx")
OUT_BASE = ROOT / "public/data/fiyat-listeleri/equsto/2026-fiyat-listesi"
OUT_IMG_BASE = ROOT / "public/images/catalog/equsto/fiyat-listesi"

ISKONTO = 0.10
KDV = 0.20
ETICARET_DERINLIK_CM = 70
ETICARET_EN_EXTRA = (120, 140, 160, 190)
# Davlumbaz e-ticaret vitrin — 4 seri × 10 ölçü = 40 ürün (ara genişlik/derinlikler dahil)
DAVLUMBAZ_ETIC_VITRIN: dict[str, tuple[tuple[int, int], ...]] = {
    "KDAVDT01": (
        (100, 100),
        (100, 120),
        (125, 100),
        (150, 120),
        (175, 100),
        (200, 100),
        (200, 120),
        (250, 100),
        (275, 120),
        (300, 100),
    ),
    "KDAVDTF02": (
        (100, 120),
        (110, 90),
        (125, 120),
        (150, 90),
        (175, 120),
        (200, 90),
        (200, 120),
        (225, 90),
        (250, 120),
        (300, 90),
    ),
    "KDAVOT01": (
        (100, 200),
        (125, 150),
        (150, 200),
        (175, 250),
        (200, 100),
        (200, 200),
        (225, 250),
        (250, 200),
        (275, 150),
        (300, 100),
    ),
    "KDAVOTF02": (
        (100, 200),
        (110, 250),
        (125, 200),
        (150, 250),
        (175, 200),
        (200, 200),
        (225, 250),
        (250, 200),
        (275, 250),
        (300, 150),
    ),
}
MALZEME = "AISI 18/10 (304 kalite) paslanmaz çelik mamül"
KAYNAK = "equsto-fiyat-listesi-2026"
SKU_PREFIX = "EQ"


def fetch_eur_try() -> tuple[float, str, bool]:
    url = "https://www.tcmb.gov.tr/kurlar/today.xml"
    try:
        with urllib.request.urlopen(url, timeout=15) as res:
            xml = res.read()
        root = ET.fromstring(xml)
        tarih = root.attrib.get("Tarih", "")
        for cur in root.findall("Currency"):
            if cur.attrib.get("Kod") == "EUR":
                selling = cur.findtext("BanknoteSelling")
                rate = float(selling)
                if rate > 0:
                    return rate, tarih, False
    except Exception as e:
        print(f"[kur] fallback: {e}")
    return 53.05, "", True


def norm_tr(s: str) -> str:
    return (
        (s or "")
        .upper()
        .replace("İ", "I")
        .replace("Ç", "C")
        .replace("Ş", "S")
        .replace("Ğ", "G")
        .replace("Ö", "O")
        .replace("Ü", "U")
    )


def norm_sheet_kod(name: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", norm_tr(name))


def sheet_kods(sheet_name: str) -> list[str]:
    raw = norm_tr(sheet_name)
    return [re.sub(r"[^A-Z0-9]", "", k) for k in re.findall(r"K[A-Z]{1,6}\d{2}", raw)]


def primary_kod(sheet_name: str) -> str:
    parts = sheet_kods(sheet_name)
    if parts:
        return parts[0]
    return norm_sheet_kod(sheet_name)


def is_dual_product_sheet(sheet_name: str) -> bool:
    return len(sheet_kods(sheet_name)) >= 2


def slug_dir(kod: str) -> str:
    return kod.lower().replace("_", "-")


def parse_title(raw: str, kod: str) -> str:
    if not raw:
        return kod
    t = re.sub(r"\s+", " ", str(raw).strip())
    t = re.sub(r"\s*[-–.]?\s*" + re.escape(kod) + r"\s*$", "", t, flags=re.I)
    t = re.sub(r"\s*[-–.]?\s*K[ÇC]T\s*\d{2}\s*$", "", t, flags=re.I)
    return t.strip(" -.")


def dept_for_kod(kod: str) -> str:
    if kod.startswith("KDAV"):
        return "davlumbaz"
    if kod.startswith("KDUVD"):
        return "dolap"
    if kod.startswith(("KDUVR", "KSDUVR")):
        return "istif"
    return "tezgah"


def category_for_kod(kod: str, dept: str) -> str:
    if dept == "davlumbaz":
        if "OT" in kod:
            return "davlumbaz-orta-tip"
        return "davlumbaz-duvar-tipi"
    if dept == "istif":
        return "duvar-rafi"
    if dept == "dolap":
        return "duvar-dolabi"
    if kod.startswith("KHCT"):
        return "hareketli-calisma-tezgahi"
    if kod.startswith("KMERTT"):
        return "mermer-tezgah"
    if kod.startswith("KPTT"):
        return "polietilen-tezgah"
    return "calisma-tezgahi"


def is_davlumbaz_duvar(kod: str) -> bool:
    return kod.startswith("KDAV") and "OT" not in kod


def is_davlumbaz_orta(kod: str) -> bool:
    return kod.startswith("KDAV") and "OT" in kod


def eticaret_filter(rows: list[dict], kod: str) -> list[dict]:
    if is_davlumbaz_duvar(kod) or is_davlumbaz_orta(kod):
        wanted = set(DAVLUMBAZ_ETIC_VITRIN.get(kod, ()))
        if not wanted:
            return []
        return [r for r in rows if (r["en_cm"], r["derinlik_cm"]) in wanted]
    depth70 = [x for x in rows if x["derinlik_cm"] == ETICARET_DERINLIK_CM]
    if not depth70:
        return []
    min_en = min(x["en_cm"] for x in depth70)
    wanted_en = {min_en, *ETICARET_EN_EXTRA}
    return [x for x in depth70 if x["en_cm"] in wanted_en]


def sku_for(en: int, derinlik: int, kod: str) -> str:
    """Benzersiz stok kodu — EQ.{seri}.{ölçü} örn. EQ.KCEVD01.09070"""
    k = norm_tr(kod)
    return f"{SKU_PREFIX}.{k}.{en:03d}{derinlik:02d}"


def slug_for(en: int, derinlik: int, kod: str) -> str:
    k = slug_dir(norm_tr(kod))
    return f"equsto-{k}-{en:03d}{derinlik:02d}"


def fmt_try(n: float) -> str:
    v = round(n)
    s = f"{v:,}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s


def pricing_tl(satis_eur: float, kur: float) -> dict:
    net = round(satis_eur * kur)
    kdv_dahil = round(net * (1 + KDV))
    return {
        "fiyat_tl_net": net,
        "fiyat_tl": kdv_dahil,
        "price": f"₺{fmt_try(net)},00 + KDV\nKDV Dahil ₺{fmt_try(kdv_dahil)},00",
    }


def img_anchor(img) -> tuple[int, int]:
    anc = getattr(img, "anchor", None)
    if anc is not None and hasattr(anc, "_from"):
        return int(anc._from.row), int(anc._from.col)
    return 0, 0


def img_anchor_row(img) -> int:
    return img_anchor(img)[0]


def product_images(ws) -> list[tuple[int, object]]:
    images = getattr(ws, "_images", []) or []
    anchored = [(img_anchor(img), img) for img in images]
    product = [(r, img) for (r, _), img in anchored if r > 0]
    if product:
        return sorted(product, key=lambda x: x[0])
    return sorted(anchored, key=lambda x: x[0])[-1:]


def save_product_image(img, kod: str, dry_run: bool) -> str | None:
    rel = f"images/catalog/equsto/fiyat-listesi/{slug_dir(kod)}/urun.png"
    if dry_run:
        return rel
    dest = ROOT / "public" / rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(img._data())
    return rel


def extract_product_image(ws, out_dir: Path, kod: str, dry_run: bool) -> str | None:
    product_imgs = product_images(ws)
    if not product_imgs:
        return None
    return save_product_image(product_imgs[0][1], kod, dry_run)


def find_title_sections(ws, sheet_name: str) -> list[tuple[int, str, str]]:
    """Return (title_row, kod, title_text) per product family on dual sheets."""
    kods = sheet_kods(sheet_name)
    if len(kods) < 2:
        return []

    found: list[tuple[int, str]] = []
    for r in range(1, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if not raw:
            continue
        text = norm_tr(str(raw))
        for kod in kods:
            if kod in text:
                found.append((r, kod))
                break

    by_kod: dict[str, tuple[int, str]] = {}
    for r, kod in sorted(found, key=lambda x: x[0]):
        if kod not in by_kod:
            by_kod[kod] = (r, str(ws.cell(r, 1).value or ""))

    sections: list[tuple[int, str, str]] = []
    for kod in kods:
        if kod in by_kod:
            r, raw = by_kod[kod]
            sections.append((r, kod, parse_title(raw, kod)))
    return sections


def image_for_section(
    anchored_imgs: list[tuple[int, object]], start_row: int, end_row: int
) -> object | None:
    in_range = [img for row, img in anchored_imgs if start_row <= row < end_row]
    if in_range:
        return in_range[0]
    after = [(row, img) for row, img in anchored_imgs if row >= start_row]
    return after[0][1] if after else None


def parse_rows(ws, min_row: int = 4, max_row: int | None = None) -> list[dict]:
    rows: list[dict] = []
    end = max_row if max_row is not None else ws.max_row
    for r in range(min_row, end + 1):
        en, derinlik, yuk = ws.cell(r, 2).value, ws.cell(r, 4).value, ws.cell(r, 6).value
        if en is None or derinlik is None or yuk is None:
            continue
        try:
            en, derinlik, yuk = int(en), int(derinlik), int(yuk)
        except (TypeError, ValueError):
            continue
        liste = ws.cell(r, 10).value
        if liste is None:
            continue
        try:
            liste = float(liste)
        except (TypeError, ValueError):
            continue
        taban = ws.cell(r, 30).value
        cekmece = ws.cell(r, 33).value
        try:
            taban = float(taban) if taban is not None else None
        except (TypeError, ValueError):
            taban = None
        try:
            cekmece = float(cekmece) if cekmece is not None else None
        except (TypeError, ValueError):
            cekmece = None
        satis = round(liste * (1 - ISKONTO), 2)
        rows.append(
            {
                "excel_row": r,
                "en_cm": en,
                "derinlik_cm": derinlik,
                "yukseklik_cm": yuk,
                "olcu": f"{en}×{derinlik}×{yuk}",
                "olcu_excel": f"{en} x {derinlik} x {yuk} cm.",
                "olcu_mm": f"{en * 10}×{derinlik * 10}×{yuk * 10}",
                "olcu_etiket": f"{en * 10}×{derinlik * 10}×{yuk * 10} mm",
                "taban_eur": taban,
                "cekmece_eur": cekmece,
                "liste_fiyati_eur": liste,
                "satis_fiyati_eur": satis,
                "iskonto_oran": int(ISKONTO * 100),
            }
        )
    return rows


def build_product(
    row: dict,
    kod: str,
    title: str,
    kur: float,
    image: str | None,
    dept: str,
    category: str,
    eticaret: bool,
) -> dict:
    en, d = row["en_cm"], row["derinlik_cm"]
    sku = sku_for(en, d, kod)
    slug = slug_for(en, d, kod)
    px = pricing_tl(row["satis_fiyati_eur"], kur)
    name = f"{title} {row['olcu_etiket']}"
    teknik = [
        f"Malzeme: {MALZEME}",
        f"Model: {kod}",
        f"Ebat: {row['olcu']} cm ({row['olcu_mm']} mm)",
    ]
    if row.get("taban_eur") is not None:
        teknik.append(f"Taban fiyat (EUR): {row['taban_eur']}")
    if row.get("cekmece_eur") is not None:
        teknik.append(f"Ek modül (EUR): {row['cekmece_eur']}")
    specs = (
        f"{name}\n\n"
        f"Marka: Equsto\n"
        f"Malzeme: {MALZEME}\n\n"
        f"Ürün kodu: {sku}\n"
        f"Model serisi: {kod}\n"
        f"Ebat: {row['olcu']} cm ({row['olcu_mm']} mm)\n\n"
        f"Liste fiyatı (EUR): {row['liste_fiyati_eur']}\n"
        f"Satış fiyatı (EUR, %{int(ISKONTO * 100)} iskonto): {row['satis_fiyati_eur']}\n"
        f"Kur: 1 EUR = {kur} TRY (KDV %{int(KDV * 100)})\n\n"
        f"Kaynak: EQUSTO Fiyat Listesi 2026 — {kod}"
    )
    return {
        "id": f"equsto__{slug}",
        "sku": sku,
        "model": sku,
        "urun_kodu": sku,
        "kod": kod,
        "dept": dept,
        "category": category,
        "brand": "Equsto",
        "name": name,
        "baslik": name,
        "malzeme": MALZEME,
        "en_cm": en,
        "derinlik_cm": d,
        "yukseklik_cm": row["yukseklik_cm"],
        "olcu": row["olcu"],
        "olcu_excel": row.get("olcu_excel", row["olcu"]),
        "olcu_mm": row["olcu_mm"],
        "olcu_etiket": row["olcu_etiket"],
        "olculer": {
            "genislik_mm": en * 10,
            "derinlik_mm": d * 10,
            "yukseklik_mm": row["yukseklik_cm"] * 10,
        },
        "liste_fiyati_eur": row["liste_fiyati_eur"],
        "taban_eur": row.get("taban_eur"),
        "cekmece_eur": row.get("cekmece_eur"),
        "satis_fiyati_eur": row["satis_fiyati_eur"],
        "satis_eur_indirimli": row["satis_fiyati_eur"],
        "iskonto_oran": row["iskonto_oran"],
        "kur_eur_try": kur,
        "fiyat_bekleniyor": False,
        "eticaret": eticaret,
        "pfos": True,
        "images": [image] if image else [],
        "kaynak": KAYNAK,
        "kaynak_dosya": str(XLSX),
        "specs": specs,
        "aciklama": f"{title}. {MALZEME}.",
        "teknik_ozellikler": teknik,
        **px,
    }


def build_family_result(
    kod: str,
    sheet_name: str,
    title: str,
    rows: list[dict],
    image: str | None,
    kur: float,
    dual_sheet: bool,
) -> dict | None:
    if not rows:
        return None
    dept = dept_for_kod(kod)
    category = category_for_kod(kod, dept)
    etic_rows = eticaret_filter(rows, kod)
    etic_keys = {(r["en_cm"], r["derinlik_cm"]) for r in etic_rows}
    pfos_products = [
        build_product(row, kod, title, kur, image, dept, category, (row["en_cm"], row["derinlik_cm"]) in etic_keys)
        for row in rows
    ]
    etic_products = [p for p in pfos_products if p["eticaret"]]
    return {
        "kod": kod,
        "sheet": sheet_name,
        "dual_sheet": dual_sheet,
        "urun_adi": title,
        "dept": dept,
        "category": category,
        "gorsel_urun": image,
        "pfos": {"count": len(pfos_products), "urunler": pfos_products},
        "eticaret": {"count": len(etic_products), "urunler": etic_products},
    }


def process_dual_sheet(ws, sheet_name: str, kur: float, dry_run: bool) -> list[dict]:
    sections = find_title_sections(ws, sheet_name)
    if len(sections) < 2:
        return []

    anchored_imgs = product_images(ws)
    results: list[dict] = []
    for i, (title_row, kod, title) in enumerate(sections):
        end_row = sections[i + 1][0] if i + 1 < len(sections) else ws.max_row + 1
        rows = parse_rows(ws, min_row=title_row, max_row=end_row - 1)
        img_obj = image_for_section(anchored_imgs, title_row, end_row)
        image = save_product_image(img_obj, kod, dry_run) if img_obj else None
        result = build_family_result(kod, sheet_name, title or kod, rows, image, kur, True)
        if result:
            results.append(result)
    return results


def process_sheet(ws, sheet_name: str, kur: float, dry_run: bool) -> list[dict]:
    if is_dual_product_sheet(sheet_name):
        dual = process_dual_sheet(ws, sheet_name, kur, dry_run)
        if dual:
            return dual

    kod = primary_kod(sheet_name)
    title_raw = ws.cell(2, 1).value
    title = parse_title(str(title_raw or ""), kod)
    if not title:
        title = kod
    rows = parse_rows(ws)
    if not rows:
        return []

    image = extract_product_image(ws, OUT_IMG_BASE / slug_dir(kod), kod, dry_run)
    result = build_family_result(kod, sheet_name, title, rows, image, kur, False)
    return [result] if result else []


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sheet", help="Yalnızca bu sekme (ör. KCT02)")
    args = ap.parse_args()

    if not XLSX.exists():
        raise SystemExit(f"Excel bulunamadı: {XLSX}")

    kur, tcmb_date, kur_fallback = fetch_eur_try()
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    sheet_filter = norm_sheet_kod(args.sheet) if args.sheet else None
    index_entries: list[dict] = []
    all_pfos: list[dict] = []
    all_etic: list[dict] = []
    stats = {
        "excel_sheets": 0,
        "catalog_entries": 0,
        "dual_splits": 0,
        "pfos": 0,
        "eticaret": 0,
        "images": 0,
    }

    for sheet_name in wb.sheetnames:
        if sheet_filter:
            sheet_norm = norm_sheet_kod(sheet_name)
            kods = sheet_kods(sheet_name)
            if sheet_norm != sheet_filter and sheet_filter not in kods and primary_kod(sheet_name) != sheet_filter:
                continue
        ws = wb[sheet_name]
        results = process_sheet(ws, sheet_name, kur, args.dry_run)
        if not results:
            print(f"[atla] {sheet_name}: ölçü yok")
            continue

        stats["excel_sheets"] += 1
        if len(results) > 1:
            stats["dual_splits"] += len(results) - 1

        for result in results:
            stats["catalog_entries"] += 1
            stats["pfos"] += result["pfos"]["count"]
            stats["eticaret"] += result["eticaret"]["count"]
            if result["gorsel_urun"]:
                stats["images"] += 1

            all_pfos.extend(result["pfos"]["urunler"])
            all_etic.extend(result["eticaret"]["urunler"])

            index_entries.append(
                {
                    "kod": result["kod"],
                    "sheet": result["sheet"],
                    "dual_sheet": result["dual_sheet"],
                    "urun_adi": result["urun_adi"],
                    "dept": result["dept"],
                    "category": result["category"],
                    "pfos_count": result["pfos"]["count"],
                    "eticaret_count": result["eticaret"]["count"],
                    "gorsel": result["gorsel_urun"],
                    "slug": slug_dir(result["kod"]),
                }
            )

            if args.dry_run:
                tag = " [dual]" if result["dual_sheet"] else ""
                print(
                    f"{result['kod']:14} PFOS={result['pfos']['count']:3} "
                    f"ETC={result['eticaret']['count']:2} dept={result['dept']}{tag}"
                )
                continue

            out_dir = OUT_BASE / slug_dir(result["kod"])
            out_dir.mkdir(parents=True, exist_ok=True)
            catalog = {
                "liste": "EQUSTO Fiyat Listesi 2026",
                "kod": result["kod"],
                "sheet": result["sheet"],
                "dual_sheet": result["dual_sheet"],
                "urun_adi": result["urun_adi"],
                "dept": result["dept"],
                "category": result["category"],
                "malzeme": MALZEME,
                "marka": "Equsto",
                "iskonto_oran": int(ISKONTO * 100),
                "kur_eur_try": kur,
                "kur_tcmb_tarih": tcmb_date,
                "kur_fallback": kur_fallback,
                "kaynak_dosya": str(XLSX),
                "imported_at": datetime.now(timezone.utc).isoformat(),
                "gorsel_urun": result["gorsel_urun"],
                "pfos": result["pfos"],
                "eticaret": result["eticaret"],
            }
            (out_dir / "catalog.json").write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
            (out_dir / "pfos-urunler.json").write_text(
                json.dumps(result["pfos"]["urunler"], ensure_ascii=False, indent=2), encoding="utf-8"
            )
            (out_dir / "eticaret-urunler.json").write_text(
                json.dumps(result["eticaret"]["urunler"], ensure_ascii=False, indent=2), encoding="utf-8"
            )

    if args.dry_run:
        print(json.dumps(stats, indent=2))
        return

    OUT_BASE.mkdir(parents=True, exist_ok=True)
    master = {
        "liste": "EQUSTO Fiyat Listesi 2026",
        "malzeme": MALZEME,
        "marka": "Equsto",
        "kaynak": KAYNAK,
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "kur_eur_try": kur,
        "sheet_count": stats["catalog_entries"],
        "excel_sheet_count": stats["excel_sheets"],
        "dual_split_count": stats["dual_splits"],
        "pfos_count": stats["pfos"],
        "eticaret_count": stats["eticaret"],
        "sheets": index_entries,
    }
    (OUT_BASE / "index.json").write_text(json.dumps(master, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_BASE / "pfos-tum-urunler.json").write_text(
        json.dumps(all_pfos, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_BASE / "eticaret-tum-urunler.json").write_text(
        json.dumps(all_etic, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(
        f"Tamam: {stats['excel_sheets']} Excel sekmesi -> {stats['catalog_entries']} katalog "
        f"({stats['dual_splits']} cift urun ayrimi), {stats['pfos']} PFOS, "
        f"{stats['eticaret']} e-ticaret, {stats['images']} gorsel"
    )
    print(f"Index: {OUT_BASE / 'index.json'}")


if __name__ == "__main__":
    main()
