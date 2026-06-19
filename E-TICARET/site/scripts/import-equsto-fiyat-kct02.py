#!/usr/bin/env python3
"""
EQUSTO Fiyat Listesi — ilk sekme KÇT02 → PFOS + e-ticaret JSON + görseller

Kaynak: c:\\D Disk\\FİYAT LİSTELERİ\\EQUSTO - FİYAT LİSTESİ.xlsx

  python scripts/import-equsto-fiyat-kct02.py
  python scripts/import-equsto-fiyat-kct02.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(r"c:\D Disk\FİYAT LİSTELERİ\EQUSTO - FİYAT LİSTESİ.xlsx")
OUT_DATA = ROOT / "public/data/fiyat-listeleri/equsto/2026-fiyat-listesi/kct02"
OUT_IMG = ROOT / "public/images/catalog/equsto/kct02"
PREVIEW_HTML = ROOT / "public/equsto-fiyat-kct02-preview.html"

SHEET_INDEX = 0
KOD = "KCT02"
KATEGORI = "calisma-tezgahi"
ISKONTO = 0.10
KDV = 0.20
ETICARET_DERINLIK_CM = 70
ETICARET_EN_EXTRA = (120, 140, 160, 190)
FAMILY_SUFFIX = "02"  # KÇT02


def fetch_eur_try() -> tuple[float, str, bool]:
    url = "https://www.tcmb.gov.tr/kurlar/today.xml"
    try:
        with urllib.request.urlopen(url, timeout=15) as res:
            xml = res.read()
        root = ET.fromstring(xml)
        tarih = root.attrib.get("Tarih", "")
        for cur in root.findall("Currency"):
            if cur.attrib.get("Kod") == "EUR":
                selling = cur.findtext("BanknoteSelling")
                rate = float(selling)
                if rate > 0:
                    return rate, tarih, False
    except Exception as e:
        print(f"[kur] fallback: {e}")
    return 53.05, "", True


def norm_sheet_kod(name: str) -> str:
    s = (name or "").upper()
    s = s.replace("İ", "I").replace("Ç", "C").replace("Ş", "S").replace("Ğ", "G").replace("Ö", "O").replace("Ü", "U")
    return re.sub(r"[^A-Z0-9]", "", s)


def parse_title(raw: str) -> str:
    if not raw:
        return ""
    t = re.sub(r"\s+", " ", str(raw).strip())
    t = re.sub(r"\s*[-–]?\s*K[ÇC]T\s*0?2\s*$", "", t, flags=re.I)
    return t.strip(" -")


def sku_for(en: int, derinlik: int) -> str:
    return f"EQUSTO.{en:03d}{derinlik:02d}.{FAMILY_SUFFIX}"


def slug_for(en: int, derinlik: int) -> str:
    return f"equsto-{en:03d}{derinlik:02d}-{FAMILY_SUFFIX.lower()}"


def fmt_try(n: float) -> str:
    v = round(n)
    s = f"{v:,}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s


def pricing_tl(satis_eur: float, kur: float) -> dict:
    net = round(satis_eur * kur)
    kdv_dahil = round(net * (1 + KDV))
    return {
        "fiyat_tl_net": net,
        "fiyat_tl": kdv_dahil,
        "price": f"₺{fmt_try(net)},00 + KDV\nKDV Dahil ₺{fmt_try(kdv_dahil)},00",
    }


def img_anchor_row(img) -> int:
    anc = getattr(img, "anchor", None)
    if anc is not None and hasattr(anc, "_from"):
        return int(anc._from.row)
    return 0


def is_product_image(img) -> bool:
    """Excel'de ürün çizimi satır 38 civarında; satır 0'daki JPEG'ler alakasız gömülü dosyalar."""
    return img_anchor_row(img) >= 20


def extract_images(ws, dry_run: bool) -> tuple[list[str], list[str]]:
    """(urun_gorselleri, atilan_gorseller)"""
    images = getattr(ws, "_images", []) or []
    product_paths: list[str] = []
    skipped: list[str] = []

    if not dry_run:
        OUT_IMG.mkdir(parents=True, exist_ok=True)
        for old in OUT_IMG.glob("kct02-sheet-*"):
            old.unlink(missing_ok=True)

    product_idx = 0
    for i, img in enumerate(images):
        if not is_product_image(img):
            skipped.append(f"embedded#{i + 1}:row{img_anchor_row(img)}")
            continue
        product_idx += 1
        ext = ".png"
        if hasattr(img, "format") and img.format:
            ext = f".{str(img.format).lower()}"
        fname = "kct02-urun.png" if product_idx == 1 else f"kct02-urun-{product_idx:02d}{ext}"
        rel = f"images/catalog/equsto/kct02/{fname}"
        product_paths.append(rel)
        if dry_run:
            continue
        dest = ROOT / "public" / rel
        dest.write_bytes(img._data())

    return product_paths, skipped


def parse_rows(ws) -> tuple[str, list[dict]]:
    title_raw = ws.cell(2, 1).value
    title = parse_title(str(title_raw or ""))
    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):
        en, derinlik, yuk = ws.cell(r, 2).value, ws.cell(r, 4).value, ws.cell(r, 6).value
        if en is None or derinlik is None or yuk is None:
            continue
        try:
            en, derinlik, yuk = int(en), int(derinlik), int(yuk)
        except (TypeError, ValueError):
            continue
        liste = ws.cell(r, 10).value
        if liste is None:
            continue
        try:
            liste = float(liste)
        except (TypeError, ValueError):
            continue
        taban = ws.cell(r, 30).value  # AD
        cekmece = ws.cell(r, 33).value  # AG
        try:
            taban = float(taban) if taban is not None else None
        except (TypeError, ValueError):
            taban = None
        try:
            cekmece = float(cekmece) if cekmece is not None else None
        except (TypeError, ValueError):
            cekmece = None
        satis = round(liste * (1 - ISKONTO), 2)
        rows.append(
            {
                "en_cm": en,
                "derinlik_cm": derinlik,
                "yukseklik_cm": yuk,
                "olcu": f"{en}×{derinlik}×{yuk}",
                "olcu_excel": f"{en} x {derinlik} x {yuk} cm.",
                "olcu_mm": f"{en * 10}×{derinlik * 10}×{yuk * 10}",
                "olcu_etiket": f"{en * 10}×{derinlik * 10}×{yuk * 10} mm",
                "taban_eur": taban,
                "cekmece_eur": cekmece,
                "liste_fiyati_eur": liste,
                "satis_fiyati_eur": satis,
                "iskonto_oran": int(ISKONTO * 100),
            }
        )
    return title, rows


def eticaret_filter(rows: list[dict]) -> list[dict]:
    depth70 = [x for x in rows if x["derinlik_cm"] == ETICARET_DERINLIK_CM]
    if not depth70:
        return []
    min_en = min(x["en_cm"] for x in depth70)
    wanted = {min_en, *ETICARET_EN_EXTRA}
    return [x for x in depth70 if x["en_cm"] in wanted]


def display_name(row: dict) -> str:
    return f"Çalışma Tezgahı KÇT02 — Tek Çekmeceli {row['olcu_etiket']}"


def build_product(row: dict, title: str, kur: float, images: list[str], eticaret: bool) -> dict:
    en, d = row["en_cm"], row["derinlik_cm"]
    sku = sku_for(en, d)
    slug = slug_for(en, d)
    px = pricing_tl(row["satis_fiyati_eur"], kur)
    name = display_name(row)
    img_list = images[:1] if images else []
    return {
        "id": f"equsto__{slug}",
        "sku": sku,
        "model": sku,
        "urun_kodu": sku,
        "kod": KOD,
        "dept": "tezgah",
        "category": KATEGORI,
        "brand": "Equsto",
        "name": name,
        "baslik": name,
        "en_cm": en,
        "derinlik_cm": d,
        "yukseklik_cm": row["yukseklik_cm"],
        "olcu": row["olcu"],
        "olcu_excel": row.get("olcu_excel", row["olcu"]),
        "olcu_mm": row["olcu_mm"],
        "olcu_etiket": row["olcu_etiket"],
        "olculer": {
            "genislik_mm": en * 10,
            "derinlik_mm": d * 10,
            "yukseklik_mm": row["yukseklik_cm"] * 10,
        },
        "liste_fiyati_eur": row["liste_fiyati_eur"],
        "taban_eur": row.get("taban_eur"),
        "cekmece_eur": row.get("cekmece_eur"),
        "satis_fiyati_eur": row["satis_fiyati_eur"],
        "satis_eur_indirimli": row["satis_fiyati_eur"],
        "iskonto_oran": row["iskonto_oran"],
        "kur_eur_try": kur,
        "fiyat_bekleniyor": False,
        "eticaret": eticaret,
        "pfos": True,
        "images": img_list,
        "kaynak": "equsto-fiyat-listesi-2026",
        "kaynak_dosya": str(XLSX),
        "specs": (
            f"{name}\n\n"
            f"Ürün kodu: {sku}\n"
            f"Model: {KOD}\n"
            f"Ebat: {row['olcu']} cm ({row['olcu_mm']} mm)\n\n"
            f"Liste fiyatı (EUR): {row['liste_fiyati_eur']}\n"
            f"Satış fiyatı (EUR, %{int(ISKONTO * 100)} iskonto): {row['satis_fiyati_eur']}\n"
            f"Kur: 1 EUR = {kur} TRY (KDV %{int(KDV * 100)})\n\n"
            f"Kaynak: EQUSTO Fiyat Listesi 2026 — {KOD}\n"
            f"Marka: Equsto"
        ),
        **px,
    }


def write_preview_html() -> None:
    html = """<!DOCTYPE html>
<html lang="tr">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>KÇT02 — EQUSTO Fiyat Listesi</title>
  <style>
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      background: #c8c8c8;
      font-family: Calibri, "Segoe UI", Arial, sans-serif;
      font-size: 11pt;
      color: #000;
    }
    .wrap { max-width: 1100px; margin: 24px auto; padding: 0 16px 40px; }
    .sheet {
      background: #fff;
      border: 1px solid #999;
      box-shadow: 0 2px 12px rgba(0,0,0,.15);
      padding: 18px 20px 24px;
    }
    .sheet-title {
      font-size: 13pt;
      font-weight: 700;
      margin: 0 0 16px;
      letter-spacing: .2px;
    }
    .sheet-kod {
      display: inline-block;
      margin-left: 8px;
      font-size: 10pt;
      font-weight: 600;
      color: #444;
    }
    .layout {
      display: grid;
      grid-template-columns: minmax(280px, 360px) 1fr;
      gap: 20px;
      align-items: start;
    }
    @media (max-width: 800px) {
      .layout { grid-template-columns: 1fr; }
    }
    table.excel {
      width: 100%;
      border-collapse: collapse;
      font-size: 11pt;
    }
    table.excel td {
      border: 1px solid #d0d0d0;
      padding: 3px 8px;
      white-space: nowrap;
    }
    table.excel tr.eticaret td { background: #e2efda; }
    table.excel tr.head td {
      background: #f2f2f2;
      font-weight: 700;
      border-color: #bfbfbf;
    }
    table.excel td.num { text-align: right; font-variant-numeric: tabular-nums; }
    table.excel td.x { text-align: center; width: 18px; border-left: none; border-right: none; padding: 3px 2px; }
    table.excel td.unit { border-left: none; color: #333; }
    .drawing {
      border: 1px solid #d0d0d0;
      background: #fafafa;
      min-height: 280px;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 12px;
    }
    .drawing img { max-width: 100%; max-height: 520px; object-fit: contain; }
    .legend {
      margin-top: 14px;
      font-size: 10pt;
      color: #444;
      display: flex;
      flex-wrap: wrap;
      gap: 16px;
    }
    .swatch { display: inline-block; width: 14px; height: 14px; background: #e2efda; border: 1px solid #a9d18e; vertical-align: middle; margin-right: 6px; }
    .meta { margin-top: 10px; font-size: 10pt; color: #666; }
    .eticaret-box {
      margin-top: 20px;
      border-top: 2px solid #e0e0e0;
      padding-top: 16px;
    }
    .eticaret-box h2 { font-size: 12pt; margin: 0 0 10px; }
    .ec-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 10px; }
    .ec-card {
      border: 1px solid #a9d18e;
      background: #f6fff3;
      padding: 10px 12px;
      font-size: 10pt;
    }
    .ec-card b { display: block; font-size: 11pt; margin-bottom: 4px; }
    .loading, .err { text-align: center; padding: 48px; color: #666; }
    .err { color: #b00020; }
    .hidden { display: none; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="sheet" id="sheet" hidden>
      <h1 class="sheet-title"><span id="title"></span><span class="sheet-kod" id="kod"></span></h1>
      <div class="layout">
        <div>
          <table class="excel">
            <tbody id="tbody"></tbody>
          </table>
          <div class="legend">
            <span><span class="swatch"></span> E-ticaret (derinlik 70 cm)</span>
            <span id="price-note"></span>
          </div>
        </div>
        <div class="drawing">
          <img id="drawing" alt="KÇT02 ürün çizimi"/>
        </div>
      </div>
      <div class="eticaret-box">
        <h2>E-ticaret seçimi (4 ölçü)</h2>
        <div class="ec-grid" id="ec-grid"></div>
      </div>
      <div class="meta" id="meta"></div>
    </div>
    <div class="loading" id="loading">Excel sayfası yükleniyor…</div>
    <div class="err hidden" id="err"></div>
  </div>
  <script>
    const CATALOG = "/data/fiyat-listeleri/equsto/2026-fiyat-listesi/kct02/catalog.json";
    fetch(CATALOG).then(r => {
      if (!r.ok) throw new Error(r.status + " " + r.statusText);
      return r.json();
    }).then(cat => {
      document.getElementById("loading").classList.add("hidden");
      document.getElementById("sheet").hidden = false;
      document.getElementById("title").textContent = cat.urun_adi;
      document.getElementById("kod").textContent = "-" + cat.sheet;
      const img = cat.gorsel_urun || (cat.gorseller && cat.gorseller[0]);
      if (img) document.getElementById("drawing").src = "/" + img;
      document.getElementById("price-note").textContent =
        "Liste € = taban + çekmece (" + (cat.cekmece_eur || 15) + " €)";
      document.getElementById("meta").textContent =
        "Kaynak: EQUSTO Fiyat Listesi | Kur: 1 EUR = " + cat.kur_eur_try + " TRY | Satış = Liste × 0,90";
      const tb = document.getElementById("tbody");
      const head = document.createElement("tr");
      head.className = "head";
      head.innerHTML = "<td>En</td><td></td><td>Der.</td><td></td><td>Yük.</td><td></td><td class=\\"num\\">Liste €</td><td class=\\"num\\">Satış €</td>";
      tb.appendChild(head);
      cat.pfos.urunler.forEach(p => {
        const tr = document.createElement("tr");
        if (p.eticaret) tr.classList.add("eticaret");
        const parts = p.olcu_excel ? p.olcu_excel.replace(" cm.", "").split(" x ") : p.olcu.replace(/×/g, " x ").split(" x ");
        tr.innerHTML =
          "<td class=\\"num\\">" + parts[0] + "</td><td class=\\"x\\">x</td>" +
          "<td class=\\"num\\">" + parts[1] + "</td><td class=\\"x\\">x</td>" +
          "<td class=\\"num\\">" + parts[2] + "</td><td class=\\"unit\\">cm.</td>" +
          "<td class=\\"num\\">" + p.liste_fiyati_eur + "</td>" +
          "<td class=\\"num\\">" + p.satis_fiyati_eur + "</td>";
        tb.appendChild(tr);
      });
      const eg = document.getElementById("ec-grid");
      cat.eticaret.urunler.forEach(p => {
        const d = document.createElement("div");
        d.className = "ec-card";
        d.innerHTML = "<b>" + p.olcu + " cm</b>Liste: " + p.liste_fiyati_eur + " €<br>Satış: " + p.satis_fiyati_eur + " €<br>₺" + (p.fiyat_tl_net||0).toLocaleString("tr-TR") + " + KDV";
        eg.appendChild(d);
      });
    }).catch(e => {
      document.getElementById("loading").classList.add("hidden");
      const el = document.getElementById("err");
      el.textContent = "Hata: " + e.message + " — npm run dev (port 3099)";
      el.classList.remove("hidden");
    });
  </script>
</body>
</html>
"""
    PREVIEW_HTML.write_text(html, encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not XLSX.exists():
        raise SystemExit(f"Excel bulunamadı: {XLSX}")

    kur, tcmb_date, kur_fallback = fetch_eur_try()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sheet_name = wb.sheetnames[SHEET_INDEX]
    ws = wb[sheet_name]
    kod_sheet = norm_sheet_kod(sheet_name)
    if kod_sheet != KOD:
        print(f"[uyarı] sekme adı {sheet_name!r} → {kod_sheet}, beklenen {KOD}")

    title, rows = parse_rows(ws)
    if not rows:
        raise SystemExit("Ölçü satırı bulunamadı")

    images, skipped_images = extract_images(ws, args.dry_run)
    etic_rows = eticaret_filter(rows)
    etic_keys = {(r["en_cm"], r["derinlik_cm"]) for r in etic_rows}
    cekmece_eur = next((r["cekmece_eur"] for r in rows if r.get("cekmece_eur")), 15)

    pfos_products = []
    for row in rows:
        key = (row["en_cm"], row["derinlik_cm"])
        pfos_products.append(build_product(row, title, kur, images, key in etic_keys))

    etic_products = [p for p in pfos_products if p["eticaret"]]

    catalog = {
        "liste": "EQUSTO Fiyat Listesi 2026",
        "kod": KOD,
        "sheet": sheet_name,
        "urun_adi": title,
        "kategori": KATEGORI,
        "iskonto_oran": int(ISKONTO * 100),
        "cekmece_eur": cekmece_eur,
        "kur_eur_try": kur,
        "kur_tcmb_tarih": tcmb_date,
        "kur_fallback": kur_fallback,
        "kaynak_dosya": str(XLSX),
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "gorsel_urun": images[0] if images else None,
        "gorseller": images,
        "gorseller_atildi": skipped_images,
        "eticaret_filter": {
            "derinlik_cm": ETICARET_DERINLIK_CM,
            "en_cm": sorted({r["en_cm"] for r in etic_rows}),
        },
        "pfos": {"count": len(pfos_products), "urunler": pfos_products},
        "eticaret": {"count": len(etic_products), "urunler": etic_products},
    }

    if args.dry_run:
        print(json.dumps({
            "sheet": sheet_name,
            "title": title,
            "rows": len(rows),
            "eticaret": len(etic_products),
            "images": len(images),
            "skipped_images": skipped_images,
            "kur": kur,
        }, ensure_ascii=False, indent=2))
        return

    OUT_DATA.mkdir(parents=True, exist_ok=True)
    (OUT_DATA / "catalog.json").write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_DATA / "pfos-urunler.json").write_text(
        json.dumps(pfos_products, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_DATA / "eticaret-urunler.json").write_text(
        json.dumps(etic_products, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_preview_html()

    print(f"KÇT02: {len(pfos_products)} PFOS, {len(etic_products)} e-ticaret, {len(images)} ürün görseli")
    if skipped_images:
        print(f"  atılan gömülü görsel: {len(skipped_images)} (Excel artığı)")
    print(f"JSON: {OUT_DATA}")
    print(f"Önizleme: http://localhost:3099/equsto-fiyat-kct02-preview.html")


if __name__ == "__main__":
    main()
