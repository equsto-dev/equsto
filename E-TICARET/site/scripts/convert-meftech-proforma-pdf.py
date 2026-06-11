# -*- coding: utf-8 -*-
"""MEFTECH proforma PDF (Grup/Poz satırlı) → Excel."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROW_START = re.compile(r"^\d{2}$")
POZ = re.compile(r"^\d{3}$")
NUM = re.compile(r"^[\d.,]+$")
DIM_PART = re.compile(r"^[\d.,]+$|^[Xx]$")


def extract_text_pdf_parse(pdf_path: Path) -> str:
    site = Path(__file__).resolve().parent.parent
    js = f"""
const fs=require('fs');
const pdf=require('pdf-parse/lib/pdf-parse.js');
pdf(fs.readFileSync({json.dumps(str(pdf_path))})).then(r=>process.stdout.write(r.text||''));
"""
    return subprocess.check_output(
        ["node", "-e", js],
        cwd=str(site),
        text=True,
        encoding="utf-8",
        errors="replace",
    )


def is_noise(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    if s in {"EUR", "X", "x"}:
        return True
    if s.startswith("-") and len(s) > 3:
        return True
    if s.startswith("■"):
        return True
    if "Proforma" in s or s.startswith("Tari"):
        return True
    if "YETKİLİ" in s or "MEFTECH" in s or "ZEKER" in s:
        return True
    if "Grup" in s and "Poz" in s:
        return True
    if s in {"MUTFAK", "ASICAK MUTFAK", "SICAK MUTFAK"}:
        return True
    if re.match(r"^[A-ZİĞÜŞÖÇ]$", s) and len(s) == 1:
        return False
    return False


def parse_rows(text: str) -> tuple[str, str, list[dict]]:
    lines = [ln.strip() for ln in text.splitlines()]
    proforma_no = ""
    isin_adi = ""
    for ln in lines:
        if re.search(r"5000114466|\d{10}", ln):
            m = re.search(r"(\d{10,})", ln)
            if m and not proforma_no:
                proforma_no = m.group(1)
        if "SPOR KULUB" in ln.upper() or "KULÜB" in ln:
            isin_adi = ln

    rows: list[dict] = []
    i = 0
    current_section = ""

    while i < len(lines):
        ln = lines[i]
        if ln in {"SICAK MUTFAK", "ASICAK MUTFAK", "SOĞUK DEPO", "BULAŞIK YIKAMA"} or (
            len(ln) > 4 and ln.isupper() and "MUTFAK" in ln
        ):
            current_section = ln.replace("A", "", 1).strip() if ln.startswith("A") else ln

        if not ROW_START.match(ln):
            i += 1
            continue

        if i + 2 >= len(lines):
            break
        grup = ln
        bolum = lines[i + 1].strip()
        poz = lines[i + 2].strip()
        if not re.match(r"^[A-Z]$", bolum) or not POZ.match(poz):
            i += 1
            continue

        i += 3
        desc_parts: list[str] = []
        stok_parts: list[str] = []

        while i < len(lines):
            cur = lines[i].strip()
            if ROW_START.match(cur):
                break
            if NUM.match(cur.replace(" ", "")) and i + 4 < len(lines):
                nxt = [lines[i + j].strip() for j in range(1, 5)]
                if len(nxt) >= 3 and nxt[0].upper() == "X" and nxt[2].upper() == "X":
                    boy, en, yuk = cur, nxt[1], nxt[4] if len(nxt) > 4 else nxt[3]
                    i += 5
                    adet = lines[i].strip() if i < len(lines) else "1"
                    i += 1
                    prices: list[str] = []
                    while i < len(lines) and lines[i].strip() != "EUR":
                        if NUM.match(lines[i].strip().replace(" ", "")) or re.search(r"[\d.,]", lines[i]):
                            prices.append(lines[i].strip())
                        i += 1
                    if i < len(lines) and lines[i].strip() == "EUR":
                        i += 1
                    tanim = " ".join(desc_parts).strip()
                    stok = " ".join(stok_parts).strip()
                    satis = prices[0] if prices else ""
                    toplam = prices[1] if len(prices) > 1 else ""
                    rows.append(
                        {
                            "Bölüm": current_section,
                            "Grup": grup,
                            "Poz": f"{bolum} {poz}",
                            "Stok/Kaynak": stok,
                            "Tanım": tanim,
                            "Boy": boy,
                            "En": en,
                            "Yük": yuk,
                            "Adet": adet,
                            "Satış": satis,
                            "Toplam Satış": toplam,
                            "Döviz": "EUR",
                        }
                    )
                    break
            if re.match(r"^[A-Z]{2,4}-$", cur) or re.match(r"^[A-Z]{2,4}$", cur):
                stok_parts.append(cur)
            elif cur and not is_noise(cur):
                desc_parts.append(cur)
            i += 1
        else:
            continue

    return proforma_no, isin_adi, rows


def write_excel(out_path: Path, proforma_no: str, isin_adi: str, rows: list[dict]) -> Path:
    headers = [
        "Bölüm",
        "Grup",
        "Poz",
        "Stok/Kaynak",
        "Tanım",
        "Boy",
        "En",
        "Yük",
        "Adet",
        "Satış",
        "Toplam Satış",
        "Döviz",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = f"Proforma {proforma_no} — {isin_adi}".strip(" —")
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = title_fill
    t.alignment = Alignment(horizontal="center")

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill

    for ri, row in enumerate(rows, 3):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=row.get(h, ""))
            ws.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical="top")

    for ci in range(1, len(headers) + 1):
        max_len = len(headers[ci - 1])
        for row in ws.iter_rows(min_row=3, min_col=ci, max_col=ci):
            v = row[0].value
            if v:
                max_len = max(max_len, min(len(str(v)), 80))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A3"
    wb.save(out_path)
    return out_path


def main() -> None:
    pdf = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"c:\Users\adema\Downloads\2026-007-1.pdf")
    if not pdf.exists():
        print(f"PDF bulunamadi: {pdf}")
        sys.exit(1)

    text = extract_text_pdf_parse(pdf)
    proforma_no, isin_adi, rows = parse_rows(text)
    out = pdf.with_suffix(".xlsx")
    write_excel(out, proforma_no, isin_adi, rows)
    print(f"OK: {out}")
    print(f"Proforma: {proforma_no} | Isin: {isin_adi} | Satir: {len(rows)}")


if __name__ == "__main__":
    main()
