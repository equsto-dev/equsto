# -*- coding: utf-8 -*-
"""İnoksan F-220 proforma PDF → Excel (orijinal kolon yapısı)."""
from __future__ import annotations

import glob
import re
import sys
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Böl.",
    "Grup",
    "Poz",
    "EK",
    "Stok no",
    "Tanımı",
    "Kaynak",
    "Boy",
    "En",
    "Yük.",
    "Adet",
    "Satış",
    "Toplam Satış",
    "Döviz",
]

NOISE = re.compile(
    r"^(Form No|Tarih|Proforma No|İşin Adı|PROFORMA FATURA|Böl\.|Grup|Poz|"
    r"EK|Stok no|Tanımı|Kaynak|Boy|En|Yük\.|Adet|Satış|Toplam Satış|Döviz|\d+)$",
    re.I,
)
ROW_START = re.compile(r"^(\d{2})$")
GRUP = re.compile(r"^[A-Z]$")
POZ = re.compile(r"^(\d{2})$")
STOK = re.compile(r"^[A-Z]{2,4}-[A-Z0-9][A-Z0-9\-\./]*$", re.I)
PRICE = re.compile(r"^[\d.,]+$")
DIM_LINE = re.compile(
    r"^(?P<kaynak>[^\d]+?)\s+(?P<boy>[\d.,]+)\s+X\s+(?P<en>[\d.,]+)\s+X\s+(?P<yuk>[\d.,]+)\.?\s*$",
    re.I,
)
DIM_ONLY = re.compile(
    r"^\s*(?P<boy>[\d.,]+)\s+X\s+(?P<en>[\d.,]+)\s+X\s+(?P<yuk>[\d.,]+)\.?\s*$",
    re.I,
)
SECTION = re.compile(r"^(\d{2})\.\s*(.+)$")
GROUP = re.compile(r"^([A-Z])\.\s*(.+)$", re.I)


def parse_num(s: str) -> float | None:
    t = str(s).strip().replace(" ", "")
    if not t or not PRICE.match(t):
        return None
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    else:
        t = t.replace(",", "")
    try:
        return float(t)
    except ValueError:
        return None


def extract_pdf_text(pdf_path: Path) -> str:
    doc = fitz.open(str(pdf_path))
    parts = []
    for page in doc:
        parts.append(page.get_text())
    return "\n".join(parts)


def parse_meta(lines: list[str]) -> dict:
    meta = {
        "form_no": "F-220 (D:01.2018)",
        "tarih": "",
        "proforma_no": "",
        "isin_adi": "",
        "section": "",
        "group_label": "",
    }
    for i, ln in enumerate(lines):
        s = ln.strip()
        if s.startswith("Form No"):
            meta["form_no"] = s.split(":", 1)[-1].strip() or meta["form_no"]
        elif s == "Tarih" and i + 2 < len(lines):
            meta["tarih"] = lines[i + 2].strip()
        elif s == "Proforma No" and i + 2 < len(lines):
            meta["proforma_no"] = lines[i + 2].strip()
        elif s == "İşin Adı" and i + 2 < len(lines):
            meta["isin_adi"] = lines[i + 2].strip()
        m = SECTION.match(s)
        if m and "MUTFAK" in m.group(2).upper():
            meta["section"] = f"{m.group(1)}. {m.group(2).strip()}"
        m = GROUP.match(s)
        if m:
            meta["group_label"] = f"{m.group(1).upper()}. {m.group(2).strip()}"
    return meta


def is_row_start(i: int, lines: list[str]) -> bool:
    if i + 2 >= len(lines):
        return False
    bol, grup, poz = lines[i].strip(), lines[i + 1].strip(), lines[i + 2].strip()
    return bool(ROW_START.match(bol) and GRUP.match(grup) and POZ.match(poz))


def parse_items(text: str) -> tuple[dict, list[dict]]:
    raw_lines = [ln.rstrip() for ln in text.splitlines()]
    lines = [ln.strip() for ln in raw_lines if ln.strip()]
    meta = parse_meta(lines)

    items: list[dict] = []
    i = 0
    while i < len(lines):
        if not is_row_start(i, lines):
            i += 1
            continue

        bol = lines[i]
        grup = lines[i + 1]
        poz = lines[i + 2]
        i += 3

        if i >= len(lines) or not STOK.match(lines[i]):
            continue
        stok = lines[i]
        i += 1

        title_parts: list[str] = []
        specs: list[str] = []
        kaynak = boy = en = yuk = ""
        adet = 1
        satis = toplam = None
        doviz = "EUR"
        phase = "title"

        while i < len(lines):
            cur = lines[i]
            if cur.upper().startswith("TOPLAM") or cur.upper().startswith("GENEL TOPLAM"):
                break
            if is_row_start(i, lines):
                break
            if NOISE.match(cur) or cur.startswith("Form No") or cur in {"Tarih", "Proforma No", "İşin Adı"}:
                i += 1
                continue
            if cur == "PROFORMA FATURA":
                i += 1
                continue
            if cur == ":" or cur == meta.get("isin_adi", ""):
                i += 1
                continue
            if SECTION.match(cur) or GROUP.match(cur):
                i += 1
                continue
            if cur.startswith("*"):
                break

            dm = DIM_LINE.match(cur)
            if dm and phase in {"title", "spec"}:
                kaynak = dm.group("kaynak").strip()
                boy = dm.group("boy").strip()
                en = dm.group("en").strip()
                yuk = dm.group("yuk").strip()
                phase = "dims"
                i += 1
                continue

            dm2 = DIM_ONLY.match(cur)
            if dm2 and phase in {"title", "spec"}:
                kaynak = kaynak or "İnoksan"
                boy = dm2.group("boy").strip()
                en = dm2.group("en").strip()
                yuk = dm2.group("yuk").strip()
                phase = "dims"
                i += 1
                continue

            if phase == "dims" and PRICE.match(cur.replace(" ", "")):
                if adet == 1 and re.fullmatch(r"\d{1,4}", cur):
                    adet = int(cur)
                    i += 1
                    continue
                val = parse_num(cur)
                if val is not None:
                    if satis is None:
                        satis = val
                    elif toplam is None:
                        toplam = val
                    phase = "money"
                i += 1
                continue

            if phase == "money" and cur.upper() == "EUR":
                phase = "spec"
                i += 1
                continue

            if phase == "money" and PRICE.match(cur.replace(" ", "")):
                val = parse_num(cur)
                if val is not None and toplam is None:
                    toplam = val
                i += 1
                continue

            if phase == "title":
                title_parts.append(cur)
                phase = "title"
            else:
                if cur.lower() in {"opsiyonlar:", "opsiyonlar:,"}:
                    i += 1
                    continue
                specs.append(cur)
            i += 1

        tanim_title = " ".join(title_parts).strip()
        tanim_spec = "\n".join(specs).strip()
        tanim = tanim_title
        if tanim_spec:
            tanim = f"{tanim_title}\n{tanim_spec}" if tanim_title else tanim_spec

        if toplam is None and satis is not None:
            toplam = round(satis * adet, 2)

        items.append(
            {
                "bol": bol,
                "grup": grup,
                "poz": poz,
                "ek": "",
                "stok": stok,
                "tanim": tanim,
                "kaynak": kaynak or "İnoksan",
                "boy": boy,
                "en": en,
                "yuk": yuk,
                "adet": adet,
                "satis": satis,
                "toplam": toplam,
                "doviz": doviz,
            }
        )

    return meta, items


def parse_terms(text: str) -> list[str]:
    seen: set[str] = set()
    terms: list[str] = []
    for ln in text.splitlines():
        s = ln.strip()
        if s.startswith("*"):
            t = s.lstrip("* ").strip()
            if t and t not in seen:
                seen.add(t)
                terms.append(t)
    return terms


def write_excel(out_path: Path, meta: dict, items: list[dict], terms: list[str]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "PROFORMA"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    sec_fill = PatternFill("solid", fgColor="E2EFDA")
    title_fill = PatternFill("solid", fgColor="1F4E79")

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"Form No: {meta.get('form_no', 'F-220 (D:01.2018)')}"
    c.font = Font(size=10)

    ws["A2"], ws["B2"] = "Tarih:", meta.get("tarih", "")
    ws["D2"], ws["E2"] = "Proforma No:", meta.get("proforma_no", "")
    ws["A3"], ws["B3"] = "İşin Adı:", meta.get("isin_adi", "")

    ws.merge_cells("A5:N5")
    t = ws["A5"]
    t.value = "PROFORMA FATURA"
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = title_fill
    t.alignment = Alignment(horizontal="center")

    hdr_row = 7
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = hdr_row + 1
    if meta.get("section"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        cell = ws.cell(row=row, column=1, value=meta["section"])
        cell.font = Font(bold=True)
        cell.fill = sec_fill
        row += 1
    if meta.get("group_label"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        cell = ws.cell(row=row, column=1, value=meta["group_label"])
        cell.font = Font(bold=True, italic=True)
        row += 1

    for item in items:
        vals = [
            item["bol"],
            item["grup"],
            item["poz"],
            item["ek"],
            item["stok"],
            item["tanim"],
            item["kaynak"],
            item["boy"],
            item["en"],
            item["yuk"],
            item["adet"],
            item["satis"],
            item["toplam"],
            item["doviz"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci in (11, 12, 13) and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
        row += 1

    grand = sum((it["toplam"] or 0) for it in items)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1, value="GENEL TOPLAM :").font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=13, value=round(grand, 2)).font = Font(bold=True)
    ws.cell(row=row, column=13).number_format = "#,##0.00"
    ws.cell(row=row, column=14, value="EUR").font = Font(bold=True)
    row += 2

    if terms:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        ws.cell(row=row, column=1, value="Şartlar ve Notlar").font = Font(bold=True)
        row += 1
        for term in terms:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
            ws.cell(row=row, column=1, value=f"* {term}")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1

    widths = [5, 5, 5, 4, 14, 55, 10, 8, 8, 8, 6, 12, 14, 6]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A8"
    wb.save(out_path)
    return out_path


def main() -> None:
    if len(sys.argv) > 1:
        pdf_path = Path(sys.argv[1])
    else:
        hits = glob.glob(r"c:\Users\adema\Downloads\COCA COLA*ISPARTA.pdf")
        if not hits:
            print("PDF bulunamadi")
            sys.exit(1)
        pdf_path = Path(hits[0])

    if not pdf_path.exists():
        print(f"PDF bulunamadi: {pdf_path}")
        sys.exit(1)

    text = extract_pdf_text(pdf_path)
    meta, items = parse_items(text)
    terms = parse_terms(text)
    out = pdf_path.with_suffix(".xlsx")
    write_excel(out, meta, items, terms)
    print(f"OK: {out}")
    print(
        f"Proforma: {meta.get('proforma_no')} | "
        f"Isin: {meta.get('isin_adi')} | Satir: {len(items)} | "
        f"Toplam: {sum((it['toplam'] or 0) for it in items):,.2f} EUR"
    )


if __name__ == "__main__":
    main()
