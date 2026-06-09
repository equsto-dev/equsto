# -*- coding: utf-8 -*-
"""
İnoksan 2026 Yurtiçi Bayi Fiyatları R1.xlsx → dept katalog (yalnız INO-* ana ürünler).
Aksesuar / Genel Aksesuar (201–209 önekleri) dahil değil.

Alış: liste × 0,77 (%23 iskonto) · Satış: alış × 1,10 (%10 kar)

  python scripts/sync-inoksan-fiyat-2026.py
  python scripts/sync-inoksan-fiyat-2026.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEPT_DIR = ROOT / "public/data/dept"
KUR_EUR_TRY = 53.2979
KDV_ORAN = 20
BAYI_ORAN = 0.77
KAR_ORAN = 1.10
BRAND = "İnoksan"
KAYNAK = "inoksan-fiyat-listesi-2026-r1"

DEPT_BY_H1 = {
    "Pişirici Cihazlar": "pisirme",
    "Fırınlar": "pisirme",
    "Bulaşık Yıkama Makineleri": "yikama",
    "Soğutucular": "sogutma",
    "Arabalar": "araba",
    "Servis Hatları": "tezgah",
    "Depolama ve İstifleme Üniteleri": "istif",
    "Tezgahlar": "tezgah",
}


def nf(s: str) -> str:
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()


def find_xlsx() -> Path:
    base = Path(r"c:/D Disk")
    for d in base.iterdir():
        if "fiyat" in nf(d.name) or ("yat" in nf(d.name) and "list" in nf(d.name)):
            for f in d.iterdir():
                if f.suffix.lower() != ".xlsx":
                    continue
                n = nf(f.name)
                if "inoksan" in n and "2026" in n and "bayi" in n:
                    return f
    raise FileNotFoundError("İNOKSAN 2026 Yurtiçi Bayi Fiyatları R1.xlsx bulunamadı")


def slugify(s: str) -> str:
    t = (
        str(s or "")
        .lower()
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ş", "s")
        .replace("ö", "o")
        .replace("ç", "c")
        .replace("ı", "i")
        .replace("İ", "i")
    )
    t = re.sub(r"[^a-z0-9]+", "-", t)
    return t.strip("-")[:80] or "diger"


def fmt_try(n: float) -> str:
    v = int(round(float(n)))
    return f"₺{v:,}".replace(",", ".") + ",00"


def is_main_sku(sku: str) -> bool:
    s = str(sku or "").strip().upper()
    return s.startswith("INO-")


def product_id(sku: str) -> str:
    return "inoksan__" + slugify(sku)


def pricing_fields(liste_eur: float, sku: str, short_name: str, cat_label: str) -> dict:
    liste = round(float(liste_eur), 2)
    alis = round(liste * BAYI_ORAN, 2)
    satis = round(alis * KAR_ORAN, 2)
    fiyat_tl_net = round(satis * KUR_EUR_TRY)
    fiyat_tl = round(fiyat_tl_net * (1 + KDV_ORAN / 100))
    price = f"{fmt_try(fiyat_tl)} KDV dahil"
    title = str(short_name or sku).strip()
    specs = "\n".join(
        [
            title,
            "",
            f"Ürün kodu: {sku}",
            f"Liste fiyatı (EUR): {liste}",
            f"Bayi iskonto: %23 (ödeme oranı {BAYI_ORAN})",
            f"Bayi net alış (EUR): {alis}",
            f"Equsto kar: %10",
            f"Equsto satış (EUR): {satis}",
            f"Hesap: liste × {BAYI_ORAN} × {KAR_ORAN}",
            f"Equsto satış (TL, KDV dahil): {fmt_try(fiyat_tl)}",
            f"Kur: 1 EUR = {KUR_EUR_TRY} TRY (KDV %{KDV_ORAN})",
            f"Kategori: {cat_label}",
            f"Kaynak: İnoksan 2026 Yurtiçi Bayi Fiyatları R1",
        ]
    )
    return {
        "price": price,
        "specs": specs,
        "aciklama": f"{title}\n\nKategori: {cat_label}",
        "liste_fiyati": liste,
        "liste_fiyati_eur": liste,
        "alis_fiyati": alis,
        "alis_fiyati_eur": alis,
        "satis_fiyati_eur": satis,
        "satis_eur_indirimli": satis,
        "iskontolu_fiyat": satis,
        "bayi_iskonto": 0.23,
        "equsto_kar_oran": 0.10,
        "para_birimi": "EUR",
        "fiyat_kaynagi": KAYNAK,
        "kaynak": KAYNAK,
        "kaynak_fiyat_listesi": KAYNAK,
        "kur_eur_try": KUR_EUR_TRY,
        "fiyat_tl_net": fiyat_tl_net,
        "fiyat_tl": fiyat_tl,
        "kdv_oran": KDV_ORAN,
        "fiyat_bekleniyor": False,
    }


def load_excel_rows() -> list[dict]:
    xlsx = find_xlsx()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        v = [c if c is not None else "" for c in row]
        sku = str(v[3]).strip()
        if not sku or not is_main_sku(sku):
            continue
        liste = v[8]
        if not isinstance(liste, (int, float)) or float(liste) <= 0:
            continue
        h1, h2, h3 = str(v[0]).strip(), str(v[1]).strip(), str(v[2]).strip()
        dept = DEPT_BY_H1.get(h1, "pisirme")
        cat = slugify(h3 or h2 or h1)
        short = str(v[4]).strip()
        g, d, y = v[5], v[6], v[7]
        olculer = {}
        if isinstance(g, (int, float)) and g > 0:
            olculer["uzunluk_mm"] = int(g)
        if isinstance(d, (int, float)) and d > 0:
            olculer["genislik_mm"] = int(d)
        if isinstance(y, (int, float)) and y > 0:
            olculer["yukseklik_mm"] = int(y)
        cat_label = " / ".join(x for x in [h1, h2, h3] if x)
        name = short.upper() if short else sku
        if not name.startswith("İNOKSAN") and not name.startswith("INOKSAN"):
            name = f"İNOKSAN {name}"
        px = pricing_fields(float(liste), sku, short or sku, cat_label)
        row_obj = {
            "category": cat,
            "brand": BRAND,
            "name": name,
            "sku": sku,
            "model": sku,
            "urun_kodu": sku,
            "stok_no": sku,
            "dept": dept,
            "oem_brand": BRAND,
            "vitrin_arka_plan": False,
            "id": product_id(sku),
            "images": [],
            "keywords": [BRAND, sku, cat, h1, h2, h3, short],
            "teknik_ozellikler": [],
            "olculer": olculer or None,
            "inoksan_h1": h1,
            "inoksan_h2": h2,
            "inoksan_h3": h3,
        }
        row_obj.update(px)
        out.append(row_obj)
    wb.close()
    return out


def is_inoksan_row(row: dict) -> bool:
    return (
        row.get("brand") == BRAND
        or row.get("kaynak_fiyat_listesi") == KAYNAK
        or str(row.get("id") or "").startswith("inoksan__")
    )


def yikama_vitrin_ok(row: dict) -> bool:
    """Bulaşık yıkama (yikama dept) sitede yok — hiçbir İnoksan yıkama satırı import edilmez."""
    if row.get("dept") == "yikama":
        return False
    return True


PRESERVE_KEYS = (
    "images",
    "inoksan_web_id",
    "inoksan_web_title",
    "inoksan_slug",
    "inoksan_url",
    "inoksan_match_via",
    "inoksan_image_source",
    "inoksan_image_url",
    "inoksan_enriched",
    "inoksan_enriched_at",
)


def load_old_inoksan_by_sku() -> dict[str, dict]:
    idx: dict[str, dict] = {}
    for dept_file in DEPT_DIR.glob("*.json"):
        data = json.loads(dept_file.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            continue
        for row in data:
            if is_inoksan_row(row) and row.get("sku"):
                idx[str(row["sku"])] = row
    return idx


def merge_into_depts(rows: list[dict], dry_run: bool) -> dict:
    by_dept: dict[str, list[dict]] = {}
    for r in rows:
        by_dept.setdefault(r["dept"], []).append(r)

    stats = {"added": 0, "removed": 0, "depts": 0}
    for dept_file in sorted(DEPT_DIR.glob("*.json")):
        data = json.loads(dept_file.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            continue
        before = len(data)
        data = [r for r in data if not is_inoksan_row(r)]
        removed = before - len(data)
        dept = dept_file.stem
        add = by_dept.pop(dept, [])
        if removed or add:
            stats["depts"] += 1
            stats["removed"] += removed
            stats["added"] += len(add)
            data.extend(add)
            print(f"  {dept_file.name}: -{removed} +{len(add)}")
            if not dry_run:
                dept_file.write_text(
                    json.dumps(data, ensure_ascii=False, separators=(",", ":")),
                    encoding="utf-8",
                )

    if by_dept:
        for dept, add in by_dept.items():
            path = DEPT_DIR / f"{dept}.json"
            data = []
            if path.exists():
                data = json.loads(path.read_text(encoding="utf-8"))
                if not isinstance(data, list):
                    data = []
            data = [r for r in data if not is_inoksan_row(r)]
            data.extend(add)
            stats["depts"] += 1
            stats["removed"] += 0
            stats["added"] += len(add)
            print(f"  {path.name}: +{len(add)} (yeni dept dosyası)")
            if not dry_run:
                path.write_text(
                    json.dumps(data, ensure_ascii=False, separators=(",", ":")),
                    encoding="utf-8",
                )
    return stats


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    all_rows = load_excel_rows()
    rows = [r for r in all_rows if yikama_vitrin_ok(r)]
    skipped_yikama = sum(
        1 for r in all_rows if r.get("dept") == "yikama" and not yikama_vitrin_ok(r)
    )
    print(f"[inoksan] Excel ana ürün (INO-*): {len(all_rows)}")
    if skipped_yikama:
        print(f"[inoksan] Yıkama vitrin dışı (silindi): {skipped_yikama}")
    if not rows:
        sys.exit(1)

    old_by_sku = load_old_inoksan_by_sku()
    for row in rows:
        prev = old_by_sku.get(str(row.get("sku") or ""))
        if not prev:
            continue
        for key in PRESERVE_KEYS:
            if prev.get(key):
                row[key] = prev[key]

    stats = merge_into_depts(rows, args.dry_run)
    print(f"\nÖzet: +{stats['added']} ürün | -{stats['removed']} eski | {stats['depts']} dept")
    if args.dry_run:
        print("(dry-run)")
        return

    subprocess.run(
        ["node", "scripts/rebuild-ekipmanlar-from-dept.mjs"],
        cwd=ROOT,
        check=True,
    )


if __name__ == "__main__":
    main()
