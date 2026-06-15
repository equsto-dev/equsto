#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""02- BALIKÇI ekipman listesi → steakhouse (2018-199-3) PROFORMA + Sayfa1 şablonu."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
OUT_DESKTOP = Path.home() / "Desktop" / "02-balikci.xlsx"
OUT_DATA = ROOT / "public" / "data" / "geo" / "02-balikci.xlsx"

# (bölüm slug PROFORMA, bölüm başlık, poz, ad, ölçü, adet)
ROWS: list[tuple[str, str, str, str, str, int]] = [
    ("kuru depo", "A- KURU DEPO", "A1", "İSTİF RAFI", "152*46*160", 2),
    ("soğuk oda", "B- SOĞUK ODA", "B1", "PANEL TİP SOĞUK ODA", "250*190*240", 1),
    ("soğuk oda", "B- SOĞUK ODA", "B2", "İSTİF RAFI", "152*46*160", 2),
    (
        "deepfreeze depo",
        "C- DEEPFREEZE DEPO",
        "C1",
        "PANEL TİP SOĞUK ODA",
        "250*190*240",
        1,
    ),
    ("deepfreeze depo", "C- DEEPFREEZE DEPO", "C2", "İSTİF RAFI", "137*46*160", 3),
    (
        "balık hazırlık",
        "D- BALIK HAZIRLIK",
        "D1",
        "DİZDEN KUMANDALI EL YIKAMA LAVABOSU",
        "-",
        1,
    ),
    (
        "balık hazırlık",
        "D- BALIK HAZIRLIK",
        "D2",
        "ÇALIŞMA TEZGAHI, TABAN RAFLI",
        "150*70*85",
        1,
    ),
    ("balık hazırlık", "D- BALIK HAZIRLIK", "D3", "DUVAR RAFI", "Ø40*50", 1),
    (
        "balık hazırlık",
        "D- BALIK HAZIRLIK",
        "D4",
        "BALIK AYIKLAMA TEZGAHI",
        "188*70*85",
        1,
    ),
    ("balık hazırlık", "D- BALIK HAZIRLIK", "D5", "DUVAR RAFI", "188*30*3", 1),
    ("balık hazırlık", "D- BALIK HAZIRLIK", "D6", "ÇÖP ARABASI", "Ø40*50", 1),
    (
        "balık hazırlık",
        "D- BALIK HAZIRLIK",
        "D7",
        "BIÇAK STERİL DOLABI",
        "10 BIÇAKLIK",
        1,
    ),
    (
        "pişirme",
        "E- PİŞİRME",
        "E1",
        "KONVEKSİYONLU FIRIN, ELK., RATIONAL SCC61E",
        "6 GN 1/1",
        1,
    ),
    ("pişirme", "E- PİŞİRME", "E2", "FIRIN TEZGAHI, TABAN RAFLI", "85*75*85", 1),
    (
        "pişirme",
        "E- PİŞİRME",
        "E3",
        "6 AÇIK ALEVLİ KUZİNE, GAZLI, MARENO",
        "120*90*85",
        1,
    ),
    (
        "pişirme",
        "E- PİŞİRME",
        "E4",
        "FRİTÖZ, ÇİFT SEPETLİ, ELK., SETÜSTÜ, MARENO FQE61L",
        "60*65*29",
        1,
    ),
    (
        "pişirme",
        "E- PİŞİRME",
        "E5",
        "LAVTAŞLI IZGARA, GAZLI, SETÜSTÜ, GPL 68G",
        "80*65*29",
        1,
    ),
    ("pişirme", "E- PİŞİRME", "E6", "NÖTR ARA TEZGAH, SETÜSTÜ", "40*65*29", 1),
    (
        "pişirme",
        "E- PİŞİRME",
        "E7",
        "TEZGAH ALTI BUZDOLABI, 3 KAPILI",
        "180*70*85",
        1,
    ),
    (
        "pişirme",
        "E- PİŞİRME",
        "E8",
        "ÇALIŞMA TEZGAHI, TABAN VE ARA RAFLI",
        "70*70*85",
        1,
    ),
    ("pişirme", "E- PİŞİRME", "E9", "TOST MAKİNASI", "-", 1),
    (
        "pişirme",
        "E- PİŞİRME",
        "E10",
        "SALAMANDER, ELK., BERTO'S SA/E 60",
        "-",
        1,
    ),
    (
        "pişirme",
        "E- PİŞİRME",
        "E11",
        "SERVİS TEZGAHI, TABAN VE ARA RAFLI, SICAK DOLAPLI",
        "170*70*85",
        1,
    ),
    ("pişirme", "E- PİŞİRME", "E12", "SERVİS TEZGAHI", "170*40*45", 1),
    (
        "pişirme",
        "E- PİŞİRME",
        "E13",
        "DAVLUMBAZ, DUVAR TİPİ, FİLTRELİ",
        "480*117*50",
        1,
    ),
    (
        "soğuk hazırlık",
        "F- SOĞUK HAZIRLIK",
        "F1",
        "ÇİFT EVYELİ ÇALIŞMA TEZGAHI",
        "190*70*85",
        1,
    ),
    ("soğuk hazırlık", "F- SOĞUK HAZIRLIK", "F2", "DUVAR RAFI", "140*30*3", 1),
    ("soğuk hazırlık", "F- SOĞUK HAZIRLIK", "F3", "ÇÖP ARABASI", "Ø40*50", 1),
    (
        "soğuk hazırlık",
        "F- SOĞUK HAZIRLIK",
        "F4",
        "BIÇAK STERİL DOLABI",
        "10 BIÇAKLIK",
        1,
    ),
    (
        "soğuk hazırlık",
        "F- SOĞUK HAZIRLIK",
        "F5",
        "MAKE-UP ÜNİTESİ, 3 KAPILI",
        "180*70*85",
        1,
    ),
    (
        "soğuk hazırlık",
        "F- SOĞUK HAZIRLIK",
        "F6",
        "ÇALIŞMA TEZGAHI, TABAN VE ARA RAFLI",
        "70*70*85",
        1,
    ),
    (
        "soğuk hazırlık",
        "F- SOĞUK HAZIRLIK",
        "F7",
        "KARBUZ MAKİNASI, BREMA G250",
        "250 kg/gün",
        1,
    ),
    ("soğuk hazırlık", "F- SOĞUK HAZIRLIK", "F8", "SİNEK ÖLDÜRÜCÜ", "-", 1),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G1",
        "TEK EVYELİ ÇALIŞMA TEZGAHI",
        "165*60*85",
        1,
    ),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G2",
        "BUZ MAKİNASI, BREMA CB416",
        "42 kg/gün",
        1,
    ),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G3",
        "BARDAK YIKAMA MAKİNASI, NEO DW53E",
        "60*60*85",
        1,
    ),
    ("ön mutfak - bar", "G- ÖN MUTFAK - BAR", "G4", "ÇAY OTOMATI", "23 lt.", 1),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G5",
        "CAM KAPILI TEZGAH TİPİ BUZDOLABI, 3 KAPILI",
        "188*60*85",
        2,
    ),
    ("ön mutfak - bar", "G- ÖN MUTFAK - BAR", "G6", "KASA BANKOSU", "110*100*85", 1),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G7",
        "BALIK TEŞHİR ÜNİTESİ",
        "200*75*125",
        1,
    ),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G8",
        "SALATA MEZE TEŞHİR ÜNİTESİ",
        "190*75*125",
        1,
    ),
    (
        "ön mutfak - bar",
        "G- ÖN MUTFAK - BAR",
        "G9",
        "TATLI TEŞHİR ÜNİTESİ",
        "150*75*125",
        1,
    ),
    (
        "bulaşıkhane",
        "H- BULAŞIK YIKAMA",
        "H1",
        "BULAŞIK SIYIRMA TEZGAHI",
        "200*70*85",
        1,
    ),
    ("bulaşıkhane", "H- BULAŞIK YIKAMA", "H2", "BASKET RAFI", "200*40*45", 1),
    ("bulaşıkhane", "H- BULAŞIK YIKAMA", "H3", "ÇÖP ARABASI", "Ø40*50", 1),
    (
        "bulaşıkhane",
        "H- BULAŞIK YIKAMA",
        "H4",
        "TEK EVYELİ ÇALIŞMA TEZGAHI",
        "230*70*85",
        1,
    ),
    (
        "bulaşıkhane",
        "H- BULAŞIK YIKAMA",
        "H5",
        "BULAŞIK YIKAMA MAKİNASI",
        "500 Tb/saat",
        1,
    ),
    ("bulaşıkhane", "H- BULAŞIK YIKAMA", "H6", "DUVAR RAFI", "230*30*3", 1),
    ("bulaşıkhane", "H- BULAŞIK YIKAMA", "H7", "YAĞ TUTUCU", "1,6 lt/sn", 1),
    ("bulaşıkhane", "H- BULAŞIK YIKAMA", "H8", "İSTİF RAFI", "152*46*160", 3),
    ("yer ızgarası", "Y- YER IZGARASI", "Y1", "YER IZGARASI", "30*30*14", 2),
    ("yer ızgarası", "Y- YER IZGARASI", "Y2", "YER IZGARASI", "95*30*14", 4),
]


def write_workbook(path: Path) -> None:
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "PROFORMA"

    ws_p["C1"] = "02- BALIKÇI"
    ws_p["C1"].font = Font(bold=True, size=14, color="FF0000")

    row = 3
    cur_slug: str | None = None
    for slug, _heading, _poz, ad, _olcu, _adet in ROWS:
        if slug != cur_slug:
            ws_p.cell(row=row, column=3, value=slug)
            row += 1
            cur_slug = slug
        ws_p.cell(row=row, column=5, value="-")
        ws_p.cell(row=row, column=7, value=ad)
        row += 1

    ws_s = wb.create_sheet("Sayfa1")
    headers = [
        "Poz",
        "Ürün",
        None,
        None,
        None,
        "Ölçü",
        "Adet",
        "Birim EUR",
        "Tutar",
        "İndirimli",
        "İnd. birim",
        "İnd. tutar",
    ]
    for col, h in enumerate(headers, 1):
        if h:
            ws_s.cell(row=1, column=col, value=h)

    for i, (slug, _heading, poz, ad, olcu, adet) in enumerate(ROWS, 2):
        ws_s.cell(row=i, column=1, value=poz)
        ws_s.cell(row=i, column=2, value=ad)
        ws_s.cell(row=i, column=6, value=olcu if olcu != "-" else None)
        ws_s.cell(row=i, column=7, value=adet)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    write_workbook(OUT_DESKTOP)
    write_workbook(OUT_DATA)
    print(f"OK {OUT_DESKTOP}")
    print(f"OK {OUT_DATA} ({len(ROWS)} kalem)")


if __name__ == "__main__":
    main()
