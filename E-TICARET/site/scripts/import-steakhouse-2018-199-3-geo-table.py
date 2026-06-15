#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2018-199-3.xlsx → public/data/geo/steakhouse-2018-199-3-table.json"""
from __future__ import annotations

import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_JSON = ROOT / "public" / "data" / "geo" / "steakhouse-2018-199-3-table.json"
OUT_XLSX = ROOT / "public" / "data" / "geo" / "2018-199-3.xlsx"
DEFAULT_XLSX = Path.home() / "Desktop" / "2018-199-3.xlsx"


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().upper())


def title_zone(z: str) -> str:
    z = z.strip()
    return z[0].upper() + z[1:] if z else z


def to_item(row: dict) -> dict:
    keys = (
        "ad",
        "olcu",
        "adet",
        "listeBirimEur",
        "listeTutarEur",
        "satisBirimEur",
        "satisTutarEur",
    )
    return {k: row[k] for k in keys if row.get(k) is not None}


def convert(xlsx: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws1 = wb["Sayfa1"]
    queues: dict[str, list] = defaultdict(list)
    all_rows: list[dict] = []

    for r in ws1.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]:
            continue
        ad = str(r[1]).strip()
        if not isinstance(r[6], (int, float)):
            continue
        row = {
            "ad": ad,
            "olcu": str(r[5]).strip() if r[5] else None,
            "adet": int(r[6]),
            "listeBirimEur": round(float(r[7]), 2) if r[7] else None,
            "listeTutarEur": round(float(r[8]), 2) if r[8] else None,
            "satisBirimEur": round(float(r[10]), 2) if r[10] else None,
            "satisTutarEur": round(float(r[11]), 2) if r[11] else None,
            "n": norm(ad),
        }
        all_rows.append(row)
        queues[row["n"]].append(row)

    def pick_item(name: str) -> dict | None:
        cands = queues.get(norm(name), [])
        return cands.pop(0) if cands else None

    ws = wb["PROFORMA"]
    zones_out: list[dict] = []
    matched: set[int] = set()

    for r in ws.iter_rows(values_only=True):
        z = r[2]
        if z and isinstance(z, str) and z.strip() and not r[6]:
            zt = z.strip()
            if zt.lower() not in ("p.no",):
                zones_out.append({"zone": title_zone(zt), "items": []})
        name = r[6]
        if not name or str(name).strip() in ("", "-", "URUN ADI", "ÜRÜN ADI"):
            continue
        row = pick_item(str(name))
        if row:
            matched.add(id(row))
            zones_out[-1]["items"].append(to_item(row))

    extras = [to_item(r) for r in all_rows if id(r) not in matched]
    if extras:
        zones_out.append({"zone": "Ek kalemler", "items": extras})

    return {
        "proformaNo": "2018-199-3",
        "label": "Steakhouse referans proforma (2018-199-3)",
        "kaynakDosya": "2018-199-3.xlsx",
        "yukleme": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z"),
        "zones": zones_out,
        "ozet": {
            "kalemSayisi": sum(len(z["items"]) for z in zones_out),
            "listeToplamEur": round(
                sum(
                    x.get("listeTutarEur", 0) or 0
                    for z in zones_out
                    for x in z["items"]
                ),
                2,
            ),
            "satisToplamEur": round(
                sum(
                    x.get("satisTutarEur", 0) or 0
                    for z in zones_out
                    for x in z["items"]
                ),
                2,
            ),
        },
    }


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"Excel bulunamadı: {xlsx}")

    data = convert(xlsx)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(xlsx, OUT_XLSX)
    print(
        f"OK {OUT_JSON.name}: {data['ozet']['kalemSayisi']} kalem, "
        f"liste {data['ozet']['listeToplamEur']} €, "
        f"proforma {data['ozet']['satisToplamEur']} €"
    )


if __name__ == "__main__":
    main()
