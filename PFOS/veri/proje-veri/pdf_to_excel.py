# -*- coding: utf-8 -*-
"""Convert PFOS project PDF equipment lists to Excel files."""

import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(r"c:\D Disk\EQUSTO-WORK\PFOS\veri\proje-veri")

PDF_FILES = [
    "06 SUSHI.pdf",
    "7 ŞARKÜTERİ.pdf",
    "8 HAMBURGER.pdf",
    "11 BIRAHANE.pdf",
    "13 HOTDOG.pdf",
    "14-PASTANE.pdf",
    "17 TAVUKCU.pdf",
    "19 THEHOUSE CAFE.pdf",
    "20 DONDURMACI - KREP.pdf",
    "PIDECI.pdf",
    "RESTORAN.pdf",
    "03-italyan.pdf",
]

HEADERS = ["Bölüm", "P.NO", "ÜRÜN ADI", "ÖLÇÜ", "AD."]
SECTION_PATTERN = re.compile(r"^[A-ZİĞÜŞÖÇ]-\s+.+")
PROJECT_PATTERN = re.compile(r"^\d{2,3}[-\s]|^RESTAURANT$|^RESTORAN$", re.I)
PNO_PATTERN = re.compile(r"^[A-Z]\d+[A-Z]?$|^\d+$", re.I)


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def parse_pdf_tables(pdf_path: Path) -> tuple[str, list[dict]]:
    project_title = ""
    current_section = ""
    rows: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table:
                    if not raw_row:
                        continue

                    cells = [clean(c) for c in raw_row]
                    while len(cells) < 4:
                        cells.append("")

                    first = cells[0]
                    if not first:
                        continue

                    if first.upper() in {"P.NO", "P.NO ÜRÜN ADI ÖLÇÜ AD."}:
                        continue

                    if not project_title and (
                        PROJECT_PATTERN.match(first) or first.upper() in {"RESTAURANT", "RESTORAN"}
                    ):
                        project_title = first
                        continue

                    if SECTION_PATTERN.match(first) and not PNO_PATTERN.match(first):
                        current_section = first
                        continue

                    if PNO_PATTERN.match(first):
                        rows.append(
                            {
                                "section": current_section,
                                "P.NO": first,
                                "ÜRÜN ADI": cells[1],
                                "ÖLÇÜ": cells[2],
                                "AD.": cells[3],
                            }
                        )

    return project_title, rows


def group_by_section(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    groups: list[tuple[str, list[dict]]] = []
    current_section = None
    current_items: list[dict] = []

    for row in rows:
        section = row.get("section", "")
        if section != current_section:
            if current_items:
                groups.append((current_section or "", current_items))
            current_section = section
            current_items = [row]
        else:
            current_items.append(row)

    if current_items:
        groups.append((current_section or "", current_items))

    return groups


def autosize_columns(ws, start_row: int = 2):
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def write_excel(pdf_path: Path, project_title: str, rows: list[dict]) -> Path:
    xlsx_path = pdf_path.with_suffix(".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürün Listesi"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = project_title or pdf_path.stem
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 3
    groups = group_by_section(rows)

    for section_name, items in groups:
        if section_name:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            section_cell = ws.cell(row=row_idx, column=1, value=section_name)
            section_cell.font = Font(bold=True, size=11)
            section_cell.fill = section_fill
            section_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        for item in items:
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value=item.get("P.NO", ""))
            ws.cell(row=row_idx, column=3, value=item.get("ÜRÜN ADI", ""))
            ws.cell(row=row_idx, column=4, value=item.get("ÖLÇÜ", ""))
            ad_value = item.get("AD.", "")
            ws.cell(row=row_idx, column=5, value=ad_value)

            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)

            row_idx += 1

    ws.freeze_panes = "A3"
    autosize_columns(ws, start_row=2)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        alt_path = xlsx_path.with_name(f"{xlsx_path.stem}_yeni{xlsx_path.suffix}")
        wb.save(alt_path)
        return alt_path
    return xlsx_path


def main():
    created = []
    for pdf_name in PDF_FILES:
        pdf_path = BASE_DIR / pdf_name
        if not pdf_path.exists():
            print(f"ATLANDI (bulunamadi): {pdf_name}")
            continue
        project_title, rows = parse_pdf_tables(pdf_path)
        out = write_excel(pdf_path, project_title, rows)
        created.append((pdf_name, len(rows), out.name))
        print(f"OK: {out.name} ({len(rows)} urun, {len(group_by_section(rows))} bolum)")

    print(f"\nToplam {len(created)} Excel dosyasi olusturuldu.")


if __name__ == "__main__":
    main()
