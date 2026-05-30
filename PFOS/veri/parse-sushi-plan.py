# -*- coding: utf-8 -*-
"""SUSHI.-PLANpdf.pdf + 06-SUSHI-ekipman-listesi.xlsx → JSON + birleşik Excel."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERI = Path(__file__).resolve().parent
PROJE_VERI = VERI / "proje-veri"
PLAN_PDF = PROJE_VERI / "SUSHI.-PLANpdf.pdf"
EKIPMAN_XLSX = PROJE_VERI / "06-SUSHI-ekipman-listesi.xlsx"

OUT_JSON = VERI / "SUSHI-PLAN.json"
OUT_ALAN_XLSX = VERI / "SUSHI-PLAN-alan-listesi.xlsx"
OUT_BIRLESIK_XLSX = VERI / "SUSHI-PLAN-birlesik.xlsx"

TOTAL_M2_PLAN = 20.1
KONSEPT = "06-sushi"

POZ_RE = re.compile(r"^C\d{1,2}$|^A\d$|^B\d$")

# Ekipman bölümleri (PDF listesi ile uyumlu)
MAHALS = [
    {"id": "kuru-depo", "label": "Kuru depo", "bolum": "A", "pozPrefix": "A"},
    {"id": "soguk-oda", "label": "Soğuk oda", "bolum": "B", "pozPrefix": "B"},
    {"id": "pisirme-teshir", "label": "Pişirme - Teşhir", "bolum": "C", "pozPrefix": "C"},
]


def read_pdf_text(path: Path) -> str:
    import fitz

    doc = fitz.open(path)
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    return text


def extract_plan_meta(text: str) -> dict:
    m2 = None
    for ln in text.splitlines():
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*m\s*2", ln, re.I)
        if m:
            m2 = float(m.group(1).replace(",", "."))
    poz_set: set[str] = set()
    for tok in re.findall(r"\b[A-Z]\d{1,2}\b", text):
        if POZ_RE.match(tok) or re.match(r"^C\d{1,2}$", tok):
            poz_set.add(tok)
    return {"totalM2": m2, "pozlar": sorted(poz_set, key=lambda p: (p[0], int(p[1:])))}


def read_ekipman_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    bolum = ""
    bolum_ad = ""
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not any(row):
            continue
        a, b, c, d, e = (row + (None,) * 5)[:5]
        if a and not b and isinstance(a, str) and "-" in a:
            bolum_ad = str(a).strip()
            bolum = bolum_ad.split("-", 1)[0].strip()
            continue
        if b and c:
            poz = str(b).strip()
            if not re.match(r"^[A-Z]\d", poz):
                continue
            adet = e
            if isinstance(adet, float):
                adet = int(adet)
            rows.append({
                "bolum": bolum,
                "bolumAd": bolum_ad,
                "poz": poz,
                "ad": str(c).strip(),
                "olcu": str(d).strip() if d else "—",
                "adet": adet if adet is not None else "—",
            })
    wb.close()
    return rows


def assign_mahal(poz: str) -> dict:
    for m in MAHALS:
        if poz.startswith(m["pozPrefix"]):
            return m
    return {"id": "diger", "label": "Diğer", "bolum": poz[0] if poz else "?"}


def build(plan_meta: dict, ekipman: list[dict]) -> dict:
    plan_poz = set(plan_meta["pozlar"])
    by_mahal: dict[str, dict] = {}
    for m in MAHALS:
        by_mahal[m["id"]] = {**m, "pozlar": [], "pozDetay": []}

    for eq in ekipman:
        m = assign_mahal(eq["poz"])
        mid = m["id"]
        if mid not in by_mahal:
            by_mahal[mid] = {**m, "pozlar": [], "pozDetay": []}
        by_mahal[mid]["pozlar"].append(eq["poz"])
        by_mahal[mid]["pozDetay"].append({
            "poz": eq["poz"],
            "planda": eq["poz"] in plan_poz,
            "ekipman": {k: eq[k] for k in ("ad", "olcu", "adet", "bolumAd")},
        })

    # Alan: plan tek blok 20.1 m²; A/B planda yoksa not
    planda_c = [p for p in plan_poz if p.startswith("C")]
    planda_ab = [p for p in plan_poz if p.startswith(("A", "B"))]

    mahaller = []
    for m in MAHALS:
        block = by_mahal.get(m["id"], {**m, "pozlar": [], "pozDetay": []})
        on_plan = any(p["planda"] for p in block.get("pozDetay", []))
        if m["bolum"] == "C":
            plan_m2 = TOTAL_M2_PLAN
        else:
            plan_m2 = None
        mahaller.append({
            **m,
            "planM2": plan_m2,
            "plandaGorunur": on_plan,
            "pozSayisi": len(block.get("pozlar", [])),
            "pozlar": block.get("pozlar", []),
        })

    only_plan = sorted(plan_poz - {e["poz"] for e in ekipman})
    only_eq = sorted({e["poz"] for e in ekipman} - plan_poz)

    return {
        "version": "1.0",
        "konsept": KONSEPT,
        "label": "SUSHI",
        "kaynak": {
            "plan": PLAN_PDF.name,
            "ekipman": EKIPMAN_XLSX.name,
        },
        "meta": {
            "totalM2": TOTAL_M2_PLAN,
            "totalM2Kaynak": "plan-pdf",
            "planPozSayisi": len(plan_poz),
            "ekipmanKalem": len(ekipman),
            "eslesen": len(ekipman) - len(only_eq),
        },
        "mahaller": mahaller,
        "mahallerDetay": list(by_mahal.values()),
        "planPozlari": plan_meta["pozlar"],
        "ekipman": ekipman,
        "uyumsuzluk": {
            "plandaEkipmandaYok": only_plan,
            "ekipmandaPlandaYok": only_eq,
        },
        "notlar": [
            "Plan PDF yalnızca C (pişirme-teşhir) pozlarını gösteriyor; A/B ekipman listede var.",
            f"Planda görülen C poz: {len(planda_c)} · Listede A+B+C toplam: {len(ekipman)} kalem.",
        ],
    }


def _styles():
    thin = Side(style="thin", color="CCCCCC")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "hdr_fill": PatternFill("solid", fgColor="001E50"),
        "sec_fill": PatternFill("solid", fgColor="E8EEF5"),
        "title_font": Font(name="Calibri", size=14, bold=True, color="001E50"),
        "hdr_font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        "sec_font": Font(name="Calibri", size=11, bold=True, color="001E50"),
        "body_font": Font(name="Calibri", size=10),
    }


def write_alan_xlsx(data: dict):
    st = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Alan"
    ws["A1"] = "06 — SUSHI · Plan alan özeti"
    ws["A1"].font = st["title_font"]
    ws.merge_cells("A1:F1")
    ws["A2"], ws["B2"] = "Toplam m² (plan)", data["meta"]["totalM2"]
    ws["A3"], ws["B3"] = "Kaynak plan", data["kaynak"]["plan"]

    row = 5
    for col, h in enumerate(("Mahal", "Böl.", "Plan m²", "Planda", "Poz sayısı", "Pozlar"), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = st["hdr_font"]
        c.fill = st["hdr_fill"]
        c.border = st["border"]
    row += 1
    for m in data["mahaller"]:
        ws.cell(row=row, column=1, value=m["label"]).border = st["border"]
        ws.cell(row=row, column=2, value=m["bolum"]).border = st["border"]
        ws.cell(row=row, column=3, value=m.get("planM2") or "—").border = st["border"]
        ws.cell(row=row, column=4, value="Evet" if m.get("plandaGorunur") else "Hayır").border = st["border"]
        ws.cell(row=row, column=5, value=m["pozSayisi"]).border = st["border"]
        ws.cell(row=row, column=6, value=", ".join(m["pozlar"])).border = st["border"]
        row += 1
    for i, w in enumerate((22, 6, 10, 10, 10, 40), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(OUT_ALAN_XLSX)


def write_birlesik_xlsx(data: dict):
    st = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan + Ekipman"
    ws["A1"] = "06 — SUSHI · Plan poz ↔ ekipman"
    ws["A1"].font = st["title_font"]
    ws.merge_cells("A1:G1")

    row = 3
    for col, h in enumerate(
        ("Mahal", "Böl.", "Poz", "Planda", "Ürün adı", "Ölçü", "Ad."), 1
    ):
        c = ws.cell(row=row, column=col, value=h)
        c.font = st["hdr_font"]
        c.fill = st["hdr_fill"]
        c.border = st["border"]
    row += 1

    for eq in data["ekipman"]:
        m = assign_mahal(eq["poz"])
        ws.cell(row=row, column=1, value=m["label"]).border = st["border"]
        ws.cell(row=row, column=2, value=eq["bolum"]).border = st["border"]
        ws.cell(row=row, column=3, value=eq["poz"]).border = st["border"]
        planda = "✓" if eq["poz"] in data["planPozlari"] else "—"
        ws.cell(row=row, column=4, value=planda).border = st["border"]
        ws.cell(row=row, column=5, value=eq["ad"]).border = st["border"]
        ws.cell(row=row, column=6, value=eq["olcu"]).border = st["border"]
        ws.cell(row=row, column=7, value=eq["adet"]).border = st["border"]
        row += 1

    for i, w in enumerate((20, 6, 8, 8, 48, 18, 6), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    wb.save(OUT_BIRLESIK_XLSX)


def main():
    plan_text = read_pdf_text(PLAN_PDF)
    plan_meta = extract_plan_meta(plan_text)
    ekipman = read_ekipman_xlsx(EKIPMAN_XLSX)
    data = build(plan_meta, ekipman)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_alan_xlsx(data)
    write_birlesik_xlsx(data)
    print(OUT_JSON)
    print(OUT_ALAN_XLSX)
    print(OUT_BIRLESIK_XLSX)
    print(
        f"plan {len(data['planPozlari'])} poz | ekipman {len(ekipman)} | "
        f"planda yok (liste): {data['uyumsuzluk']['ekipmandaPlandaYok']} | "
        f"planda fazla: {data['uyumsuzluk']['plandaEkipmandaYok']}"
    )


if __name__ == "__main__":
    main()
