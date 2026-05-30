# -*- coding: utf-8 -*-
"""PFOS veri/*.pdf → ayrı ekipman Excel dosyaları (Equsto stili)."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERI = Path(__file__).resolve().parent
PROJE_VERI = VERI / "proje-veri"

PDF_JOBS = [
    "03-italyan.pdf",
    "06 SUSHI.pdf",
    "7 ŞARKÜTERİ.pdf",
    "8 HAMBURGER.pdf",
    "11 BIRAHANE.pdf",
    "13 HOTDOG.pdf",
    "14-PASTANE.pdf",
    "17 TAVUKCU.pdf",
    "19 THEHOUSE CAFE.pdf",
    "20 DONDURMACI - KREP.pdf",
    "PIDECI.pdf",
    "RESTORAN.pdf",
]

PLAN_PDF = "ITALYAN-PLAN.pdf"

POZ_LINE = re.compile(r"^[A-Z]\d{1,2}A?$|^Y\d$|^\d{1,3}$")
SECTION_LINE = re.compile(r"^[A-Z]- .+|^[A-Z] - .+")
TITLE_LINE = re.compile(r"^(\d{1,2})[-\s].+|^RESTAURANT$", re.I)


def read_pdf(path: Path) -> str:
    import fitz

    doc = fitz.open(path)
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    return text


def slug_title(raw: str) -> str:
    s = re.sub(r"[^\w\s-]", "", raw, flags=re.UNICODE)
    s = re.sub(r"\s+", "-", s.strip())
    return s[:60] or "proje"


def detect_title(lines: list[str], fallback: str) -> str:
    for ln in lines[:12]:
        ln = ln.strip()
        if TITLE_LINE.match(ln) or re.match(r"^\d{2}-", ln):
            return ln.replace("\t", " ").strip()
        if ln.upper() in ("RESTAURANT", "RESTORAN"):
            return "RESTORAN"
    return fallback


def parse_tab_rows(text: str) -> list[dict] | None:
    """Tek satırda poz + ad + ölçü + adet (Read aracı çıktısı)."""
    rows: list[dict] = []
    bolum = ""
    bolum_ad = ""
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("--") or line.startswith("P.NO"):
            continue
        if SECTION_LINE.match(line):
            bolum = line.split("-", 1)[0].strip()
            bolum_ad = line
            continue
        if re.match(r"^\d{2}-", line) or line.upper() == "RESTAURANT":
            continue
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t") if p.strip()]
            if len(parts) >= 4 and POZ_LINE.match(parts[0]):
                rows.append({
                    "bolum": bolum,
                    "bolumAd": bolum_ad,
                    "poz": parts[0],
                    "ad": parts[1],
                    "olcu": parts[2] or "—",
                    "adet": int(parts[3]) if parts[3].isdigit() else parts[3],
                })
            elif len(parts) >= 3 and POZ_LINE.match(parts[0]):
                adet = parts[-1]
                rows.append({
                    "bolum": bolum,
                    "bolumAd": bolum_ad,
                    "poz": parts[0],
                    "ad": " ".join(parts[1:-2]) if len(parts) > 3 else parts[1],
                    "olcu": parts[-2] if len(parts) > 3 else "—",
                    "adet": int(adet) if str(adet).isdigit() else adet,
                })
    return rows if rows else None


def parse_line_rows(text: str) -> list[dict]:
    """PyMuPDF: poz / ad / ölçü / adet ayrı satırlarda."""
    rows: list[dict] = []
    bolum = ""
    bolum_ad = ""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    skip_hdr = {"P.NO", "ÜRÜN ADI", "ÖLÇÜ", "AD."}
    i = 0
    while i < len(lines):
        line = lines[i]
        if line in skip_hdr or line.startswith("--"):
            i += 1
            continue
        if re.match(r"^\d{2}-", line) or line.upper() in ("RESTAURANT", "RESTORAN"):
            i += 1
            continue
        if SECTION_LINE.match(line):
            bolum = line.split("-", 1)[0].strip()
            bolum_ad = line
            i += 1
            continue
        if POZ_LINE.match(line):
            poz = line
            i += 1
            ad_parts: list[str] = []
            while i < len(lines):
                nxt = lines[i]
                if POZ_LINE.match(nxt) or SECTION_LINE.match(nxt):
                    break
                if nxt in skip_hdr or re.match(r"^\d{2}-", nxt):
                    i += 1
                    continue
                ad_parts.append(nxt)
                i += 1
            if len(ad_parts) >= 2:
                olcu = ad_parts[-2]
                adet_raw = ad_parts[-1]
                ad = " ".join(ad_parts[:-2]) if len(ad_parts) > 2 else ad_parts[0]
                adet: int | str = int(adet_raw) if adet_raw.isdigit() else adet_raw
                rows.append({
                    "bolum": bolum,
                    "bolumAd": bolum_ad,
                    "poz": poz,
                    "ad": ad,
                    "olcu": olcu or "—",
                    "adet": adet,
                })
            continue
        i += 1
    return rows


def parse_equipment_pdf(path: Path) -> tuple[str, list[dict]]:
    text = read_pdf(path)
    tab = parse_tab_rows(text)
    rows = tab if tab else parse_line_rows(text)
    title = detect_title(text.splitlines(), path.stem)
    return title, rows


def write_ekipman_xlsx(out: Path, title: str, rows: list[dict], kaynak: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ekipman"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="001E50")
    sec_fill = PatternFill("solid", fgColor="E8EEF5")
    title_font = Font(name="Calibri", size=14, bold=True, color="001E50")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sec_font = Font(name="Calibri", size=11, bold=True, color="001E50")
    body_font = Font(name="Calibri", size=10)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{title} · Ekipman listesi"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 28
    ws["A2"], ws["B2"] = "Kaynak", kaynak
    ws["A2"].font = Font(bold=True, size=10, color="64748B")

    start = 4
    for col, h in enumerate(("Böl.", "Poz", "Ürün adı", "Ölçü", "Ad."), 1):
        c = ws.cell(row=start, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = start + 1
    total = 0
    last_sec = None
    for row in rows:
        sec = row["bolum"]
        if sec and sec != last_sec:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            hdr = next((x.get("bolumAd") or sec for x in rows if x["bolum"] == sec), sec)
            c = ws.cell(row=r, column=1, value=hdr)
            c.font = sec_font
            c.fill = sec_fill
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = border
            r += 1
            last_sec = sec

        ws.cell(row=r, column=1, value=row["bolum"]).font = body_font
        ws.cell(row=r, column=2, value=row["poz"]).font = body_font
        ws.cell(row=r, column=3, value=row["ad"]).font = body_font
        ws.cell(row=r, column=4, value=row["olcu"]).font = body_font
        ad_cell = ws.cell(row=r, column=5, value=row["adet"])
        ad_cell.font = body_font
        ad_cell.alignment = Alignment(horizontal="center")
        if isinstance(row["adet"], int):
            total += row["adet"]
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = border
            if col == 3:
                ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="center")
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="TOPLAM ADET").font = Font(bold=True, size=10)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=5, value=total).font = Font(bold=True, size=10)
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F8FAFC")

    for i, w in enumerate((8, 8, 52, 18, 8), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start + 1}"
    if r > start + 1:
        ws.auto_filter.ref = f"A{start}:E{r - 1}"
    wb.save(out)


def parse_plan_pdf(path: Path) -> list[dict]:
    """ITALYAN-PLAN: poz listesi + plan metin alanları."""
    text = read_pdf(path)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    pozlar: list[str] = []
    seen: set[str] = set()
    alanlar: list[tuple[str, str]] = []
    for ln in lines:
        if POZ_LINE.match(ln) and ln not in seen:
            seen.add(ln)
            pozlar.append(ln)
        m = re.match(r"^(\d+)\s*m2", ln, re.I)
        if m:
            alanlar.append((ln, m.group(1)))

    rows: list[dict] = []
    for p in sorted(pozlar, key=lambda x: (x[0], int(re.search(r"\d+", x).group() or 0))):
        rows.append({
            "bolum": "PLAN",
            "bolumAd": "PLAN — Yerleşim poz",
            "poz": p,
            "ad": "Plan poz etiketi",
            "olcu": "—",
            "adet": 1,
        })
    return rows


def write_plan_xlsx(out: Path, rows: list[dict], kaynak: str):
    write_ekipman_xlsx(out, "İTALYAN · Yerleşim planı (poz)", rows, kaynak)


def out_name(pdf: Path, title: str) -> Path:
    if pdf.name == PLAN_PDF:
        return VERI / "ITALYAN-PLAN-poz-listesi.xlsx"
    stem = pdf.stem.strip()
    m = re.match(r"^(\d{1,2})", stem)
    if m:
        num = m.group(1).zfill(2) if len(m.group(1)) == 1 else m.group(1)
        rest = re.sub(r"^\d{1,2}[\s\-]*", "", stem).strip()
        base = f"{num}-{slug_title(rest)}" if rest else num
    else:
        base = slug_title(stem)
    return PROJE_VERI / f"{base}-ekipman-listesi.xlsx"


def main():
    made: list[str] = []
    for name in PDF_JOBS:
        path = PROJE_VERI / name
        if not path.exists():
            print("SKIP (yok):", name)
            continue
        title, rows = parse_equipment_pdf(path)
        out = out_name(path, title)
        write_ekipman_xlsx(out, title, rows, name)
        made.append(f"{name} -> {out.name} ({len(rows)} kalem)")
        print(made[-1])

    plan = PROJE_VERI / PLAN_PDF
    if plan.exists():
        rows = parse_plan_pdf(plan)
        out = VERI / "ITALYAN-PLAN-poz-listesi.xlsx"
        write_plan_xlsx(out, rows, PLAN_PDF)
        made.append(f"{PLAN_PDF} -> {out.name} ({len(rows)} poz)")
        print(made[-1])

    print("\nToplam:", len(made), "dosya")


if __name__ == "__main__":
    main()
